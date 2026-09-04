# -*- coding: utf-8 -*-
"""Relatorio de folha (holerite) AGOSTO/2026 - lojas CENTRO e TRANCOSO.
Pagamento em 05/09/2026. Mesma estrutura usada na loja ARRAIAL."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = "/home/user/site-farmacia/relatorios/contabilidade/"
F = "Arial"
MONEY = 'R$ #,##0.00;-R$ #,##0.00;"-"'

def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

FILL_TIT   = PatternFill("solid", fgColor="1F3864")
FILL_SEC   = PatternFill("solid", fgColor="D9E2F3")
FILL_HDR   = PatternFill("solid", fgColor="2E5496")
FILL_IN    = PatternFill("solid", fgColor="FFF2CC")
FILL_CONF  = PatternFill("solid", fgColor="FCE4D6")
FILL_TOT   = PatternFill("solid", fgColor="E2EFDA")
FILL_LIQ   = PatternFill("solid", fgColor="C6E0B4")
FILL_ALERT = PatternFill("solid", fgColor="FFF0F0")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BLUE, GREEN = "0000FF", "008000"

# ------------------------------------------------------------------ CONFIG
CENTRO = dict(
 loja="CENTRO", arquivo="RELATORIO_HOLERITE_AGOSTO_2026_CENTRO_pgto_05-09-2026.xlsx",
 mult=1.0,
 EMP=["ARIANE", "ELIANA", "GENECIR", "RENALDO", "THAYANE", "PEDRO"],
 # nome: (cod, venda bruta, descontos, venda liquida, comissao, inj, vit, outros)
 INOVA={"ARIANE":  (211, 16556.20,  5188.17, 11368.03,  257.46, 120.0,  0.0,   5.0),
        "ELIANA":  ( 81, 60046.70, 20690.58, 39356.12, 1511.36, 170.0, 20.0,  35.0),
        "GENECIR": (241,   188.86,     5.99,   182.87,    0.62,   0.0,  0.0,   0.0),  # ferias 01 a 30/08
        "RENALDO": (151, 49529.57, 21504.43, 28025.14,  847.94,  80.0, 10.0,   6.0),
        "THAYANE": ( 61,  6772.99,   199.94,  6573.05,   57.33,   0.0,  0.0,  10.0),
        "PEDRO":   ( 41,  1689.51,   310.63,  1378.88,   12.18,   0.0,  0.0,   0.0)},
 OUTROS_COD=[("11", "GILSON MOURA SANTOS", 7066.28, 26.43, 31.0),
             ("13", "UILLIAN SILVA SANTOS", 2134.49, 0.0, 0.0),
             ("31", "FABIANA RODRIGUES", 241.91, 5.83, 0.0),
             ("43", "SARA", 365.03, 2.65, 0.0)],
 OUTROS_NOTA="GILSON, UILLIAN e FABIANA não recebem comissão. O saldo da SARA (loja Arraial) foi somado à folha dela no relatório do Arraial.",
 FOLGUISTAS=[("SERGIO", "51", 150.00, 6023.97, 128.80, 40.0),
             ("ANA CELIA", "131", 100.00, 3924.63, 83.56, 0.0)],
 GERAL=dict(bruta=154540.14, desc=52653.74, liq=101886.40, com=2934.17, inc=527.0),
 SALARIO={"ARIANE": 5648.41, "ELIANA": 1621.0, "GENECIR": 1621.0,
          "RENALDO": 1621.0, "THAYANE": 1621.0, "PEDRO": 1621.0},
 NOTURNO={}, AUXGER={}, METACX={},
 VT6={"ELIANA": -84.29, "RENALDO": -74.29, "THAYANE": -84.29},
 SEM_COMISSAO=["ARIANE", "PEDRO"],
 SALARIO_ZERO={"GENECIR": "férias de 01 a 30/08/2026, pagas em recibo próprio"},
 JUL={"COMISSÃO PRODUTOS + PDV": {"ELIANA": 1400.0, "GENECIR": 100.0, "RENALDO": 1200.0, "THAYANE": 60.0},
      "PRÊMIO COTA GERAL":       {"ELIANA": 300.0, "GENECIR": 100.0, "RENALDO": 150.0, "THAYANE": 100.0},
      "PRÊMIO PRÉ-VENCIDOS":     {"ELIANA": 250.0, "RENALDO": 100.0},
      "REPOUSO REMUNERADO / DSR": {"ELIANA": 269.23, "GENECIR": 19.23, "RENALDO": 230.77, "THAYANE": 11.54},
      "SALÁRIO BASE":            {"ARIANE": 5648.41, "ELIANA": 1621.0, "GENECIR": 1621.0,
                                  "RENALDO": 1621.0, "THAYANE": 1621.0, "PEDRO": 1621.0},
      "CONVÊNIO":                {"GENECIR": -400.0, "RENALDO": -400.0, "THAYANE": -200.0},
      "DESCONTO VALE TRANSPORTE 6%": {"ELIANA": -84.29, "GENECIR": -84.29, "RENALDO": -74.29, "THAYANE": -84.29},
      "DESCONTO INSS":           {"ARIANE": -592.27, "ELIANA": -349.41, "GENECIR": -141.30,
                                  "RENALDO": -284.80, "THAYANE": -137.00},
      "DESCONTO IRRF":           {"ARIANE": -251.04},
      "ADIANTAMENTO / VALES":    {"ELIANA": -1300.0, "GENECIR": -700.0, "RENALDO": -1000.0}},
 JUL_SALDO={"ARIANE": 4805.10, "ELIANA": 2106.53, "GENECIR": 514.64,
            "RENALDO": 1542.68, "THAYANE": 1371.25, "PEDRO": 1621.00},
 JUL_VT={"ELIANA": 270.0, "RENALDO": 260.0, "THAYANE": 260.0, "PEDRO": 350.0},
 JUL_VA={},
 JUL_ABA="CENTRO 05JULHO",
 FERIAS_OBS="GENECIR: férias de 01 a 30/08/2026, pagas em recibo próprio — informar o valor.",
 BONIF={"ELIANA": 600.0, "THAYANE": 100.0, "RENALDO": 250.0},
 DSR_HE={}, SAL_FAMILIA={}, INSUF={"GENECIR": 3.24},
 VALES={"ELIANA": -2192.0, "THAYANE": -200.0, "RENALDO": -1955.0},
 VT_DESC={"ELIANA": -84.29, "THAYANE": -84.29, "RENALDO": -68.08, "GENECIR": -3.24},
 INSS={"ELIANA": -371.35, "ARIANE": -592.27, "THAYANE": -136.72, "RENALDO": -235.06},
 IRRF={"ARIANE": -251.04},
 LIQ_BETEL={"ELIANA": 1375.37, "ARIANE": 4805.10, "THAYANE": 1368.35,
            "RENALDO": 623.87, "GENECIR": 0.0},
 PONTO=[],
 DEFINIDO=["ARIANE e PEDRO: não recebem comissão sobre vendas — a rubrica fica zerada e o apurado deles aparece só na aba BASE.",
           "GENECIR: férias de 01 a 30/08/2026, já pagas em recibo próprio — salário zerado neste holerite; confirmar o dia 31/08.",
           "SERGIO e ANA CELIA são folguistas (diária de R$ 150,00 e R$ 100,00) — ficam na aba FOLGUISTAS, para o contas a pagar, fora do holerite.",
           "GILSON, UILLIAN e FABIANA não recebem comissão pela loja Centro.",
           "SARA: o saldo dela na loja Centro foi somado à folha do Arraial."],
 PENDENCIAS=["PEDRO não aparece no holerite de agosto emitido pela Betel para esta loja — verificar em qual empresa ele foi lançado.",
             "GENECIR: a comissão de R$ 0,62 não entrou no holerite (ele passou o mês de férias) — conferir se vale lançar.",
             "GENECIR: as férias vão até 30/08 e o mês tem 31 dias — confirmar se o dia 31/08 foi trabalhado e lançar o salário desse dia (R$ 1.621,00 ÷ 30 = R$ 54,03).",
             "Preencher a quantidade de diárias de SERGIO e ANA CELIA na aba FOLGUISTAS.",
             "Confirmar se os folguistas recebem a comissão apurada no código deles (SERGIO R$ 128,80 e ANA CELIA R$ 83,56) além da diária."],
)

TRANCOSO = dict(
 loja="TRANCOSO", arquivo="RELATORIO_HOLERITE_AGOSTO_2026_TRANCOSO_pgto_05-09-2026.xlsx",
 mult=2.0,
 EMP=["UILLIAN", "MANOEL", "VALDICK", "INIURLE", "TAMILES"],
 INOVA={"UILLIAN": ( 13, 36429.09, 11079.03, 25350.06, 589.92,  10.0,  0.0, 15.0),
        "MANOEL":  ( 92, 44540.66, 17241.62, 27299.04, 825.38,  20.0, 20.0,  9.0),
        "VALDICK": (162, 48961.39, 20272.67, 28688.72, 886.30, 110.0, 35.0, 115.0),
        "INIURLE": ( 83, 16724.49,  3314.09, 13410.40, 235.42,   0.0,  0.0,  0.0),
        "TAMILES": (103, 12924.71,  2016.24, 10908.47, 125.58,   0.0,  0.0,  0.0)},
 OUTROS_COD=[("11", "GILSON MOURA SANTOS", 21.99, 0.0, 0.0)],
 GERAL=dict(bruta=159602.33, desc=53945.60, liq=105656.73, com=2662.60, inc=334.0),
 SALARIO={"UILLIAN": 5648.41, "MANOEL": 1621.0, "VALDICK": 1621.0,
          "INIURLE": 1621.0, "TAMILES": 1621.0},
 NOTURNO={}, AUXGER={}, METACX={}, VT6={},
 SEM_COMISSAO=["UILLIAN"],
 JUL={"COMISSÃO PRODUTOS + PDV": {"INIURLE": 265.0, "MANOEL": 850.0, "TAMILES": 122.11, "VALDICK": 908.70},
      "HORAS EXTRAS":            {"INIURLE": 442.09},
      "PRÊMIO COTA GERAL":       {"INIURLE": 500.0, "MANOEL": 1500.0, "TAMILES": 300.0, "VALDICK": 1350.0},
      "PRÊMIO PRÉ-VENCIDOS":     {"VALDICK": 100.0},
      "REPOUSO REMUNERADO / DSR": {"INIURLE": 221.0, "MANOEL": 163.46, "TAMILES": 23.48, "VALDICK": 174.75},
      "SALÁRIO BASE":            {"UILLIAN": 5648.41, "MANOEL": 1621.0, "VALDICK": 1621.0,
                                  "INIURLE": 1621.0, "TAMILES": 1621.0},
      "CONVÊNIO":                {"INIURLE": -350.0, "MANOEL": -500.0, "TAMILES": -200.0, "VALDICK": -290.0},
      "DESCONTO INSS":           {"UILLIAN": -843.31, "MANOEL": -384.72, "VALDICK": -375.12,
                                  "INIURLE": -254.48, "TAMILES": -161.67}},
 JUL_SALDO={"UILLIAN": 4805.10, "MANOEL": 3249.74, "VALDICK": 3489.33,
            "INIURLE": 2444.61, "TAMILES": 1704.92},
 JUL_VT={"MANOEL": 750.0, "VALDICK": 400.0},
 JUL_VA={},
 JUL_ABA="TRANCOSO JULHO",
 BONIF={"VALDICK": 100.0, "MANOEL": 200.0, "INIURLE": 100.0, "TAMILES": 100.0},
 DSR_HE={"INIURLE": 85.02}, SAL_FAMILIA={}, INSUF={},
 VALES={"UILLIAN": -810.0, "VALDICK": -580.0, "MANOEL": -465.0},
 VT_DESC={},
 INSS={"UILLIAN": -592.27, "VALDICK": -348.72, "MANOEL": -343.29,
       "INIURLE": -236.18, "TAMILES": -157.52},
 IRRF={"UILLIAN": -251.04},
 LIQ_BETEL={"UILLIAN": 3995.10, "VALDICK": 2905.76, "MANOEL": 2980.92,
            "INIURLE": 2658.33, "TAMILES": 1862.94},
 FERIAS_OBS="Não houve férias nesta loja em agosto/2026. RENALDO (Centro) e VALDICK (Trancoso) entram de férias em 01/09 — ver a folha de setembro.",
 PONTO=[("INIURLE", "HE50", 40, "40 horas extras normais em agosto/2026.")],
 PONTO_PROX=[("TAMILES", "Atestado médico de 7 dias, entregue em 02/09/2026 — cobre de 02/09 a 08/09/2026.",
              "Cai na competência SETEMBRO/2026 (holerite pago em 05/10/2026), não nesta folha de agosto. "
              "Atestado de até 15 dias é abonado pela empresa: falta justificada, sem desconto de salário nem de DSR. "
              "A partir do 16º dia seria auxílio-doença pelo INSS. Anexar o atestado à pasta da funcionária.")],
 JUL_NOTA="OBSERVAÇÃO: em julho/2026 a comissão foi lançada por uma MÉDIA, e não pelo apurado do sistema. A partir de agosto/2026 a loja passa a pagar a comissão em dobro, e a empresa vai reduzir nas premiações a diferença gerada por esse aumento.",
 DEFINIDO=["Os prêmios, vales, INSS e IRRF vieram do holerite de agosto emitido pela Betel Contabilidade. A linha DIFERENÇA fecha em zero para os cinco.",
           "TRANCOSO PAGA O DOBRO DA COMISSÃO APURADA: a linha COMISSÃO PRODUTOS já multiplica por 2 o valor do InovaFarma (multiplicador na aba PARÂMETROS).",
           "Os incentivos são simples: entram pelo valor apurado, sem dobrar.",
           "Em julho/2026 a comissão foi lançada por média. Com a comissão dobrada a partir de agosto, a empresa vai reduzir nas premiações a diferença.",
           "UILLIAN: não recebe comissão sobre vendas (mesmo critério de julho/2026) — a rubrica fica zerada e o apurado dele aparece só na aba BASE."],
 PENDENCIAS=["TAMILES: atestado de 02/09 a 08/09/2026 — lançar na folha de SETEMBRO como falta justificada (abonada), sem desconto. Não entra nesta folha de agosto.",
             "Lançar na linha PRÊMIO COTA GERAL o valor já com a redução combinada, para compensar o aumento da comissão dobrada.",
             "UILLIAN também registra vendas no relatório da loja Centro — conferir se alguma comissão deve ser rateada entre as lojas."],
)

# ------------------------------------------------------------------ BUILDER
def build(cfg):
    EMP = cfg["EMP"]; NCOL = len(EMP)
    C0 = 2; CT = C0 + NCOL; CO = CT + 1
    LO = get_column_letter(CO)
    INOVA = cfg["INOVA"]
    wb = openpyxl.Workbook()

    def title_block(ws, text, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text); c.font = font(14, True, "FFFFFF"); c.fill = FILL_TIT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(2, 1, sub); c.font = font(10, True, "1F3864")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

    # ---------------------------------------------------------- PARÂMETROS
    ws = wb.active; ws.title = "PARÂMETROS"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABC", [46, 18, 74]):
        ws.column_dimensions[col].width = w
    title_block(ws, "PARÂMETROS DE CÁLCULO",
                f"Farmácia Tropical Multi Econômica — loja {cfg['loja']} — competência AGOSTO/2026", 3)
    P = [("SEÇÃO", "COMPETÊNCIA / PAGAMENTO", ""),
         ("Competência (mês de referência)", "AGOSTO/2026", "Vendas de 01/08/2026 a 31/08/2026 (relatório InovaFarma)."),
         ("Data do pagamento", "05/09/2026", "Folha a ser lançada no holerite pago em 05/09/2026."),
         ("Data de extração do InovaFarma", "03/09/2026", f"Relatório de comissões da loja {cfg['loja']}."),
         ("SEÇÃO", "CALENDÁRIO DO MÊS (base do DSR)", ""),
         ("Dias do mês", 31, "Agosto/2026."),
         ("Domingos + feriados", 5, "Domingos: 02, 09, 16, 23 e 30/08/2026. Sem feriado nacional em agosto."),
         ("Dias úteis (inclui sábados)", None, "Dias do mês menos domingos e feriados."),
         ("Fator DSR (domingos ÷ dias úteis)", None, "Domingos ÷ dias úteis. Usado na rubrica REPOUSO REMUNERADO / DSR."),
         ("SEÇÃO", "VALORES FIXOS / RECORRENTES", ""),
         ("Salário base (piso 2026)", 1621.00, "Valor praticado em jul/2026."),
         ("Salário do gerente", 5648.41, "Conforme jul/2026."),
         ("Horas mensais (base do salário-hora)", 220, "Salário-hora = salário base ÷ 220."),
         ("Adicional de hora extra", 0.5, "50% sobre a hora normal (hora extra = salário-hora × 1,5)."),
         ("Adicional noturno", 0.2, "20% sobre a hora normal, nas horas entre 22h e 5h."),
         ("Dias do mês para o salário-dia", 30, "Salário-dia = salário base ÷ 30. Usado no feriado trabalhado."),
         ("Multiplicador da comissão", cfg["mult"],
          "TRANCOSO paga o dobro do valor apurado no InovaFarma." if cfg["mult"] != 1
          else "1 = paga o valor apurado no InovaFarma, sem multiplicador."),
         ("SEÇÃO", "INSS / IRRF", ""),
         ("INSS", "a calcular", "Calculado pela contabilidade sobre o total de proventos (tabela progressiva vigente)."),
         ("IRRF", "a calcular", "Calculado pela contabilidade após dedução do INSS e dependentes.")]
    r = 4; PROW = {}
    for a, b, c in P:
        if a == "SEÇÃO":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            cc = ws.cell(r, 1, b); cc.font = font(11, True, "1F3864"); cc.fill = FILL_SEC
            ws.row_dimensions[r].height = 20
        else:
            ws.cell(r, 1, a).font = font(10, True)
            cc = ws.cell(r, 2, b); cc.font = font(10, False, BLUE); cc.fill = FILL_IN
            cc.alignment = Alignment(horizontal="center"); cc.border = BOX
            if isinstance(b, float):
                cc.number_format = MONEY if b > 100 else "0.00"
            PROW[a] = r
            ws.cell(r, 3, c).font = font(9, it=True)
            ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    r_dias = PROW["Dias do mês"]; r_dom = PROW["Domingos + feriados"]
    r_ute = PROW["Dias úteis (inclui sábados)"]; r_fat = PROW["Fator DSR (domingos ÷ dias úteis)"]
    ws.cell(r_ute, 2).value = f"=B{r_dias}-B{r_dom}"
    cf = ws.cell(r_fat, 2, f"=B{r_dom}/B{r_ute}")
    cf.number_format = "0.000000"; cf.font = font(10, True); cf.fill = FILL_TOT
    cf.alignment = Alignment(horizontal="center"); cf.border = BOX
    FATOR = f"PARÂMETROS!$B${r_fat}"
    MULT = f"PARÂMETROS!$B${PROW['Multiplicador da comissão']}"

    # ---------------------------------------------------------- BASE INOVA
    ws = wb.create_sheet("BASE INOVAFARMA AGO.26")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([8, 26, 16, 16, 16, 15, 13, 13, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    title_block(ws, "APURAÇÃO DE COMISSÕES E INCENTIVOS — INOVAFARMA",
                f"Loja {cfg['loja']} · vendas de 01/08/2026 a 31/08/2026 · extraído em 03/09/2026", 10)
    hdr = ["CÓD.", "VENDEDOR", "VENDA BRUTA", "DESCONTOS", "VENDA LÍQUIDA", "COMISSÃO",
           "INCENT. APLIC.", "INCENT. VITAM.", "OUTROS INCENT.", "TOTAL INCENT."]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
    ws.row_dimensions[4].height = 28
    r = 5; first = r
    for n in EMP:
        cod, vb, ds, vl, com, inj, vit, out = INOVA[n]
        ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, n).font = font(10, True)
        for i, v in enumerate([vb, ds, vl, com, inj, vit, out], 3):
            c = ws.cell(r, i, v); c.number_format = MONEY; c.font = font(10, False, BLUE)
        ws.cell(r, 10, f"=SUM(G{r}:I{r})").number_format = MONEY
        ws.cell(r, 10).font = font(10, True)
        for i in range(1, 11):
            ws.cell(r, i).border = BOX
        r += 1
    last = r - 1
    ws.cell(r, 2, "TOTAL FUNCIONÁRIOS").font = font(10, True)
    for i in range(3, 11):
        L = get_column_letter(i)
        c = ws.cell(r, i, f"=SUM({L}{first}:{L}{last})")
        c.number_format = MONEY; c.font = font(10, True)
    for i in range(1, 11):
        ws.cell(r, i).fill = FILL_TOT; ws.cell(r, i).border = BOX
    ws.auto_filter.ref = f"A4:J{last}"
    ws.freeze_panes = "A5"
    r += 2
    ws.cell(r, 1, "Códigos sem vínculo de folha nesta loja:").font = font(9, True, "1F3864")
    if cfg.get("OUTROS_NOTA"):
        ws.cell(r, 3, cfg["OUTROS_NOTA"]).font = font(9, it=True)
    r += 1
    for cod, nome, vb, com, inc in cfg["OUTROS_COD"]:
        ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, nome).font = font(9)
        ws.cell(r, 3, vb).number_format = MONEY; ws.cell(r, 3).font = font(9)
        ws.cell(r, 6, com).number_format = MONEY; ws.cell(r, 6).font = font(9)
        ws.cell(r, 10, inc).number_format = MONEY; ws.cell(r, 10).font = font(9)
        r += 1
    r += 1
    g = cfg["GERAL"]
    notas = [f"CONFERÊNCIA — total geral do relatório: venda bruta R$ {g['bruta']:,.2f} · descontos R$ {g['desc']:,.2f} · "
             f"venda líquida R$ {g['liq']:,.2f} · comissão R$ {g['com']:,.2f} · incentivo R$ {g['inc']:,.2f}."
             .replace(",", "X").replace(".", ",").replace("X", "."),
             "INCENT. APLIC. = incentivo do grupo INJETÁVEIS (aplicações). INCENT. VITAM. = grupo APLICAÇÃO E VITAMINAS INCENTIVO. OUTROS INCENT. = populares, oficinais e similar normal.",
             "A comissão desta aba é a APURADA pelo sistema, sem multiplicador. O valor que vai para a folha está na aba HOLERITE AGO.26."]
    for n in cfg["SEM_COMISSAO"]:
        notas.append(f"{n} não recebe comissão em folha: no caso dele(a) a comissão apurada acima é apenas informativa.")
    if cfg["mult"] != 1:
        notas.append("ATENÇÃO: nesta loja a comissão paga é o DOBRO do valor apurado acima.")
    for t in notas:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(r, 1, t); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1
    BASE = "'BASE INOVAFARMA AGO.26'"
    BROW = {n: first + i for i, n in enumerate(EMP)}

    # ---------------------------------------------------------- helpers folha
    def cabecalho(ws, titulo, sub):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 38
        for i in range(C0, CT + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13.5
        ws.column_dimensions[LO].width = 62
        ws.freeze_panes = "B5"
        title_block(ws, titulo, sub, CO)
        ws.cell(4, 1, "RUBRICA").font = font(10, True, "FFFFFF")
        ws.cell(4, 1).fill = FILL_HDR
        for i, n in enumerate(EMP):
            c = ws.cell(4, C0 + i, n); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(4, CT, "TOTAL"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(4, CO, "ORIGEM / OBSERVAÇÃO"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
        for i in range(1, CO + 1):
            ws.cell(4, i).border = BOX
        ws.row_dimensions[4].height = 20
        return 5

    def sec(ws, r, texto):
        for i in range(1, CO + 1):
            ws.cell(r, i).fill = FILL_SEC
        c = ws.cell(r, 1, texto); c.font = font(11, True, "1F3864")
        ws.row_dimensions[r].height = 19
        return r + 1

    def linha(ws, r, rotulo, valores, obs="", kind="in", bold=False, fill=None):
        c = ws.cell(r, 1, rotulo); c.font = font(10, bold); c.border = BOX
        for i, n in enumerate(EMP):
            cell = ws.cell(r, C0 + i)
            v = valores.get(n)
            if v is not None:
                cell.value = v
            cell.number_format = MONEY; cell.border = BOX
            if kind == "in":
                cell.font = font(10, bold, BLUE); cell.fill = fill or FILL_IN
            elif kind == "link":
                cell.font = font(10, bold, GREEN)
                if fill: cell.fill = fill
            else:
                cell.font = font(10, bold)
                if fill: cell.fill = fill
        t = ws.cell(r, CT, f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(CT-1)}{r})")
        t.number_format = MONEY; t.font = font(10, True); t.border = BOX
        t.fill = fill or FILL_TOT
        o = ws.cell(r, CO, obs); o.font = font(9, it=True); o.border = BOX
        o.alignment = Alignment(wrap_text=True, vertical="center")
        return r + 1

    # ---------------------------------------------------------- HOLERITE AGO
    ws = wb.create_sheet("HOLERITE AGO.26")
    r = cabecalho(ws, "RELATÓRIO PARA A CONTABILIDADE — LANÇAMENTO EM HOLERITE",
                  f"Loja {cfg['loja']} · competência AGOSTO/2026 (01/08 a 31/08/2026) · pagamento em 05/09/2026")
    rows = {}
    sem = cfg["SEM_COMISSAO"]
    r = sec(ws, r, "PROVENTOS")
    rows["SALÁRIO BASE"] = r
    sal = dict(cfg["SALARIO"])
    obs_sal = "Piso 2026 R$ 1.621,00; gerente conforme jul/2026. CONFERIR admissões, afastamentos e férias do mês."
    for n, motivo in cfg.get("SALARIO_ZERO", {}).items():
        sal[n] = 0.0
        obs_sal += f" {n}: zerado — {motivo}."
    r = linha(ws, r, "SALÁRIO BASE", sal, obs_sal, kind="in", fill=FILL_CONF)
    rows["ADICIONAL NOTURNO"] = r
    r = linha(ws, r, "ADICIONAL NOTURNO", cfg["NOTURNO"], "Conforme apontamento do ponto.", kind="in")
    rows["HORAS EXTRAS"] = r
    r = linha(ws, r, "HORAS EXTRAS", {}, "PREENCHER com o apurado no ponto de agosto/2026.", kind="in")
    rows["COM INOVA"] = r
    obs_com = "Apurado no InovaFarma (aba BASE INOVAFARMA AGO.26)"
    if cfg["mult"] != 1:
        obs_com += ", JÁ MULTIPLICADO POR 2 (multiplicador na aba PARÂMETROS)"
    if sem:
        obs_com += ". " + " e ".join(sem) + " não recebe(m) comissão — lançado zero"
    r = linha(ws, r, "COMISSÃO PRODUTOS (INOVAFARMA)",
              {n: (0 if n in sem else f"=ROUND({BASE}!F{BROW[n]}*{MULT},2)") for n in EMP},
              obs_com + ".", kind="link")
    rows["COM PDV"] = r
    r = linha(ws, r, "COMPLEMENTO / COMISSÃO PDV", {},
              "PREENCHER: PDV antigos e complementos de comissão acordados, se houver.", kind="in")
    rows["COMISSÃO TOTAL"] = r
    r = linha(ws, r, "COMISSÃO TOTAL (soma)",
              {n: f"=SUM({get_column_letter(C0+i)}{rows['COM INOVA']}:{get_column_letter(C0+i)}{rows['COM PDV']})"
               for i, n in enumerate(EMP)},
              "Base do DSR junto com as horas extras.", kind="calc", bold=True, fill=FILL_TOT)
    rows["DSR"] = r
    r = linha(ws, r, "REPOUSO REMUNERADO / DSR",
              {n: (f"=ROUND(({get_column_letter(C0+i)}{rows['COMISSÃO TOTAL']}"
                   f"+{get_column_letter(C0+i)}{rows['HORAS EXTRAS']})*{FATOR},2)") for i, n in enumerate(EMP)},
              "Cálculo: (comissão total + horas extras) × 5 domingos ÷ 26 dias úteis de agosto/2026.", kind="calc")
    rows["DSR HE"] = r
    r = linha(ws, r, "DSR SOBRE HORAS EXTRAS", cfg.get("DSR_HE", {}),
              "Rubrica 057 do holerite da Betel. ATENÇÃO: a rubrica 420 acima já leva as horas extras na base, "
              "então este valor paga o DSR das horas extras uma segunda vez — conferir com a contabilidade.",
              kind="in", fill=FILL_ALERT)
    rows["SALÁRIO FAMÍLIA"] = r
    r = linha(ws, r, "SALÁRIO FAMÍLIA", cfg.get("SAL_FAMILIA", {}), "Rubrica 599 do holerite da Betel.",
              kind="in", fill=FILL_CONF)
    rows["INSUFICIÊNCIA"] = r
    r = linha(ws, r, "INSUFICIÊNCIA DE SALDO", cfg.get("INSUF", {}),
              "Rubrica 998 do holerite da Betel — contrapartida do vale-transporte de quem ficou sem saldo.",
              kind="in", fill=FILL_CONF)
    rows["AUXÍLIO GERÊNCIA"] = r
    r = linha(ws, r, "AUXÍLIO GERÊNCIA", cfg["AUXGER"], "Conforme acordo da loja.", kind="in")
    rows["META CAIXA"] = r
    r = linha(ws, r, "ADICIONAL PRÊMIO META CAIXA", cfg["METACX"], "Conforme apuração do caixa.", kind="in")
    rows["PRÊMIO COTA GERAL"] = r
    obs_cota = "PREENCHER conforme a tabela de metas da loja (aba METAS da planilha original)."
    if cfg["mult"] != 1:
        obs_cota += " Lançar já com a redução combinada para compensar o aumento da comissão dobrada."
    obs_cota = ("Rubrica 011 (Bonificação/Prêmios) do holerite da Betel — já engloba cota geral e pré-vencidos. "
                + ("Lançar já com a redução combinada pela comissão dobrada." if cfg["mult"] != 1 else ""))
    r = linha(ws, r, "PRÊMIO COTA GERAL / BONIFICAÇÃO", cfg.get("BONIF", {}), obs_cota, kind="in", fill=FILL_CONF)
    rows["PRÊMIO PRÉ-VENCIDOS"] = r
    r = linha(ws, r, "PRÊMIO PRÉ-VENCIDOS", {}, "Vem somado na linha acima, na rubrica 011 do holerite.", kind="in")
    rows["FÉRIAS"] = r
    r = linha(ws, r, "FÉRIAS (DIAS GOZADOS)", {}, "PREENCHER se houve férias no mês; informar o período.", kind="in")
    rows["1/3 FÉRIAS"] = r
    r = linha(ws, r, "1/3 CONSTITUCIONAL DE FÉRIAS", {}, "Calculado pela contabilidade sobre o valor das férias.", kind="in")
    rows["INC APLIC"] = r
    inc_obs = "Incentivo de injetáveis apurado no InovaFarma."
    if cfg["mult"] != 1:
        inc_obs += " Incentivo é simples: NÃO entra em dobro."
    r = linha(ws, r, "INCENTIVO APLICAÇÕES", {n: f"={BASE}!G{BROW[n]}" for n in EMP}, inc_obs, kind="link")
    rows["INC VIT"] = r
    r = linha(ws, r, "INCENTIVO VITAMINAS", {n: f"={BASE}!H{BROW[n]}" for n in EMP},
              "Grupo APLICAÇÃO E VITAMINAS INCENTIVO no InovaFarma.", kind="link")
    rows["INC OUTROS"] = r
    r = linha(ws, r, "OUTROS INCENTIVOS", {n: f"={BASE}!I{BROW[n]}" for n in EMP},
              "Populares, oficinais e similar normal.", kind="link")
    rows["TOTAL PROVENTOS"] = r
    r = linha(ws, r, "TOTAL DE PROVENTOS",
              {n: (f"=SUM({get_column_letter(C0+i)}{rows['SALÁRIO BASE']}:{get_column_letter(C0+i)}{rows['INC OUTROS']})"
                   f"-{get_column_letter(C0+i)}{rows['COM INOVA']}-{get_column_letter(C0+i)}{rows['COM PDV']}")
               for i, n in enumerate(EMP)},
              "Soma das rubricas acima (a linha COMISSÃO TOTAL já engloba a comissão do InovaFarma e o complemento).",
              kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "DESCONTOS  (lançar com sinal negativo)")
    rows["VALES"] = r
    r = linha(ws, r, "ADIANTAMENTO SALARIAL / VALES", cfg.get("VALES", {}),
              "Rubrica 630 (Desconto - verbas / Adiantamento) do holerite da Betel.", kind="in", fill=FILL_CONF)
    rows["VALES INC"] = r
    r = linha(ws, r, "ADIANT. VALES INCENT. E APLIC.",
              {n: (f"=-({get_column_letter(C0+i)}{rows['INC APLIC']}+{get_column_letter(C0+i)}{rows['INC VIT']}"
                   f"+{get_column_letter(C0+i)}{rows['INC OUTROS']})") for i, n in enumerate(EMP)},
              "Estorno dos incentivos já adiantados em dinheiro no mês. AJUSTAR se algum incentivo não foi adiantado.",
              kind="calc", fill=FILL_CONF)
    rows["CONVÊNIO"] = r
    r = linha(ws, r, "CONVÊNIO", {}, "PREENCHER conforme o convênio de agosto.", kind="in")
    rows["FALTAS"] = r
    r = linha(ws, r, "DESCONTO FALTAS / ATRASOS", {}, "PREENCHER conforme o ponto.", kind="in")
    rows["VT6"] = r
    r = linha(ws, r, "DESCONTO VALE TRANSPORTE 6%", cfg.get("VT_DESC", cfg["VT6"]),
              "Rubrica 604 do holerite da Betel.", kind="in", fill=FILL_CONF)
    rows["INSS"] = r
    r = linha(ws, r, "DESCONTO INSS", cfg.get("INSS", {}), "Rubrica 903 do holerite da Betel.", kind="in", fill=FILL_CONF)
    rows["IRRF"] = r
    r = linha(ws, r, "DESCONTO IRRF", cfg.get("IRRF", {}), "Rubrica 914 do holerite da Betel.", kind="in", fill=FILL_CONF)
    rows["TOTAL DESC"] = r
    r = linha(ws, r, "TOTAL DE DESCONTOS",
              {n: f"=SUM({get_column_letter(C0+i)}{rows['VALES']}:{get_column_letter(C0+i)}{rows['IRRF']})"
               for i, n in enumerate(EMP)}, "", kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "LÍQUIDO")
    rows["LIQ"] = r
    r = linha(ws, r, "LÍQUIDO A RECEBER (05/09/2026)",
              {n: f"={get_column_letter(C0+i)}{rows['TOTAL PROVENTOS']}+{get_column_letter(C0+i)}{rows['TOTAL DESC']}"
               for i, n in enumerate(EMP)},
              "Total de proventos menos descontos. Só fica definitivo depois do INSS/IRRF da contabilidade.",
              kind="calc", bold=True, fill=FILL_LIQ)
    if cfg.get("LIQ_BETEL"):
        rows["LIQ BETEL"] = r
        r = linha(ws, r, "LÍQUIDO DO HOLERITE (BETEL)", cfg["LIQ_BETEL"],
                  "Valor líquido do demonstrativo emitido pela Betel Contabilidade.", kind="in", fill=FILL_CONF)
        rows["DIF BETEL"] = r
        r = linha(ws, r, "DIFERENÇA (planilha − holerite)",
                  {n: f"=ROUND({get_column_letter(C0+i)}{rows['LIQ']}-{get_column_letter(C0+i)}{rows['LIQ BETEL']},2)"
                   for i, n in enumerate(EMP)},
                  "Deve ser zero. Os incentivos entram como provento e como desconto, então não afetam o líquido.",
                  kind="calc", bold=True, fill=FILL_TOT)

    r = sec(ws, r, "INFORMATIVO — PAGO PELA EMPRESA, NÃO ENTRA NO HOLERITE")
    rows["VT"] = r
    r = linha(ws, r, "VALE TRANSPORTE (compra set/26)", {}, "PREENCHER com as passagens compradas para setembro/2026.", kind="in")
    rows["VA"] = r
    r = linha(ws, r, "VALE ALIMENTAÇÃO FERIADOS E DOMINGOS", {}, "PREENCHER conforme escala de domingos e feriados.", kind="in")
    rows["FERIAS PARTE"] = r
    r = linha(ws, r, "FÉRIAS PAGAS À PARTE (CONTAS A PAGAR)", {},
              "Recibo de férias pago fora do holerite — lembrete para o contas a pagar. "
              + cfg.get("FERIAS_OBS", "Preencher se houve férias no mês."),
              kind="in", fill=FILL_CONF)
    ws.auto_filter.ref = f"A4:{LO}{r-1}"
    r += 1
    for a, b in [("LEGENDA", ""),
                 ("Célula amarela, texto azul", "valor digitado — PREENCHER ou conferir antes de enviar."),
                 ("Célula laranja", "valor repetido de julho/2026 — CONFERIR se continua válido em agosto."),
                 ("Texto verde", "valor puxado da aba BASE INOVAFARMA AGO.26 (não digitar por cima)."),
                 ("Texto preto em célula verde", "resultado de fórmula — não alterar.")]:
        ws.cell(r, 1, a).font = font(9, True, "1F3864")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=CO)
        ws.cell(r, 2, b).font = font(9, it=True)
        r += 1
    HOL = "'HOLERITE AGO.26'"

    # ---------------------------------------------------------- JULHO REVISADO
    ws = wb.create_sheet("JULHO.26 REVISADO")
    r = cabecalho(ws, "FOLHA REVISADA — COMPETÊNCIA JULHO/2026 (paga em 05/08/2026)",
                  f"Loja {cfg['loja']} · mesmos valores pagos, reorganizados e com fórmulas conferidas · confronto do DSR")
    JUL = cfg["JUL"]; jr = {}
    ordem_prov = ["SALÁRIO BASE", "ADICIONAL NOTURNO", "HORAS EXTRAS", "COMISSÃO PRODUTOS + PDV",
                  "REPOUSO REMUNERADO / DSR", "AUXÍLIO GERÊNCIA", "PRÊMIO COTA GERAL",
                  "PRÊMIO PRÉ-VENCIDOS", "INCENTIVO APLICAÇÕES", "INCENTIVO VITAMINAS"]
    obs_jul = {"REPOUSO REMUNERADO / DSR": "Pago com fator 5/26 = 0,192307 (calendário de AGOSTO). Julho/2026 tem 4 domingos e 27 dias úteis → fator correto 0,148148. Ver conferência abaixo.",
               "COMISSÃO PRODUTOS + PDV": f"Como consta na aba {cfg['JUL_ABA']} da planilha original."}
    r = sec(ws, r, "PROVENTOS")
    for rot in ordem_prov:
        jr[rot] = r
        r = linha(ws, r, rot, JUL.get(rot, {}), obs_jul.get(rot, ""), kind="in")
    jr["TOTAL PROVENTOS"] = r
    r = linha(ws, r, "TOTAL DE PROVENTOS",
              {n: f"=SUM({get_column_letter(C0+i)}{jr[ordem_prov[0]]}:{get_column_letter(C0+i)}{jr[ordem_prov[-1]]})"
               for i, n in enumerate(EMP)}, "", kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "DESCONTOS")
    ordem_desc = ["ADIANTAMENTO / VALES", "CONVÊNIO", "DESCONTO FALTAS",
                  "DESCONTO VALE TRANSPORTE 6%", "DESCONTO INSS", "DESCONTO IRRF"]
    for rot in ordem_desc:
        jr[rot] = r
        r = linha(ws, r, rot, JUL.get(rot, {}), "", kind="in")
    jr["TOTAL DESC"] = r
    r = linha(ws, r, "TOTAL DE DESCONTOS",
              {n: f"=SUM({get_column_letter(C0+i)}{jr[ordem_desc[0]]}:{get_column_letter(C0+i)}{jr[ordem_desc[-1]]})"
               for i, n in enumerate(EMP)}, "", kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "LÍQUIDO")
    jr["LIQ"] = r
    r = linha(ws, r, "LÍQUIDO PAGO (05/08/2026)",
              {n: f"={get_column_letter(C0+i)}{jr['TOTAL PROVENTOS']}+{get_column_letter(C0+i)}{jr['TOTAL DESC']}"
               for i, n in enumerate(EMP)}, "", kind="calc", bold=True, fill=FILL_LIQ)
    jr["SALDO ORIG"] = r
    r = linha(ws, r, "SALDO FINAL DA PLANILHA ORIGINAL", cfg["JUL_SALDO"],
              f"Valor que constava na aba {cfg['JUL_ABA']} do arquivo original.", kind="in")
    jr["DIF"] = r
    r = linha(ws, r, "DIFERENÇA (revisado − original)",
              {n: f"=ROUND({get_column_letter(C0+i)}{jr['LIQ']}-{get_column_letter(C0+i)}{jr['SALDO ORIG']},2)"
               for i, n in enumerate(EMP)},
              "Deve ser zero — confirma que a revisão não alterou nenhum valor pago.",
              kind="calc", bold=True, fill=FILL_TOT)
    ws.auto_filter.ref = f"A4:{LO}{r-1}"
    r = sec(ws, r, "CONFERÊNCIA DO DSR DE JULHO/2026 (não altera o que já foi pago)")
    jr["DSR PAGO"] = r
    r = linha(ws, r, "DSR pago (fator 5/26)",
              {n: f"={get_column_letter(C0+i)}{jr['REPOUSO REMUNERADO / DSR']}" for i, n in enumerate(EMP)},
              "Fator do calendário de agosto aplicado por engano na competência de julho.", kind="calc")
    jr["DSR CORR"] = r
    r = linha(ws, r, "DSR recalculado (fator 4/27 · comissão + HE)",
              {n: (f"=ROUND(({get_column_letter(C0+i)}{jr['COMISSÃO PRODUTOS + PDV']}"
                   f"+{get_column_letter(C0+i)}{jr['HORAS EXTRAS']})*4/27,2)") for i, n in enumerate(EMP)},
              "Julho/2026: 4 domingos (05, 12, 19 e 26) e 27 dias úteis.", kind="calc")
    jr["DSR DIF"] = r
    r = linha(ws, r, "Diferença de DSR (a menor/maior)",
              {n: f"=ROUND({get_column_letter(C0+i)}{jr['DSR CORR']}-{get_column_letter(C0+i)}{jr['DSR PAGO']},2)"
               for i, n in enumerate(EMP)},
              "Valor negativo = foi pago a mais em julho. Decidir com a contabilidade se compensa em set/26.",
              kind="calc", bold=True, fill=FILL_CONF)
    r = sec(ws, r, "INFORMATIVO — PAGO PELA EMPRESA, NÃO ENTRA NO HOLERITE")
    r = linha(ws, r, "VALE TRANSPORTE", cfg["JUL_VT"], "Compra registrada na planilha original.", kind="in")
    if cfg["JUL_VA"]:
        r = linha(ws, r, "VALE ALIMENTAÇÃO FERIADOS E DOMINGOS", cfg["JUL_VA"], "", kind="in")
    r += 1
    notas_jul = []
    if cfg.get("JUL_NOTA"):
        notas_jul.append(cfg["JUL_NOTA"])
    for t in notas_jul + ["AJUSTES FEITOS NESTA REVISÃO (a planilha original continua intacta):",
              "1. Rubricas separadas em PROVENTOS × DESCONTOS, com subtotais próprios.",
              "2. Nomes de rubrica padronizados e coluna de observação com a origem de cada valor.",
              "3. Conferência do DSR: em julho foi usado o fator de agosto (5/26) em vez de 4/27.",
              "4. Linha de conferência comparando o líquido revisado com o SALDO FINAL da planilha original (tem que fechar em zero)."]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=CO)
        c = ws.cell(r, 1, t)
        alerta = t.startswith("ATENÇÃO")
        c.font = font(9, t.startswith("AJUSTES") or alerta, "C00000" if alerta else ("1F3864" if t.startswith("AJUSTES") else "000000"))
        if alerta:
            c.fill = FILL_ALERT
            ws.row_dimensions[r].height = 26
        c.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1

    # ---------------------------------------------------------- LISTA
    ws = wb.create_sheet("LISTA CONTABILIDADE")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [16, 40, 14, 16, 46]):
        ws.column_dimensions[col].width = w
    title_block(ws, f"LANÇAMENTOS POR FUNCIONÁRIO — {cfg['loja']} — AGOSTO/2026 (pagamento 05/09/2026)",
                "Mesmos valores da aba HOLERITE AGO.26, em formato de lista (vinculados por fórmula)", 5)
    for i, h in enumerate(["FUNCIONÁRIO", "RUBRICA", "TIPO", "VALOR", "OBSERVAÇÃO"], 1):
        c = ws.cell(4, i, h); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR; c.border = BOX
    LISTA = [("SALÁRIO BASE", "SALÁRIO BASE", "Provento"),
             ("ADICIONAL NOTURNO", "ADICIONAL NOTURNO", "Provento"),
             ("HORAS EXTRAS", "HORAS EXTRAS", "Provento"),
             ("COMISSÃO TOTAL", "COMISSÃO SOBRE VENDAS", "Provento"),
             ("DSR", "REPOUSO REMUNERADO / DSR", "Provento"),
             ("AUXÍLIO GERÊNCIA", "AUXÍLIO GERÊNCIA", "Provento"),
             ("META CAIXA", "ADICIONAL PRÊMIO META CAIXA", "Provento"),
             ("PRÊMIO COTA GERAL", "PRÊMIO COTA GERAL", "Provento"),
             ("PRÊMIO PRÉ-VENCIDOS", "PRÊMIO PRÉ-VENCIDOS", "Provento"),
             ("FÉRIAS", "FÉRIAS", "Provento"),
             ("1/3 FÉRIAS", "1/3 CONSTITUCIONAL DE FÉRIAS", "Provento"),
             ("INC APLIC", "INCENTIVO APLICAÇÕES", "Provento"),
             ("INC VIT", "INCENTIVO VITAMINAS", "Provento"),
             ("INC OUTROS", "OUTROS INCENTIVOS", "Provento"),
             ("TOTAL PROVENTOS", "TOTAL DE PROVENTOS", "Subtotal"),
             ("VALES", "ADIANTAMENTO SALARIAL / VALES", "Desconto"),
             ("VALES INC", "ADIANT. VALES INCENT. E APLIC.", "Desconto"),
             ("CONVÊNIO", "CONVÊNIO", "Desconto"),
             ("FALTAS", "DESCONTO FALTAS / ATRASOS", "Desconto"),
             ("VT6", "DESCONTO VALE TRANSPORTE 6%", "Desconto"),
             ("INSS", "DESCONTO INSS", "Desconto"),
             ("IRRF", "DESCONTO IRRF", "Desconto"),
             ("TOTAL DESC", "TOTAL DE DESCONTOS", "Subtotal"),
             ("LIQ", "LÍQUIDO A RECEBER", "Líquido")]
    r = 5
    for i, n in enumerate(EMP):
        col = get_column_letter(C0 + i)
        for key, rot, tipo in LISTA:
            ws.cell(r, 1, n).font = font(10, tipo == "Líquido")
            ws.cell(r, 2, rot).font = font(10)
            ws.cell(r, 3, tipo).font = font(10)
            c = ws.cell(r, 4, f"={HOL}!{col}{rows[key]}")
            c.number_format = MONEY; c.font = font(10, tipo in ("Subtotal", "Líquido"), GREEN)
            if tipo == "Líquido":
                for k in range(1, 6): ws.cell(r, k).fill = FILL_LIQ
            elif tipo == "Subtotal":
                for k in range(1, 6): ws.cell(r, k).fill = FILL_TOT
            for k in range(1, 6):
                ws.cell(r, k).border = BOX
            r += 1
        r += 1
    ws.auto_filter.ref = f"A4:E{r-2}"
    ws.freeze_panes = "A5"

    # ---------------------------------------------------------- GANHOS
    wsg = wb.create_sheet("GANHOS DO COLABORADOR")
    wsg.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [4, 44, 18, 58]):
        wsg.column_dimensions[col].width = w
    title_block(wsg, "DEMONSTRATIVO DE GANHOS DO COLABORADOR",
                f"Loja {cfg['loja']} · competência AGOSTO/2026 · pagamento em 05/09/2026", 4)
    wsg.cell(4, 2, "Escolha o colaborador:").font = font(11, True, "1F3864")
    sel = wsg.cell(4, 3, EMP[0])
    sel.font = font(12, True, BLUE); sel.fill = FILL_IN; sel.border = BOX
    sel.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1='"' + ",".join(EMP) + '"', allow_blank=False)
    dv.error = "Escolha um nome da lista."
    dv.promptTitle = "Colaborador"; dv.prompt = "Selecione o colaborador para ver os ganhos dele."
    wsg.add_data_validation(dv); dv.add(sel)
    wsg.cell(4, 4, "Troque o nome aqui e a tabela abaixo muda sozinha — é o relatório para mandar para cada um.").font = font(9, it=True)
    COLREF = f"MATCH($C$4,{HOL}!$B$4:${get_column_letter(CT-1)}$4,0)"
    RECIBO = [("SEC", "PROVENTOS", None),
              ("L", "Salário base", rows["SALÁRIO BASE"]),
              ("L", "Adicional noturno", rows["ADICIONAL NOTURNO"]),
              ("L", "Horas extras", rows["HORAS EXTRAS"]),
              ("L", "Comissão sobre vendas", rows["COMISSÃO TOTAL"]),
              ("L", "Repouso remunerado / DSR", rows["DSR"]),
              ("L", "Auxílio gerência", rows["AUXÍLIO GERÊNCIA"]),
              ("L", "Adicional prêmio meta caixa", rows["META CAIXA"]),
              ("L", "Prêmio cota geral", rows["PRÊMIO COTA GERAL"]),
              ("L", "Prêmio pré-vencidos", rows["PRÊMIO PRÉ-VENCIDOS"]),
              ("L", "Férias", rows["FÉRIAS"]),
              ("L", "1/3 constitucional de férias", rows["1/3 FÉRIAS"]),
              ("L", "Incentivo aplicações", rows["INC APLIC"]),
              ("L", "Incentivo vitaminas", rows["INC VIT"]),
              ("L", "Outros incentivos", rows["INC OUTROS"]),
              ("T", "TOTAL DE PROVENTOS", rows["TOTAL PROVENTOS"]),
              ("SEC", "DESCONTOS", None),
              ("L", "Adiantamento salarial / vales", rows["VALES"]),
              ("L", "Adiantamento de incentivos e aplicações", rows["VALES INC"]),
              ("L", "Convênio", rows["CONVÊNIO"]),
              ("L", "Faltas / atrasos", rows["FALTAS"]),
              ("L", "Vale transporte 6%", rows["VT6"]),
              ("L", "INSS", rows["INSS"]),
              ("L", "IRRF", rows["IRRF"]),
              ("T", "TOTAL DE DESCONTOS", rows["TOTAL DESC"]),
              ("SEC", "LÍQUIDO", None),
              ("Q", "LÍQUIDO A RECEBER EM 05/09/2026", rows["LIQ"])]
    r = 6
    for tipo, rot, lh in RECIBO:
        if tipo == "SEC":
            for i in range(2, 5):
                wsg.cell(r, i).fill = FILL_SEC
            wsg.cell(r, 2, rot).font = font(11, True, "1F3864")
            wsg.row_dimensions[r].height = 19
            r += 1
            continue
        c = wsg.cell(r, 2, rot)
        c.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q")); c.border = BOX
        v = wsg.cell(r, 3, f"=IFERROR(INDEX({HOL}!$B${lh}:${get_column_letter(CT-1)}${lh},1,{COLREF}),0)")
        v.number_format = MONEY; v.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q"), GREEN); v.border = BOX
        if tipo == "T":
            wsg.cell(r, 2).fill = FILL_TOT; v.fill = FILL_TOT
        if tipo == "Q":
            wsg.cell(r, 2).fill = FILL_LIQ; v.fill = FILL_LIQ
            wsg.row_dimensions[r].height = 22
        r += 1
    r += 1
    wsg.cell(r, 2, "Vendas do colaborador no mês (InovaFarma):").font = font(10, True, "1F3864")
    r += 1
    for rot, colb in [("Venda bruta", "C"), ("Descontos concedidos", "D"), ("Venda líquida", "E"),
                      ("Comissão apurada", "F"), ("Total de incentivos", "J")]:
        wsg.cell(r, 2, rot).font = font(10); wsg.cell(r, 2).border = BOX
        v = wsg.cell(r, 3, f"=IFERROR(INDEX({BASE}!${colb}${first}:${colb}${last},MATCH($C$4,{BASE}!$B${first}:$B${last},0)),0)")
        v.number_format = MONEY; v.font = font(10, False, GREEN); v.border = BOX
        r += 1
    r += 1
    rodape = ['A comissão apurada acima é a do sistema. O que entra na folha é a linha "Comissão sobre vendas".']
    if cfg["mult"] != 1:
        rodape.append("Nesta loja a comissão paga é o DOBRO da apurada — por isso os dois valores são diferentes.")
    rodape += ["Os valores de INSS e IRRF são calculados pela contabilidade e só aparecem depois de preenchidos na aba HOLERITE AGO.26.",
               "Para mandar para o colaborador: selecione o nome acima e imprima esta aba em PDF (ou tire um print)."]
    for t in rodape:
        wsg.cell(r, 2, t).font = font(9, it=True)
        r += 1
    wsg.print_area = f"A1:D{r}"
    wsg.page_setup.fitToWidth = 1
    wsg.page_setup.fitToHeight = 1
    wsg.sheet_properties.pageSetUpPr.fitToPage = True

    # ---------------------------------------------------------- PONTO
    TIPOS = {"HE50": ("Horas extras 50%", "horas", "HORAS EXTRAS"),
             "NOT": ("Adicional noturno", "horas", "ADICIONAL NOTURNO"),
             "FER": ("Feriado trabalhado (dobra)", "dias", "HORAS EXTRAS"),
             "FER_COMP": ("Feriado com folga compensatória", "dias", "—"),
             "VALOR": ("Adicional noturno (valor fixo)", "valor", "ADICIONAL NOTURNO")}
    wsp = wb.create_sheet("PONTO AGO.26")
    wsp.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [14, 32, 10, 10, 14, 14, 14, 22, 62]):
        wsp.column_dimensions[col].width = w
    title_block(wsp, "APONTAMENTO DO PONTO — AGOSTO/2026",
                f"Loja {cfg['loja']} · horas extras, adicional noturno e feriado trabalhado · alimenta a aba HOLERITE AGO.26", 9)
    for i, h in enumerate(["FUNCIONÁRIO", "OCORRÊNCIA", "QTDE", "UNIDADE", "SALÁRIO BASE",
                           "VALOR UNITÁRIO", "TOTAL", "RUBRICA DESTINO", "OBSERVAÇÃO"], 1):
        c = wsp.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
    wsp.row_dimensions[4].height = 26
    SALROW = rows["SALÁRIO BASE"]
    LEMP = get_column_letter(CT - 1)
    P_HORAS = f"PARÂMETROS!$B${PROW['Horas mensais (base do salário-hora)']}"
    P_HE = f"PARÂMETROS!$B${PROW['Adicional de hora extra']}"
    P_NOT = f"PARÂMETROS!$B${PROW['Adicional noturno']}"
    P_DIAS = f"PARÂMETROS!$B${PROW['Dias do mês para o salário-dia']}"
    pr = 5; P_INI = 5
    def linha_sal(pr):
        c = wsp.cell(pr, 5, f"=IFERROR(INDEX({HOL}!$B${SALROW}:${LEMP}${SALROW},1,"
                            f"MATCH(A{pr},{HOL}!$B$4:${LEMP}$4,0)),0)")
        c.number_format = MONEY; c.font = font(10, False, GREEN)
    for func, tipo, qtd, obs in cfg.get("PONTO", []):
        rot, uni, dest = TIPOS[tipo]
        wsp.cell(pr, 1, func).font = font(10, True)
        wsp.cell(pr, 2, rot).font = font(10)
        c = wsp.cell(pr, 3, qtd if tipo != "VALOR" else 1)
        c.font = font(10, False, BLUE); c.fill = FILL_IN
        c.alignment = Alignment(horizontal="center")
        wsp.cell(pr, 4, uni).alignment = Alignment(horizontal="center")
        linha_sal(pr)
        if tipo == "HE50":
            fu = f"=ROUND(E{pr}/{P_HORAS}*(1+{P_HE}),4)"
        elif tipo == "NOT":
            fu = f"=ROUND(E{pr}/{P_HORAS}*{P_NOT},4)"
        elif tipo == "FER":
            fu = f"=ROUND(E{pr}/{P_DIAS},2)"
        elif tipo == "FER_COMP":
            fu = 0
        else:
            fu = qtd
        c = wsp.cell(pr, 6, fu); c.number_format = MONEY
        c.font = font(10, False, BLUE if tipo == "VALOR" else "000000")
        if tipo == "VALOR":
            c.fill = FILL_CONF
        c = wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)"); c.number_format = MONEY
        c.font = font(10, True); c.fill = FILL_TOT
        wsp.cell(pr, 8, dest).font = font(9, True, "1F3864")
        c = wsp.cell(pr, 9, obs); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        for i in range(1, 10):
            wsp.cell(pr, i).border = BOX
        pr += 1
    for extra in range(8):
        for i in range(1, 10):
            cell = wsp.cell(pr, i); cell.border = BOX
            if i in (1, 2, 3, 4, 8, 9):
                cell.font = font(10, False, BLUE); cell.fill = FILL_IN
        linha_sal(pr)
        wsp.cell(pr, 6).number_format = MONEY
        wsp.cell(pr, 6).font = font(10, False, BLUE); wsp.cell(pr, 6).fill = FILL_IN
        wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)").number_format = MONEY
        wsp.cell(pr, 7).font = font(10, True); wsp.cell(pr, 7).fill = FILL_TOT
        pr += 1
    P_FIM = pr - 1
    wsp.cell(pr, 2, "TOTAL").font = font(10, True)
    c = wsp.cell(pr, 7, f"=SUM(G{P_INI}:G{P_FIM})"); c.number_format = MONEY; c.font = font(10, True)
    for i in range(1, 10):
        wsp.cell(pr, i).fill = FILL_TOT; wsp.cell(pr, i).border = BOX
    wsp.auto_filter.ref = f"A4:I{P_FIM}"
    wsp.freeze_panes = "A5"
    pr += 2
    for t in ["Salário-hora = salário base ÷ 220. Hora extra = salário-hora × 1,5. Adicional noturno = salário-hora × 20%. Feriado trabalhado sem folga = 1 salário-dia a mais (salário base ÷ 30).",
              "No salário de R$ 1.621,00: 40 horas extras dão R$ 442,09 e 24 horas dão R$ 265,25 — é a mesma conta usada na planilha de julho.",
              "As linhas em branco são para acrescentar ocorrências: preencha funcionário, ocorrência, quantidade, valor unitário e a rubrica de destino (HORAS EXTRAS ou ADICIONAL NOTURNO).",
              "Os totais desta aba entram sozinhos nas linhas HORAS EXTRAS e ADICIONAL NOTURNO da aba HOLERITE AGO.26."]:
        wsp.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=9)
        c = wsp.cell(pr, 1, t); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        wsp.row_dimensions[pr].height = 24
        pr += 1
    if cfg.get("PONTO_PROX"):
        pr += 1
        for i in range(1, 10):
            wsp.cell(pr, i).fill = FILL_SEC
        c = wsp.cell(pr, 1, "OCORRÊNCIAS PARA A PRÓXIMA COMPETÊNCIA (SETEMBRO/2026) — NÃO ENTRAM NESTA FOLHA")
        c.font = font(11, True, "1F3864")
        wsp.row_dimensions[pr].height = 19
        pr += 1
        for func, ocorrencia, detalhe in cfg["PONTO_PROX"]:
            c = wsp.cell(pr, 1, func); c.font = font(10, True); c.border = BOX
            wsp.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=9)
            c = wsp.cell(pr, 2, ocorrencia); c.font = font(10, True, "C00000"); c.fill = FILL_ALERT
            c.alignment = Alignment(wrap_text=True, vertical="center")
            pr += 1
            wsp.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=9)
            c = wsp.cell(pr, 2, detalhe); c.font = font(9, it=True)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            wsp.row_dimensions[pr].height = 30
            pr += 1

    wsh = wb["HOLERITE AGO.26"]
    for rub, lh in (("HORAS EXTRAS", rows["HORAS EXTRAS"]), ("ADICIONAL NOTURNO", rows["ADICIONAL NOTURNO"])):
        for i, n in enumerate(EMP):
            c = wsh.cell(lh, C0 + i,
                         f"=SUMIFS('PONTO AGO.26'!$G${P_INI}:$G${P_FIM},'PONTO AGO.26'!$A${P_INI}:$A${P_FIM},"
                         f"{get_column_letter(C0+i)}$4,'PONTO AGO.26'!$H${P_INI}:$H${P_FIM},\"{rub}\")")
            c.number_format = MONEY; c.font = font(10, False, GREEN); c.fill = PatternFill()
        wsh.cell(lh, CO).value = "Somado automaticamente da aba PONTO AGO.26 (apontamento de agosto/2026). Para mudar, edite lá."

    # ---------------------------------------------------------- FOLGUISTAS
    if cfg.get("FOLGUISTAS"):
        wf = wb.create_sheet("FOLGUISTAS")
        wf.sheet_view.showGridLines = False
        for col, w in zip("ABCDEFGH", [22, 8, 14, 12, 16, 16, 15, 15]):
            wf.column_dimensions[col].width = w
        title_block(wf, "FOLGUISTAS — CONTAS A PAGAR",
                    f"Loja {cfg['loja']} · agosto/2026 · NÃO entram no holerite: pagamento por diária, lançar no contas a pagar", 8)
        hdr = ["COLABORADOR", "CÓD.", "VALOR DIÁRIA", "Nº DIÁRIAS", "TOTAL DIÁRIAS",
               "COMISSÃO APURADA", "INCENTIVOS", "TOTAL A PAGAR"]
        for i, h in enumerate(hdr, 1):
            c = wf.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
            c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
        wf.row_dimensions[4].height = 28
        rr = 5; ini = rr
        for nome, cod, diaria, bruta, com, inc in cfg["FOLGUISTAS"]:
            wf.cell(rr, 1, nome).font = font(10, True)
            wf.cell(rr, 2, cod).alignment = Alignment(horizontal="center")
            c = wf.cell(rr, 3, diaria); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 4); c.font = font(10, False, BLUE); c.fill = FILL_IN
            c.alignment = Alignment(horizontal="center")
            wf.cell(rr, 5, f"=C{rr}*D{rr}").number_format = MONEY
            c = wf.cell(rr, 6, com); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 7, inc); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 8, f"=E{rr}+F{rr}+G{rr}"); c.number_format = MONEY; c.font = font(10, True)
            c.fill = FILL_LIQ
            for i in range(1, 9):
                wf.cell(rr, i).border = BOX
            rr += 1
        fim = rr - 1
        wf.cell(rr, 1, "TOTAL").font = font(10, True)
        for i in (5, 6, 7, 8):
            L = get_column_letter(i)
            c = wf.cell(rr, i, f"=SUM({L}{ini}:{L}{fim})")
            c.number_format = MONEY; c.font = font(10, True)
        for i in range(1, 9):
            wf.cell(rr, i).fill = FILL_TOT; wf.cell(rr, i).border = BOX
        wf.auto_filter.ref = f"A4:H{fim}"
        wf.freeze_panes = "A5"
        rr += 2
        for t in ["PREENCHER a coluna Nº DIÁRIAS com os dias trabalhados em agosto/2026.",
                  "Valor da diária combinado: SERGIO R$ 150,00 e ANA CELIA R$ 100,00.",
                  "Comissão e incentivos vêm do relatório do InovaFarma (códigos 51 e 131) — CONFIRMAR se o folguista recebe esses valores além da diária; se não receber, zerar as colunas.",
                  "Estes valores não entram no holerite: são pagamento de prestação de serviço, para baixa no contas a pagar."]:
            wf.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
            c = wf.cell(rr, 1, t); c.font = font(9, it=True)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            rr += 1

    # ---------------------------------------------------------- CAPA
    ws = wb.create_sheet("CAPA", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 118
    title_block(ws, f"RELATÓRIO DE FOLHA PARA A CONTABILIDADE — LOJA {cfg['loja']}", "", 2)
    blocos = [("SEC", "IDENTIFICAÇÃO"),
              ("T", f"Empresa: Farmácia Tropical Multi Econômica — loja {cfg['loja']}"),
              ("T", "Competência (mês de referência): AGOSTO/2026 — vendas de 01/08/2026 a 31/08/2026"),
              ("T", "Data do pagamento: 05/09/2026"),
              ("T", "Fonte das comissões e incentivos: InovaFarma — relatório de comissão de vendedor extraído em 03/09/2026"),
              ("T", "Funcionários: " + ", ".join(EMP)),
              ("SEC", "COMO USAR ESTE ARQUIVO"),
              ("T", "1. Abra a aba HOLERITE AGO.26 — é o relatório que vai para a contabilidade."),
              ("T", "2. Preencha as células AMARELAS (horas extras, prêmios, vales, convênio, faltas)."),
              ("T", "3. Confira as células LARANJAS — são valores repetidos de julho/2026 que podem ter mudado."),
              ("T", "4. Comissões, incentivos e DSR já vêm calculados; não digite por cima."),
              ("T", "5. A aba GANHOS DO COLABORADOR monta o demonstrativo individual: escolha o nome na lista."),
              ("T", "6. INSS e IRRF ficam em branco: são calculados pela contabilidade."),
              ("SEC", "O QUE JÁ ESTÁ APURADO (AGOSTO/2026)"),
              ("T", f"Venda bruta geral da loja: R$ {g['bruta']:,.2f} · descontos concedidos: R$ {g['desc']:,.2f} · venda líquida: R$ {g['liq']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
              ("T", f"Comissão apurada no InovaFarma: R$ {g['com']:,.2f} · incentivos: R$ {g['inc']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
              ("T", "DSR de agosto/2026: 5 domingos (02, 09, 16, 23 e 30) ÷ 26 dias úteis = fator 0,192307"),
              ("SEC", "DEFINIÇÕES DA LOJA")]
    for t in cfg["DEFINIDO"]:
        blocos.append(("T", t))
    blocos.append(("SEC", "PENDÊNCIAS — CONFIRMAR ANTES DE ENVIAR"))
    for t in ["Horas extras, adicional noturno e feriados trabalhados de agosto — lançar na aba PONTO AGO.26.",
              "Prêmio cota geral e prêmio pré-vencidos, conforme a tabela de metas da loja.",
              "Vales adiantados, convênio e faltas de agosto.",
              "Férias, afastamentos e admissões que mudem o salário do mês."] + cfg["PENDENCIAS"]:
        blocos.append(("P", t))
    blocos += [("SEC", "OBSERVAÇÃO TÉCNICA"),
               ("T", "As células de total e de cálculo são fórmulas. Ao abrir no Excel ou no Google Planilhas elas aparecem calculadas; em visualizadores simples podem aparecer em branco até o arquivo ser aberto."),
               ("SEC", "ABAS DO ARQUIVO"),
               ("T", "HOLERITE AGO.26 — relatório principal da competência agosto/2026."),
               ("T", "GANHOS DO COLABORADOR — demonstrativo individual, pronto para imprimir/mandar."),
               ("T", "LISTA CONTABILIDADE — os mesmos lançamentos em formato de lista."),
               ("T", "BASE INOVAFARMA AGO.26 — apuração de comissões e incentivos por vendedor."),
               ("T", f"JULHO.26 REVISADO — a folha de julho ({cfg['JUL_ABA']}) reorganizada e conferida."),
               ("T", "PARÂMETROS — calendário do mês, fator do DSR e valores fixos."),
               ("T", "PONTO AGO.26 — horas extras, adicional noturno e feriado trabalhado; alimenta o holerite.")]
    if cfg.get("FOLGUISTAS"):
        blocos.append(("T", "FOLGUISTAS — diárias de SERGIO e ANA CELIA para o contas a pagar (fora do holerite)."))
    r = 4
    for tipo, txt in blocos:
        if tipo == "SEC":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            c = ws.cell(r, 1, txt); c.font = font(11, True, "1F3864"); c.fill = FILL_SEC
            ws.row_dimensions[r].height = 20
        else:
            ws.cell(r, 1, "•" if tipo == "P" else "").font = font(10, True, "C00000")
            c = ws.cell(r, 2, txt); c.font = font(10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if tipo == "P":
                c.fill = FILL_ALERT
            ws.row_dimensions[r].height = 17
        r += 1

    wb.save(DIR + cfg["arquivo"])
    return DIR + cfg["arquivo"]

if __name__ == "__main__":
    for cfg in (CENTRO, TRANCOSO):
        print("ok", build(cfg))
