# -*- coding: utf-8 -*-
"""Resumo consolidado das 3 lojas - competencia AGOSTO/2026, pagamento 05/09/2026."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/site-farmacia/relatorios/contabilidade/RESUMO_3_LOJAS_AGOSTO_2026_pgto_05-09-2026.xlsx"
F = "Arial"
MONEY = 'R$ #,##0.00;-R$ #,##0.00;"-"'
def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)
FILL_TIT = PatternFill("solid", fgColor="1F3864")
FILL_SEC = PatternFill("solid", fgColor="D9E2F3")
FILL_HDR = PatternFill("solid", fgColor="2E5496")
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")
FILL_LIQ = PatternFill("solid", fgColor="C6E0B4")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")
FILL_ALERT = PatternFill("solid", fgColor="FFF0F0")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

FATOR = 5 / 26          # agosto/2026: 5 domingos / 26 dias uteis
HORA = 1621 / 220
HE = round(HORA * 1.5, 4)     # 11,0523
NOT_ = round(HORA * 0.2, 4)   # 1,4736
DIA = round(1621 / 30, 2)     # 54,03

def E(nome, sal, com, he=0.0, notu=0.0, aux=0.0, meta=0.0, inc=0.0, vt6=0.0, obs=""):
    dsr = round((com + he) * FATOR, 2)
    prov = round(sal + he + notu + com + dsr + aux + meta + inc, 2)
    desc = round(-inc + vt6, 2)
    return dict(nome=nome, sal=sal, he=he, notu=notu, com=com, dsr=dsr, aux=aux,
                meta=meta, inc=inc, prov=prov, vt6=vt6, adiant_inc=-inc, adiant_sal=None,
                desc=desc, liq=round(prov + desc, 2), obs=obs)

ARRAIAL = [
 E("DEAN",    5648.41, 0.0,   inc=199.0, obs="Não recebe comissão. Dia 15/08 trabalhado com folga compensatória."),
 E("AGNOR",   1621.00, 2000.00, notu=round(NOT_*24,2), aux=810.50, inc=189.0,
   obs="Comissão fixa de R$ 2.000,00 (apurado R$ 909,51 + complemento R$ 1.090,49). 3 madrugadas = 24 h noturnas."),
 E("EDEY",    1621.00, 873.76, he=round(HE*24,2), notu=350.00, vt6=-84.29,
   inc=73.0, obs="24 horas extras. Adicional noturno fixo de R$ 350,00 — conferir com o ponto."),
 E("JOEL",    0.0,     103.31, inc=22.0,
   obs="Férias de 01 a 31/08 pagas em recibo próprio (saiu em 04/08, volta em 04/09). Comissão dos dias 01 e 02/08."),
 E("VALÉRIA", 1621.00, 1141.03, he=DIA, notu=round(NOT_*16,2), meta=307.99, inc=192.0, vt6=-84.29,
   obs="Feriado 15/08 trabalhado. 2 madrugadas = 16 h noturnas, com folga no dia seguinte."),
 E("SARA",    1621.00, 371.43, he=DIA, aux=810.50, inc=10.0,
   obs="Comissão do Arraial (R$ 368,78) + loja Centro (R$ 2,65). Feriado 15/08 trabalhado."),
 E("CAMILA",  1621.00, 241.50, he=DIA, inc=0.0,
   obs="1º mês completo após a admissão. Feriado 15/08 trabalhado."),
 E("NATI",    1621.00, 243.14, he=round(HE*17,2)+DIA, notu=round(NOT_*17,2), meta=307.99,
   inc=0.0, vt6=-84.29,
   obs="17 horas extras noturnas (hora extra + adicional). Feriado 15/08 trabalhado."),
]
CENTRO = [
 E("ARIANE",  5648.41, 0.0, inc=125.0, obs="Não recebe comissão."),
 E("ELIANA",  1621.00, 1511.36, inc=225.0, vt6=-84.29, obs=""),
 E("GENECIR", 0.0,     0.62, inc=0.0,
   obs="Férias de 01 a 30/08 pagas em recibo próprio — conferir o dia 31/08."),
 E("RENALDO", 1621.00, 847.94, inc=96.0, vt6=-74.29, obs=""),
 E("THAYANE", 1621.00, 57.33, inc=10.0, vt6=-84.29, obs=""),
 E("PEDRO",   1621.00, 0.0, obs="Não recebe comissão (apurado R$ 12,18 fica só como informação)."),
]
TRANCOSO = [
 E("UILLIAN", 5648.41, 0.0, inc=25.0, obs="Não recebe comissão."),
 E("MANOEL",  1621.00, 1650.76, inc=49.0, obs="Comissão apurada R$ 825,38 × 2."),
 E("VALDICK", 1621.00, 1772.60, inc=260.0, obs="Comissão apurada R$ 886,30 × 2."),
 E("INIURLE", 1621.00, 470.84, he=round(HE*40,2), inc=0.0,
   obs="Comissão apurada R$ 235,42 × 2. 40 horas extras."),
 E("TAMILES", 1621.00, 251.16, inc=0.0,
   obs="Comissão apurada R$ 125,58 × 2. Atestado de 02 a 08/09 é competência de setembro."),
]
LOJAS = [("ARRAIAL", ARRAIAL), ("CENTRO", CENTRO), ("TRANCOSO", TRANCOSO)]

COLS = [("FUNCIONÁRIO", "nome", 15), ("SALÁRIO", "sal", 12), ("HORAS EXTRAS", "he", 12),
        ("AD. NOTURNO", "notu", 12), ("COMISSÃO", "com", 12), ("DSR", "dsr", 11),
        ("AUX. GERÊNCIA", "aux", 12), ("PRÊMIO META CX", "meta", 12), ("INCENTIVOS", "inc", 12),
        ("TOTAL PROVENTOS", "prov", 14), ("ADIANT. INCENT.", "adiant_inc", 13),
        ("ADIANTAMENTO SALARIAL / VALES", "adiant_sal", 15),
        ("VALE TRANSP. 6%", "vt6", 13), ("TOTAL DESCONTOS", "desc", 14),
        ("LÍQUIDO PARCIAL", "liq", 14), ("OBSERVAÇÃO", "obs", 70)]

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "RESUMO 3 LOJAS AGO.26"
ws.sheet_view.showGridLines = False
for i, (_, _, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
N = len(COLS)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
c = ws.cell(1, 1, "RESUMO PARA A CONTABILIDADE — TRÊS LOJAS")
c.font = font(14, True, "FFFFFF"); c.fill = FILL_TIT
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
c = ws.cell(2, 1, "Farmácia Tropical Multi Econômica · competência AGOSTO/2026 (01/08 a 31/08) · pagamento em 05/09/2026")
c.font = font(10, True, "1F3864"); c.alignment = Alignment(horizontal="center")
r = 4
for i, (h, _, _) in enumerate(COLS, 1):
    c = ws.cell(r, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
ws.row_dimensions[r].height = 30
r += 1
ws.freeze_panes = "A5"
ini_geral = r
linhas_loja = []
for loja, emps in LOJAS:
    for i in range(1, N + 1):
        ws.cell(r, i).fill = FILL_SEC
    c = ws.cell(r, 1, f"LOJA {loja}"); c.font = font(11, True, "1F3864")
    ws.row_dimensions[r].height = 19
    r += 1
    ini = r
    KEYS = [k for _, k, _ in COLS]
    L = {k: get_column_letter(i) for i, k in enumerate(KEYS, 1)}
    for e in emps:
        for i, (_, k, _) in enumerate(COLS, 1):
            if k == "prov":
                v = f"=SUM({L['sal']}{r}:{L['inc']}{r})"
            elif k == "desc":
                v = f"=SUM({L['adiant_inc']}{r}:{L['vt6']}{r})"
            elif k == "liq":
                v = f"={L['prov']}{r}+{L['desc']}{r}"
            else:
                v = e[k]
            c = ws.cell(r, i, v)
            c.border = BOX
            if k == "nome":
                c.font = font(10, True)
            elif k == "obs":
                c.font = font(9, it=True)
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                c.number_format = MONEY
                c.font = font(10, True if k == "liq" else False)
                if k == "liq":
                    c.fill = FILL_LIQ
                elif k in ("prov", "desc"):
                    c.fill = FILL_TOT
                elif k == "adiant_sal":
                    c.font = font(10, False, "0000FF"); c.fill = FILL_IN
        r += 1
    fim = r - 1
    linhas_loja.append((loja, ini, fim))
    ws.cell(r, 1, f"TOTAL {loja}").font = font(10, True)
    for i, (_, k, _) in enumerate(COLS, 1):
        if k in ("nome", "obs"):
            continue
        L = get_column_letter(i)
        c = ws.cell(r, i, f"=SUM({L}{ini}:{L}{fim})")
        c.number_format = MONEY; c.font = font(10, True)
    for i in range(1, N + 1):
        ws.cell(r, i).fill = FILL_TOT; ws.cell(r, i).border = BOX
    linhas_loja[-1] = (loja, ini, fim, r)
    r += 1
r += 1
ws.cell(r, 1, "TOTAL GERAL — 3 LOJAS").font = font(11, True)
for i, (_, k, _) in enumerate(COLS, 1):
    if k in ("nome", "obs"):
        continue
    L = get_column_letter(i)
    partes = "+".join(f"{L}{t}" for _, _, _, t in linhas_loja)
    c = ws.cell(r, i, f"={partes}")
    c.number_format = MONEY; c.font = font(11, True)
for i in range(1, N + 1):
    ws.cell(r, i).fill = FILL_LIQ; ws.cell(r, i).border = BOX
ws.row_dimensions[r].height = 22
r += 2
avisos = [
 "O QUE AINDA NÃO ESTÁ NESTE RESUMO — precisa ser somado antes de fechar o holerite:",
 "• INSS e IRRF de cada funcionário (cálculo da contabilidade) — por isso a última coluna é LÍQUIDO PARCIAL.",
 "• Convênio e descontos de falta do mês.",
 "• A coluna ADIANTAMENTO SALARIAL / VALES está em amarelo, para preencher com o valor adiantado a cada um (em negativo). O total de descontos e o líquido se atualizam sozinhos.",
 "• Prêmio cota geral e prêmio pré-vencidos, conforme a apuração de metas de cada loja.",
 "• Diárias dos folguistas SERGIO (R$ 150,00) e ANA CELIA (R$ 100,00), na loja Centro — vão para o contas a pagar, fora do holerite.",
 "",
 "CRITÉRIOS USADOS:",
 "• DSR sobre comissão e horas extras: 5 domingos ÷ 26 dias úteis de agosto/2026 = fator 0,192307.",
 "• Salário-hora = salário ÷ 220. Hora extra = salário-hora × 1,5 (R$ 11,0523). Adicional noturno = salário-hora × 20% (R$ 1,4736).",
 "• Feriado de 15/08 trabalhado sem folga = 1 salário-dia a mais (R$ 54,03).",
 "• TRANCOSO paga o dobro da comissão apurada no InovaFarma; os incentivos entram pelo valor simples.",
 "• Os incentivos são adiantados em dinheiro durante o mês, por isso aparecem como provento e como desconto (ADIANT. INCENT.).",
 "• Comissões e incentivos vêm do relatório do InovaFarma de 01/08 a 31/08/2026, extraído em 03/09/2026.",
]
for t in avisos:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    c = ws.cell(r, 1, t)
    c.font = font(9, t.endswith(":"), "C00000" if t.startswith("O QUE") else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    if t.startswith("O QUE"):
        c.fill = FILL_ALERT
    r += 1
ws.auto_filter.ref = f"A4:{get_column_letter(N)}{linhas_loja[-1][3]}"
wb.save(OUT)
print("ok", OUT)
