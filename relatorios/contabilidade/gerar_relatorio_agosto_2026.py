# -*- coding: utf-8 -*-
"""Gera o relatorio de folha (holerite) AGOSTO/2026 - pagamento 05/09/2026."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/home/user/site-farmacia/relatorios/contabilidade/RELATORIO_HOLERITE_AGOSTO_2026_pgto_05-09-2026.xlsx"

F = "Arial"
MONEY = 'R$ #,##0.00;-R$ #,##0.00;"-"'
PCT = '0.00%'

def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

FILL_TIT   = PatternFill("solid", fgColor="1F3864")   # azul escuro
FILL_SEC   = PatternFill("solid", fgColor="D9E2F3")   # azul claro (secao)
FILL_HDR   = PatternFill("solid", fgColor="2E5496")
FILL_IN    = PatternFill("solid", fgColor="FFF2CC")   # amarelo = preencher
FILL_CONF  = PatternFill("solid", fgColor="FCE4D6")   # laranja = conferir
FILL_TOT   = PatternFill("solid", fgColor="E2EFDA")   # verde = totais
FILL_LIQ   = PatternFill("solid", fgColor="C6E0B4")
FILL_ALERT = PatternFill("solid", fgColor="FFF0F0")

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

BLUE = "0000FF"   # entrada digitada
GREEN = "008000"  # link entre abas

EMP = ["DEAN", "AGNOR", "EDEY", "JOEL", "VALÉRIA", "SARA", "CAMILA", "NATI"]
NCOL = len(EMP)
C0 = 2                       # coluna B
CT = C0 + NCOL               # coluna do TOTAL (J)
CO = CT + 1                  # coluna de observacao (K)
LT = get_column_letter(CT)
LO = get_column_letter(CO)

# ---------------------------------------------------------------- dados AGOSTO
# InovaFarma - Relatorio detalhado de comissao de vendedor, 01/08/2026 a 31/08/2026
INOVA = {          # nome_planilha: (cod, venda_bruta, desconto, venda_liq, comissao, inj, vit, outros)
 "DEAN":    (242, 53167.33, 22509.36, 30657.97,  975.65, 190.0,  0.0,  9.0),
 "AGNOR":   ( 53, 51871.06, 19353.31, 32517.75,  909.51, 150.0, 25.0, 14.0),
 "EDEY":    (272, 49696.15, 18497.95, 31198.20,  873.76,  40.0, 30.0,  3.0),
 "JOEL":    (181,  5057.69,  1361.04,  3696.65,  103.31,  20.0,  0.0,  2.0),
 "VALÉRIA": (292, 61902.54, 22544.73, 39357.81, 1141.03, 140.0, 35.0, 17.0),
 "SARA":    ( 43, 29798.09,  3068.04, 26730.05,  368.78,   0.0,  0.0, 10.0),
 "CAMILA":  (312, 21520.43,  1007.45, 20512.98,  241.50,   0.0,  0.0,  0.0),
 "NATI":    (212, 26696.39,  1321.75, 25374.64,  243.14,   0.0,  0.0,  0.0),
}
OUTROS_COD = [("11", "GILSON MOURA SANTOS", 603.28, 0.0018, 0.0),
              ("41", "PEDRO", 10.41, 0.0, 0.0),
              ("73", "BALCONISTA", 48.58, 0.0, 0.0)]

# valores recorrentes (vindos da competencia JULHO/2026) -> conferir
SALARIO   = {"DEAN": 5648.41, "AGNOR": 1621.0, "EDEY": 1621.0, "JOEL": 0.0,
             "VALÉRIA": 1621.0, "SARA": 1621.0, "CAMILA": 1621.0, "NATI": 1621.0}
NOTURNO   = {"EDEY": 350.0}
AUXGER    = {"AGNOR": 810.5, "SARA": 810.5}
METACX    = {"VALÉRIA": 307.99, "NATI": 307.99}
VT6       = {"EDEY": -84.29, "VALÉRIA": -84.29, "NATI": -84.29}

# ------------------------------------------------------------------ JULHO/2026
JUL = {  # rubrica -> {func: valor}   (exatamente como foi pago em 05/08/2026)
 "ADICIONAL PRÊMIO META CAIXA": {"VALÉRIA": 307.99, "NATI": 307.99},
 "ADICIONAL NOTURNO":           {"EDEY": 350.0},
 "COMISSÃO PRODUTOS + PDV":     {"AGNOR": 2000.0, "EDEY": 1202.0, "JOEL": 1693.68,
                                 "VALÉRIA": 1071.0, "SARA": 345.0, "CAMILA": 120.0, "NATI": 280.0},
 "AUXÍLIO GERÊNCIA":            {"AGNOR": 810.5, "SARA": 810.5},
 "HORAS EXTRAS":                {"EDEY": 442.09, "NATI": 221.05},
 "PRÊMIO COTA GERAL":           {"EDEY": 250.0, "JOEL": 250.0, "VALÉRIA": 250.0, "SARA": 150.0, "NATI": 150.0},
 "PRÊMIO PRÉ-VENCIDOS":         {"EDEY": 100.0, "JOEL": 200.0, "VALÉRIA": 250.0, "SARA": 100.0, "NATI": 100.0},
 "REPOUSO REMUNERADO / DSR":    {"AGNOR": 384.62, "EDEY": 401.19, "JOEL": 325.71,
                                 "VALÉRIA": 205.96, "SARA": 66.35, "NATI": 138.87},
 "SALÁRIO BASE":                {"DEAN": 5648.41, "AGNOR": 1621.0, "EDEY": 1621.0, "JOEL": 1621.0,
                                 "VALÉRIA": 1621.0, "SARA": 1621.0, "CAMILA": 595.0, "NATI": 1621.0},
 "INCENTIVO APLICAÇÕES":        {"AGNOR": 264.0, "EDEY": 100.0, "JOEL": 290.0, "VALÉRIA": 292.0},
 "INCENTIVO VITAMINAS":         {"VALÉRIA": 40.0},
 "CONVÊNIO":                    {"AGNOR": -300.0, "JOEL": -500.0, "VALÉRIA": -300.0, "SARA": -300.0},
 "ADIANTAMENTO / VALES":        {"AGNOR": -2300.0, "EDEY": -1500.0, "JOEL": -1300.0,
                                 "VALÉRIA": -1300.0, "SARA": -1200.0},
 "ADIANT. VALES INCENT. E APLIC.": {"AGNOR": -264.0, "EDEY": -100.0, "JOEL": -290.0, "VALÉRIA": -332.0},
 "DESCONTO FALTAS":             {},
 "DESCONTO VALE TRANSPORTE 6%": {"EDEY": -84.29, "JOEL": -84.29, "VALÉRIA": -84.29, "NATI": -84.29},
 "DESCONTO IRRF":               {"DEAN": -251.04},
 "DESCONTO INSS":               {"DEAN": -592.27, "AGNOR": -480.85, "EDEY": -426.78, "JOEL": -414.75,
                                 "VALÉRIA": -373.14, "SARA": -213.94, "NATI": -229.38},
}
JUL_SALDO = {"DEAN": 4805.10, "AGNOR": 1735.27, "EDEY": 2355.21, "JOEL": 1791.35,
             "VALÉRIA": 1648.52, "SARA": 1378.91, "CAMILA": 715.0, "NATI": 2505.24}
JUL_VT    = {"DEAN": 350.0, "EDEY": 416.0, "VALÉRIA": 500.0, "CAMILA": 135.0, "NATI": 288.0}
JUL_VA    = {"DEAN": 647.0, "AGNOR": 84.0, "EDEY": 48.0, "JOEL": 50.0, "VALÉRIA": 75.0, "SARA": 84.0}

wb = openpyxl.Workbook()

# =============================================================== aba PARÂMETROS
ws = wb.active
ws.title = "PARÂMETROS"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 74

def title_block(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text); c.font = font(14, True, "FFFFFF"); c.fill = FILL_TIT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, sub); c.font = font(10, True, "1F3864")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

title_block(ws, "PARÂMETROS DE CÁLCULO", "Farmácia Tropical Multi Econômica — competência AGOSTO/2026", 3)

P = [
 ("SEÇÃO", "COMPETÊNCIA / PAGAMENTO", ""),
 ("Competência (mês de referência)", "AGOSTO/2026", "Vendas de 01/08/2026 a 31/08/2026 (relatório InovaFarma)."),
 ("Data do pagamento", "05/09/2026", "Folha a ser lançada no holerite pago em 05/09/2026."),
 ("Data de extração do InovaFarma", "03/09/2026", "Arquivo Comissao_de_Vendedores03092026_080157.xlsx."),
 ("SEÇÃO", "CALENDÁRIO DO MÊS (base do DSR)", ""),
 ("Dias do mês", 31, "Agosto/2026."),
 ("Domingos + feriados", 5, "Domingos: 02, 09, 16, 23 e 30/08/2026. Sem feriado nacional em agosto."),
 ("Dias úteis (inclui sábados)", 26, "Dias do mês menos domingos e feriados."),
 ("Fator DSR (domingos ÷ dias úteis)", None, "Domingos ÷ dias úteis. Usado na rubrica REPOUSO REMUNERADO / DSR da aba HOLERITE AGO.26."),
 ("SEÇÃO", "VALORES FIXOS / RECORRENTES", ""),
 ("Salário base (piso 2026)", 1621.00, "Valor praticado em jul/2026 para todos, exceto DEAN."),
 ("Salário DEAN", 5648.41, "Conforme jul/2026."),
 ("Auxílio gerência", 810.50, "AGNOR e SARA (conforme jul/2026)."),
 ("Adicional prêmio meta caixa", 307.99, "VALÉRIA e NATI (conforme jul/2026)."),
 ("Adicional noturno EDEY", 350.00, "Referência de jul/2026. O valor efetivo é lançado na aba PONTO AGO.26 — conferir com o apontamento."),
 ("Comissão fixa acordada — AGNOR", 2000.00, "Valor fixo mantido pela empresa. O complemento é a diferença entre este valor e a comissão apurada no InovaFarma."),
 ("Horas mensais (base do salário-hora)", 220, "Salário-hora = salário base ÷ 220."),
 ("Adicional de hora extra", 0.5, "50% sobre a hora normal (hora extra = salário-hora × 1,5)."),
 ("Adicional noturno", 0.2, "20% sobre a hora normal, nas horas entre 22h e 5h."),
 ("Dias do mês para o salário-dia", 30, "Salário-dia = salário base ÷ 30. Usado no feriado trabalhado."),
 ("Desconto vale-transporte 6%", -84.29, "Valor praticado em jul/2026. 6% do salário base daria R$ 97,26 — CONFERIR."),
 ("SEÇÃO", "INSS / IRRF", ""),
 ("INSS", "a calcular", "Calculado pela contabilidade sobre o total de proventos (tabela progressiva vigente)."),
 ("IRRF", "a calcular", "Calculado pela contabilidade após dedução do INSS e dependentes."),
]
r = 4
PROW = {}
for a, b, c in P:
    if a == "SEÇÃO":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cc = ws.cell(r, 1, b); cc.font = font(11, True, "1F3864"); cc.fill = FILL_SEC
        ws.row_dimensions[r].height = 20
    else:
        ws.cell(r, 1, a).font = font(10, True)
        cc = ws.cell(r, 2, b)
        cc.font = font(10, False, BLUE)
        cc.fill = FILL_IN
        cc.alignment = Alignment(horizontal="center")
        cc.border = BOX
        if isinstance(b, float):
            cc.number_format = MONEY
        PROW[a] = r
        ws.cell(r, 3, c).font = font(9, it=True)
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="center")
    r += 1
r_dom = PROW["Domingos + feriados"]
r_ute = PROW["Dias úteis (inclui sábados)"]
r_fat = PROW["Fator DSR (domingos ÷ dias úteis)"]
cf = ws.cell(r_fat, 2, f"=B{r_dom}/B{r_ute}")
cf.number_format = "0.000000"; cf.font = font(10, True); cf.fill = FILL_TOT
cf.alignment = Alignment(horizontal="center"); cf.border = BOX
ws.cell(r_ute, 2).value = f"=B{PROW['Dias do mês']}-B{r_dom}"
FATOR = f"PARÂMETROS!$B${r_fat}"

# ============================================================= aba BASE INOVA
ws = wb.create_sheet("BASE INOVAFARMA AGO.26")
ws.sheet_view.showGridLines = False
widths = [8, 26, 16, 16, 16, 15, 13, 13, 13, 13]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_block(ws, "APURAÇÃO DE COMISSÕES E INCENTIVOS — INOVAFARMA",
            "Relatório detalhado de comissão de vendedor · vendas de 01/08/2026 a 31/08/2026 · extraído em 03/09/2026", 10)
hdr = ["CÓD.", "VENDEDOR", "VENDA BRUTA", "DESCONTOS", "VENDA LÍQUIDA", "COMISSÃO",
       "INCENT. APLIC.", "INCENT. VITAM.", "OUTROS INCENT.", "TOTAL INCENT."]
r = 4
for i, h in enumerate(hdr, 1):
    c = ws.cell(r, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
ws.row_dimensions[r].height = 28
r = 5
first = r
for n in EMP:
    cod, vb, ds, vl, com, inj, vit, out = INOVA[n]
    ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
    ws.cell(r, 2, n).font = font(10, True)
    for i, v in enumerate([vb, ds, vl, com, inj, vit, out], 3):
        c = ws.cell(r, i, v); c.number_format = MONEY; c.font = font(10, False, BLUE)
    ws.cell(r, 10, f"=SUM(G{r}:I{r})").number_format = MONEY
    ws.cell(r, 10).font = font(10, True)
    for i in range(1, 11):
        ws.cell(r, i).border = BOX
    r += 1
last = r - 1
ws.cell(r, 2, "TOTAL FUNCIONÁRIOS").font = font(10, True)
for i in range(3, 11):
    c = ws.cell(r, i, f"=SUM({get_column_letter(i)}{first}:{get_column_letter(i)}{last})")
    c.number_format = MONEY; c.font = font(10, True)
for i in range(1, 11):
    ws.cell(r, i).fill = FILL_TOT; ws.cell(r, i).border = BOX
tot_row = r
r += 2
ws.cell(r, 1, "Códigos sem vínculo de folha (não entram no holerite):").font = font(9, True, "1F3864")
r += 1
for cod, nome, vb, com, inc in OUTROS_COD:
    ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
    ws.cell(r, 2, nome).font = font(9)
    ws.cell(r, 3, vb).number_format = MONEY; ws.cell(r, 3).font = font(9)
    ws.cell(r, 6, com).number_format = MONEY; ws.cell(r, 6).font = font(9)
    r += 1
r += 1
notas = [
 "CONFERÊNCIA — total geral do relatório InovaFarma: venda bruta R$ 300.371,95 · descontos R$ 90.276,65 · venda líquida R$ 210.095,30 · comissão R$ 4.856,69 · incentivo R$ 685,00.",
 "INCENT. APLIC. = incentivo do grupo INJETÁVEIS (aplicações). INCENT. VITAM. = grupo APLICAÇÃO E VITAMINAS INCENTIVO. OUTROS INCENT. = populares, oficinais e similar normal.",
 "A comissão acima é a APURADA pelo sistema. Comissões de PDV antigos e valores fixos acordados são lançados à parte na aba HOLERITE AGO.26.",
 "SARA também vendeu R$ 365,03 pelo código dela na loja CENTRO, com R$ 2,65 de comissão. Esse valor foi somado à folha dela aqui, na linha COMPLEMENTO / COMISSÃO PDV.",
 "DEAN não recebe comissão em folha: no caso dele a comissão apurada acima é apenas informativa. Os incentivos dele continuam sendo lançados normalmente.",
 "JOEL: vendeu R$ 3.684,97 nos dias 01 e 02/08 e saiu de férias em 04/08 (retorno em 04/09) — daí a comissão de R$ 103,31. Constam ainda 2 itens de R$ 11,68 em 07/08 no código dele.",
]
for t in notas:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(r, 1, t); c.font = font(9, it=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26
    r += 1
ws.auto_filter.ref = f"A4:J{last}"
ws.freeze_panes = "A5"
BASE = "'BASE INOVAFARMA AGO.26'"
BASE_ROW = {n: first + i for i, n in enumerate(EMP)}

# ============================================================ aba HOLERITE AGO
def build_folha(ws, titulo, subtitulo, linhas, obs_col=True):
    """linhas: lista de tuplas (tipo, rotulo, dict/def, obs)"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    for i in range(C0, CT + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13.5
    ws.column_dimensions[LO].width = 62
    ws.freeze_panes = "B5"
    title_block(ws, titulo, subtitulo, CO)
    r = 4
    ws.cell(r, 1, "RUBRICA").font = font(10, True, "FFFFFF")
    ws.cell(r, 1).fill = FILL_HDR
    for i, n in enumerate(EMP):
        c = ws.cell(r, C0 + i, n); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(r, CT, "TOTAL"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
    c.alignment = Alignment(horizontal="center")
    c = ws.cell(r, CO, "ORIGEM / OBSERVAÇÃO"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
    for i in range(1, CO + 1):
        ws.cell(r, i).border = BOX
    ws.row_dimensions[r].height = 20
    return r + 1

def sec(ws, r, texto):
    # sem mesclar: a aba tem filtro e celulas mescladas atrapalham o filtro/ordenacao
    for i in range(1, CO + 1):
        ws.cell(r, i).fill = FILL_SEC
    c = ws.cell(r, 1, texto); c.font = font(11, True, "1F3864")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 20
    return r + 1

def linha(ws, r, rotulo, valores, obs="", kind="in", bold=False, fill=None, negativo=False):
    """valores: dict func->valor|formula. kind: in (digitado) / calc / link"""
    c = ws.cell(r, 1, rotulo); c.font = font(10, bold); c.border = BOX
    for i, n in enumerate(EMP):
        cell = ws.cell(r, C0 + i)
        v = valores.get(n)
        if v is not None:
            cell.value = v
        cell.number_format = MONEY
        cell.border = BOX
        if kind == "in":
            cell.font = font(10, bold, BLUE); cell.fill = fill or FILL_IN
        elif kind == "link":
            cell.font = font(10, bold, GREEN); cell.fill = fill or PatternFill()
        else:
            cell.font = font(10, bold); 
            if fill: cell.fill = fill
    t = ws.cell(r, CT, f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(CT-1)}{r})")
    t.number_format = MONEY; t.font = font(10, True); t.border = BOX
    t.fill = fill or FILL_TOT
    o = ws.cell(r, CO, obs); o.font = font(9, it=True); o.alignment = Alignment(wrap_text=True, vertical="center")
    o.border = BOX
    return r + 1

ws = wb.create_sheet("HOLERITE AGO.26")
r = build_folha(ws, "RELATÓRIO PARA A CONTABILIDADE — LANÇAMENTO EM HOLERITE",
                "Competência AGOSTO/2026 (01/08 a 31/08/2026) · Pagamento em 05/09/2026 · Farmácia Tropical Multi Econômica",
                None)
rows = {}
r = sec(ws, r, "PROVENTOS")
rows["SALÁRIO BASE"] = r
r = linha(ws, r, "SALÁRIO BASE", {n: SALARIO[n] for n in EMP},
          "Piso 2026 R$ 1.621,00. CAMILA: admitida em jul/26 (salário proporcional naquele mês); agosto é o 1º mês completo. JOEL: zerado — trabalhou de 01 a 03/08 mas esses dias já estão dentro do recibo de férias (01 a 31/08); se a empresa decidir pagá-los à parte, lançar aqui.",
          kind="in", fill=FILL_CONF)
rows["ADICIONAL NOTURNO"] = r
r = linha(ws, r, "ADICIONAL NOTURNO", NOTURNO, "Conforme apontamento do ponto (jul/26: EDEY R$ 350,00).", kind="in", fill=FILL_CONF)
rows["HORAS EXTRAS"] = r
r = linha(ws, r, "HORAS EXTRAS", {}, "PREENCHER com o apurado no ponto de agosto/2026.", kind="in")
rows["COMISSÃO PRODUTOS (INOVAFARMA)"] = r
r = linha(ws, r, "COMISSÃO PRODUTOS (INOVAFARMA)",
          {n: (0 if n == "DEAN" else f"={BASE}!F{BASE_ROW[n]}") for n in EMP},
          "Apurado no InovaFarma (aba BASE INOVAFARMA AGO.26). JOEL: R$ 103,31 das vendas dele nos dias 01 e 02/08, antes de sair de férias em 04/08 (o recibo de férias saiu como 01 a 31/08). DEAN NÃO RECEBE COMISSÃO — lançado zero (o apurado dele fica só como informação na aba BASE). AGNOR: mantido o fixo de R$ 2.000,00 pela linha de complemento abaixo.",
          kind="link")
rows["COMPLEMENTO / COMISSÃO PDV"] = r
r = linha(ws, r, "COMPLEMENTO / COMISSÃO PDV",
          {"AGNOR": f"=ROUND(PARÂMETROS!$B${PROW['Comissão fixa acordada — AGNOR']}-C{rows['COMISSÃO PRODUTOS (INOVAFARMA)']},2)",
           "SARA": 2.65},
          "AGNOR: complemento até a comissão fixa de R$ 2.000,00 (parâmetro na aba PARÂMETROS). SARA: R$ 2,65 de comissão apurada no código dela na loja CENTRO (venda bruta de R$ 365,03), trazida para a folha do Arraial. Demais: preencher com PDV antigos, se houver.",
          kind="in", fill=FILL_CONF)
rows["COMISSÃO TOTAL"] = r
r = linha(ws, r, "= COMISSÃO TOTAL",
          {n: f"=SUM({get_column_letter(C0+i)}{rows['COMISSÃO PRODUTOS (INOVAFARMA)']}:{get_column_letter(C0+i)}{rows['COMPLEMENTO / COMISSÃO PDV']})" for i, n in enumerate(EMP)},
          "Base do DSR junto com as horas extras.", kind="calc", bold=True, fill=FILL_TOT)
rows["DSR"] = r
r = linha(ws, r, "REPOUSO REMUNERADO / DSR",
          {n: f"=ROUND(({get_column_letter(C0+i)}{rows['COMISSÃO TOTAL']}+{get_column_letter(C0+i)}{rows['HORAS EXTRAS']})*{FATOR},2)" for i, n in enumerate(EMP)},
          "= (comissão total + horas extras) × 5 domingos ÷ 26 dias úteis de agosto/2026 (fator na aba PARÂMETROS). A linha HORAS EXTRAS inclui o feriado trabalhado vindo da aba PONTO AGO.26.",
          kind="calc")
rows["AUXÍLIO GERÊNCIA"] = r
r = linha(ws, r, "AUXÍLIO GERÊNCIA", AUXGER, "AGNOR e SARA — valor recorrente de jul/26.", kind="in", fill=FILL_CONF)
rows["META CAIXA"] = r
r = linha(ws, r, "ADICIONAL PRÊMIO META CAIXA", METACX, "VALÉRIA e NATI — valor recorrente de jul/26; confirmar se a meta foi batida em agosto.", kind="in", fill=FILL_CONF)
rows["PRÊMIO COTA GERAL"] = r
r = linha(ws, r, "PRÊMIO COTA GERAL", {}, "PREENCHER conforme apuração de metas de agosto.", kind="in")
rows["PRÊMIO PRÉ-VENCIDOS"] = r
r = linha(ws, r, "PRÊMIO PRÉ-VENCIDOS", {}, "PREENCHER conforme apuração de pré-vencidos de agosto.", kind="in")
rows["FÉRIAS"] = r
r = linha(ws, r, "FÉRIAS (DIAS GOZADOS)", {},
          "JOEL: recibo de férias de 01 a 31/08/2026, JÁ PAGO no início de agosto — NÃO lançar de novo aqui. Pelo acordo interno ele saiu em 04/08 e volta em 04/09.",
          kind="in", fill=FILL_CONF)
rows["1/3 FÉRIAS"] = r
r = linha(ws, r, "1/3 CONSTITUCIONAL DE FÉRIAS", {},
          "JOEL: também já pago com as férias no início de agosto. Linha mantida para eventuais férias de outros meses.",
          kind="in", fill=FILL_CONF)
rows["INC APLIC"] = r
r = linha(ws, r, "INCENTIVO APLICAÇÕES", {n: f"={BASE}!G{BASE_ROW[n]}" for n in EMP},
          "Incentivo de injetáveis apurado no InovaFarma.", kind="link")
rows["INC VIT"] = r
r = linha(ws, r, "INCENTIVO VITAMINAS", {n: f"={BASE}!H{BASE_ROW[n]}" for n in EMP},
          "Grupo APLICAÇÃO E VITAMINAS INCENTIVO no InovaFarma.", kind="link")
rows["INC OUTROS"] = r
r = linha(ws, r, "OUTROS INCENTIVOS", {n: f"={BASE}!I{BASE_ROW[n]}" for n in EMP},
          "Populares, oficinais e similar normal.", kind="link")
prov_ini, prov_fim = rows["SALÁRIO BASE"], rows["INC OUTROS"]
rows["TOTAL PROVENTOS"] = r
skip = rows["COMISSÃO PRODUTOS (INOVAFARMA)"], rows["COMPLEMENTO / COMISSÃO PDV"]
r = linha(ws, r, "TOTAL DE PROVENTOS",
          {n: (f"=SUM({get_column_letter(C0+i)}{prov_ini}:{get_column_letter(C0+i)}{prov_fim})"
               f"-{get_column_letter(C0+i)}{skip[0]}-{get_column_letter(C0+i)}{skip[1]}") for i, n in enumerate(EMP)},
          "Soma das rubricas acima (a linha COMISSÃO TOTAL já engloba comissão InovaFarma + complemento PDV).",
          kind="calc", bold=True, fill=FILL_TOT)

r = sec(ws, r, "DESCONTOS  (lançar com sinal negativo)")
rows["VALES"] = r
r = linha(ws, r, "ADIANTAMENTO / VALES", {}, "PREENCHER com os vales adiantados durante agosto.", kind="in")
rows["VALES INC"] = r
r = linha(ws, r, "ADIANT. VALES INCENT. E APLIC.",
          {n: f"=-({get_column_letter(C0+i)}{rows['INC APLIC']}+{get_column_letter(C0+i)}{rows['INC VIT']}+{get_column_letter(C0+i)}{rows['INC OUTROS']})" for i, n in enumerate(EMP)},
          "Estorno dos incentivos já adiantados em dinheiro no mês (mesmo critério de jul/26). AJUSTAR se algum incentivo não foi adiantado.",
          kind="calc", fill=FILL_CONF)
rows["CONVÊNIO"] = r
r = linha(ws, r, "CONVÊNIO", {}, "PREENCHER (jul/26: AGNOR -300, JOEL -500, VALÉRIA -300, SARA -300).", kind="in")
rows["FALTAS"] = r
r = linha(ws, r, "DESCONTO FALTAS / ATRASOS", {}, "PREENCHER conforme o ponto.", kind="in")
rows["VT6"] = r
r = linha(ws, r, "DESCONTO VALE TRANSPORTE 6%", VT6,
          "Valor praticado em jul/26. 6% sobre R$ 1.621,00 seria R$ 97,26 — CONFERIR a base usada. JOEL: zerado, mês de férias.", kind="in", fill=FILL_CONF)
rows["INSS"] = r
r = linha(ws, r, "DESCONTO INSS", {}, "A CALCULAR PELA CONTABILIDADE sobre o total de proventos (tabela progressiva vigente).", kind="in")
rows["IRRF"] = r
r = linha(ws, r, "DESCONTO IRRF", {}, "A CALCULAR PELA CONTABILIDADE.", kind="in")
rows["TOTAL DESC"] = r
r = linha(ws, r, "TOTAL DE DESCONTOS",
          {n: f"=SUM({get_column_letter(C0+i)}{rows['VALES']}:{get_column_letter(C0+i)}{rows['IRRF']})" for i, n in enumerate(EMP)},
          "", kind="calc", bold=True, fill=FILL_TOT)

r = sec(ws, r, "LÍQUIDO")
rows["LIQ"] = r
r = linha(ws, r, "LÍQUIDO A RECEBER (05/09/2026)",
          {n: f"={get_column_letter(C0+i)}{rows['TOTAL PROVENTOS']}+{get_column_letter(C0+i)}{rows['TOTAL DESC']}" for i, n in enumerate(EMP)},
          "Total de proventos menos descontos. Só fica definitivo depois do INSS/IRRF da contabilidade.",
          kind="calc", bold=True, fill=FILL_LIQ)

r = sec(ws, r, "INFORMATIVO — PAGO PELA EMPRESA, NÃO ENTRA NO HOLERITE")
rows["VT"] = r
r = linha(ws, r, "VALE TRANSPORTE (compra set/26)", {}, "PREENCHER com as passagens compradas para setembro/2026. JOEL: só a partir de 04/09, quando volta das férias.", kind="in")
rows["VA"] = r
r = linha(ws, r, "VALE ALIMENTAÇÃO FERIADOS E DOMINGOS", {}, "PREENCHER conforme escala de domingos/feriados de agosto.", kind="in")

ws.auto_filter.ref = f"A4:{LO}{r-1}"
r += 1
legend = [
 ("LEGENDA", ""),
 ("Célula amarela, texto azul", "valor digitado — PREENCHER ou conferir antes de enviar."),
 ("Célula laranja", "valor repetido de julho/2026 — CONFERIR se continua válido em agosto."),
 ("Texto verde", "valor puxado da aba BASE INOVAFARMA AGO.26 (não digitar por cima)."),
 ("Texto preto em célula verde", "resultado de fórmula — não alterar."),
]
for a, b in legend:
    c = ws.cell(r, 1, a); c.font = font(9, True, "1F3864")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=CO)
    c2 = ws.cell(r, 2, b); c2.font = font(9, it=True)
    r += 1
HOL = "'HOLERITE AGO.26'"

# ======================================================= aba JULHO REVISADO
ws = wb.create_sheet("JULHO.26 REVISADO")
r = build_folha(ws, "FOLHA REVISADA — COMPETÊNCIA JULHO/2026 (paga em 05/08/2026)",
                "Mesmos valores efetivamente pagos, reorganizados e com fórmulas conferidas · confronto com o cálculo correto do DSR",
                None)
jr = {}
r = sec(ws, r, "PROVENTOS")
ordem_prov = ["SALÁRIO BASE", "ADICIONAL NOTURNO", "HORAS EXTRAS", "COMISSÃO PRODUTOS + PDV",
              "REPOUSO REMUNERADO / DSR", "AUXÍLIO GERÊNCIA", "ADICIONAL PRÊMIO META CAIXA",
              "PRÊMIO COTA GERAL", "PRÊMIO PRÉ-VENCIDOS", "INCENTIVO APLICAÇÕES", "INCENTIVO VITAMINAS"]
obs_jul = {"REPOUSO REMUNERADO / DSR": "Pago com fator 5/26 = 0,192307 (calendário de AGOSTO). Julho/2026 tem 4 domingos e 27 dias úteis → fator correto 0,148148. Ver conferência abaixo.",
           "HORAS EXTRAS": "Valor lançado na folha. O DSR foi calculado sobre o dobro deste valor — verificar qual é o valor correto das horas extras.",
           "COMISSÃO PRODUTOS + PDV": "Comissão do InovaFarma somada aos PDV antigos.",
           "SALÁRIO BASE": "CAMILA proporcional (admissão no mês)."}
for rot in ordem_prov:
    jr[rot] = r
    r = linha(ws, r, rot, JUL.get(rot, {}), obs_jul.get(rot, ""), kind="in")
jr["TOTAL PROVENTOS"] = r
r = linha(ws, r, "TOTAL DE PROVENTOS",
          {n: f"=SUM({get_column_letter(C0+i)}{jr[ordem_prov[0]]}:{get_column_letter(C0+i)}{jr[ordem_prov[-1]]})" for i, n in enumerate(EMP)},
          "", kind="calc", bold=True, fill=FILL_TOT)
r = sec(ws, r, "DESCONTOS")
ordem_desc = ["ADIANTAMENTO / VALES", "ADIANT. VALES INCENT. E APLIC.", "CONVÊNIO", "DESCONTO FALTAS",
              "DESCONTO VALE TRANSPORTE 6%", "DESCONTO INSS", "DESCONTO IRRF"]
for rot in ordem_desc:
    jr[rot] = r
    r = linha(ws, r, rot, JUL.get(rot, {}), "", kind="in")
jr["TOTAL DESC"] = r
r = linha(ws, r, "TOTAL DE DESCONTOS",
          {n: f"=SUM({get_column_letter(C0+i)}{jr[ordem_desc[0]]}:{get_column_letter(C0+i)}{jr[ordem_desc[-1]]})" for i, n in enumerate(EMP)},
          "", kind="calc", bold=True, fill=FILL_TOT)
r = sec(ws, r, "LÍQUIDO")
jr["LIQ"] = r
r = linha(ws, r, "LÍQUIDO PAGO (05/08/2026)",
          {n: f"={get_column_letter(C0+i)}{jr['TOTAL PROVENTOS']}+{get_column_letter(C0+i)}{jr['TOTAL DESC']}" for i, n in enumerate(EMP)},
          "", kind="calc", bold=True, fill=FILL_LIQ)
jr["SALDO ORIG"] = r
r = linha(ws, r, "SALDO FINAL DA PLANILHA ORIGINAL", JUL_SALDO,
          "Valor que constava na aba AGOSTOJULHO.26 do arquivo original.", kind="in")
jr["DIF"] = r
r = linha(ws, r, "DIFERENÇA (revisado − original)",
          {n: f"=ROUND({get_column_letter(C0+i)}{jr['LIQ']}-{get_column_letter(C0+i)}{jr['SALDO ORIG']},2)" for i, n in enumerate(EMP)},
          "Deve ser zero — confirma que a revisão não alterou nenhum valor pago.", kind="calc", bold=True, fill=FILL_TOT)

ws.auto_filter.ref = f"A4:{LO}{r-1}"
r = sec(ws, r, "CONFERÊNCIA DO DSR DE JULHO/2026 (não altera o que já foi pago — decidir se ajusta em setembro)")
jr["DSR PAGO"] = r
r = linha(ws, r, "DSR pago (fator 5/26)",
          {n: f"={get_column_letter(C0+i)}{jr['REPOUSO REMUNERADO / DSR']}" for i, n in enumerate(EMP)},
          "Fator do calendário de agosto aplicado por engano na competência de julho.", kind="calc")
jr["DSR CORR"] = r
r = linha(ws, r, "DSR recalculado (fator 4/27 · comissão + HE)",
          {n: (f"=ROUND(({get_column_letter(C0+i)}{jr['COMISSÃO PRODUTOS + PDV']}"
               f"+{get_column_letter(C0+i)}{jr['HORAS EXTRAS']})*4/27,2)") for i, n in enumerate(EMP)},
          "Julho/2026: 4 domingos (05, 12, 19 e 26) e 27 dias úteis.", kind="calc")
jr["DSR DIF"] = r
r = linha(ws, r, "Diferença de DSR (a menor/maior)",
          {n: f"=ROUND({get_column_letter(C0+i)}{jr['DSR CORR']}-{get_column_letter(C0+i)}{jr['DSR PAGO']},2)" for i, n in enumerate(EMP)},
          "Valor negativo = foi pago a mais em julho. Decidir com a contabilidade se compensa em set/26.",
          kind="calc", bold=True, fill=FILL_CONF)

r = sec(ws, r, "INFORMATIVO — PAGO PELA EMPRESA, NÃO ENTRA NO HOLERITE")
r = linha(ws, r, "VALE TRANSPORTE (passagens de agosto)", JUL_VT, "Compra registrada na planilha original.", kind="in")
r = linha(ws, r, "VALE ALIMENTAÇÃO FERIADOS E DOMINGOS", JUL_VA, "", kind="in")
r += 1
for t in ["AJUSTES FEITOS NESTA REVISÃO (a planilha original continua intacta no arquivo enviado):",
          "1. Fórmulas de SALDO FINAL corrigidas — na original o total do DEAN pulava a linha 17 e o do AGNOR começava na linha 4, deixando linhas de fora da soma.",
          "2. Rubricas separadas em PROVENTOS × DESCONTOS, com subtotais próprios, em vez de uma soma única.",
          "3. Nomes de rubrica padronizados e coluna de observação com a origem de cada valor.",
          "4. Conferência do DSR: em julho foi usado o fator de agosto (5/26) e a base foi o dobro das horas extras lançadas.",
          "5. Linha de conferência comparando o líquido revisado com o SALDO FINAL da planilha original (tem que fechar em zero)."]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=CO)
    c = ws.cell(r, 1, t); c.font = font(9, True if t.startswith("AJUSTES") else False, "1F3864" if t.startswith("AJUSTES") else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

# ================================================== aba LISTA (contabilidade)
ws = wb.create_sheet("LISTA CONTABILIDADE")
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [16, 40, 14, 16, 46]):
    ws.column_dimensions[col].width = w
title_block(ws, "LANÇAMENTOS POR FUNCIONÁRIO — AGOSTO/2026 (pagamento 05/09/2026)",
            "Mesmos valores da aba HOLERITE AGO.26, em formato de lista (os valores são vinculados por fórmula)", 5)
r = 4
for i, h in enumerate(["FUNCIONÁRIO", "RUBRICA", "TIPO", "VALOR", "OBSERVAÇÃO"], 1):
    c = ws.cell(r, i, h); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR; c.border = BOX
r = 5
LISTA = [("SALÁRIO BASE", "SALÁRIO BASE", "Provento"),
         ("ADICIONAL NOTURNO", "ADICIONAL NOTURNO", "Provento"),
         ("HORAS EXTRAS", "HORAS EXTRAS", "Provento"),
         ("COMISSÃO TOTAL", "COMISSÃO SOBRE VENDAS", "Provento"),
         ("DSR", "REPOUSO REMUNERADO / DSR", "Provento"),
         ("AUXÍLIO GERÊNCIA", "AUXÍLIO GERÊNCIA", "Provento"),
         ("META CAIXA", "ADICIONAL PRÊMIO META CAIXA", "Provento"),
         ("PRÊMIO COTA GERAL", "PRÊMIO COTA GERAL", "Provento"),
         ("PRÊMIO PRÉ-VENCIDOS", "PRÊMIO PRÉ-VENCIDOS", "Provento"),
         ("INC APLIC", "INCENTIVO APLICAÇÕES", "Provento"),
         ("INC VIT", "INCENTIVO VITAMINAS", "Provento"),
         ("INC OUTROS", "OUTROS INCENTIVOS", "Provento"),
         ("TOTAL PROVENTOS", "TOTAL DE PROVENTOS", "Subtotal"),
         ("VALES", "ADIANTAMENTO / VALES", "Desconto"),
         ("VALES INC", "ADIANT. VALES INCENT. E APLIC.", "Desconto"),
         ("CONVÊNIO", "CONVÊNIO", "Desconto"),
         ("FALTAS", "DESCONTO FALTAS / ATRASOS", "Desconto"),
         ("VT6", "DESCONTO VALE TRANSPORTE 6%", "Desconto"),
         ("INSS", "DESCONTO INSS", "Desconto"),
         ("IRRF", "DESCONTO IRRF", "Desconto"),
         ("TOTAL DESC", "TOTAL DE DESCONTOS", "Subtotal"),
         ("LIQ", "LÍQUIDO A RECEBER", "Líquido")]
for i, n in enumerate(EMP):
    col = get_column_letter(C0 + i)
    for key, rot, tipo in LISTA:
        ws.cell(r, 1, n).font = font(10, True if tipo in ("Líquido",) else False)
        ws.cell(r, 2, rot).font = font(10)
        ws.cell(r, 3, tipo).font = font(10)
        c = ws.cell(r, 4, f"={HOL}!{col}{rows[key]}")
        c.number_format = MONEY; c.font = font(10, tipo in ("Subtotal", "Líquido"), GREEN)
        if tipo == "Líquido":
            for k in range(1, 6): ws.cell(r, k).fill = FILL_LIQ
        elif tipo == "Subtotal":
            for k in range(1, 6): ws.cell(r, k).fill = FILL_TOT
        for k in range(1, 6):
            ws.cell(r, k).border = BOX
        r += 1
    r += 1

ws.auto_filter.ref = f"A4:E{r-2}"
ws.freeze_panes = "A5"

# ==================================================== aba GANHOS DO COLABORADOR
wsg = wb.create_sheet("GANHOS DO COLABORADOR")
wsg.sheet_view.showGridLines = False
for col, w in zip("ABCD", [4, 44, 18, 58]):
    wsg.column_dimensions[col].width = w
title_block(wsg, "DEMONSTRATIVO DE GANHOS DO COLABORADOR",
            "Competência AGOSTO/2026 · Pagamento em 05/09/2026 · Farmácia Tropical Multi Econômica", 4)
wsg.cell(4, 2, "Escolha o colaborador:").font = font(11, True, "1F3864")
sel = wsg.cell(4, 3, EMP[0])
sel.font = font(12, True, BLUE); sel.fill = FILL_IN; sel.border = BOX
sel.alignment = Alignment(horizontal="center")
dv = DataValidation(type="list", formula1='"' + ",".join(EMP) + '"', allow_blank=False)
dv.error = "Escolha um nome da lista."
dv.promptTitle = "Colaborador"
dv.prompt = "Selecione o colaborador para ver os ganhos dele."
wsg.add_data_validation(dv)
dv.add(sel)
wsg.cell(4, 4, "Troque o nome aqui e a tabela abaixo muda sozinha — é o relatório para mandar para cada um.").font = font(9, it=True)

COLREF = f"MATCH($C$4,{HOL}!$B$4:${get_column_letter(CT-1)}$4,0)"
RECIBO = [("SEC", "PROVENTOS", None),
          ("L", "Salário base", rows["SALÁRIO BASE"]),
          ("L", "Adicional noturno", rows["ADICIONAL NOTURNO"]),
          ("L", "Horas extras", rows["HORAS EXTRAS"]),
          ("L", "Comissão sobre vendas", rows["COMISSÃO TOTAL"]),
          ("L", "Repouso remunerado / DSR", rows["DSR"]),
          ("L", "Auxílio gerência", rows["AUXÍLIO GERÊNCIA"]),
          ("L", "Adicional prêmio meta caixa", rows["META CAIXA"]),
          ("L", "Prêmio cota geral", rows["PRÊMIO COTA GERAL"]),
          ("L", "Prêmio pré-vencidos", rows["PRÊMIO PRÉ-VENCIDOS"]),
          ("L", "Férias", rows["FÉRIAS"]),
          ("L", "1/3 constitucional de férias", rows["1/3 FÉRIAS"]),
          ("L", "Incentivo aplicações", rows["INC APLIC"]),
          ("L", "Incentivo vitaminas", rows["INC VIT"]),
          ("L", "Outros incentivos", rows["INC OUTROS"]),
          ("T", "TOTAL DE PROVENTOS", rows["TOTAL PROVENTOS"]),
          ("SEC", "DESCONTOS", None),
          ("L", "Adiantamento / vales", rows["VALES"]),
          ("L", "Adiantamento de incentivos e aplicações", rows["VALES INC"]),
          ("L", "Convênio", rows["CONVÊNIO"]),
          ("L", "Faltas / atrasos", rows["FALTAS"]),
          ("L", "Vale transporte 6%", rows["VT6"]),
          ("L", "INSS", rows["INSS"]),
          ("L", "IRRF", rows["IRRF"]),
          ("T", "TOTAL DE DESCONTOS", rows["TOTAL DESC"]),
          ("SEC", "LÍQUIDO", None),
          ("Q", "LÍQUIDO A RECEBER EM 05/09/2026", rows["LIQ"])]
r = 6
for tipo, rot, linha_hol in RECIBO:
    if tipo == "SEC":
        for i in range(2, 5):
            wsg.cell(r, i).fill = FILL_SEC
        c = wsg.cell(r, 2, rot); c.font = font(11, True, "1F3864")
        wsg.row_dimensions[r].height = 19
        r += 1
        continue
    c = wsg.cell(r, 2, rot)
    c.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q")); c.border = BOX
    v = wsg.cell(r, 3, f"=IFERROR(INDEX({HOL}!$B${linha_hol}:${get_column_letter(CT-1)}${linha_hol},1,{COLREF}),0)")
    v.number_format = MONEY; v.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q"), GREEN)
    v.border = BOX
    if tipo == "T":
        wsg.cell(r, 2).fill = FILL_TOT; v.fill = FILL_TOT
    if tipo == "Q":
        wsg.cell(r, 2).fill = FILL_LIQ; v.fill = FILL_LIQ
        wsg.row_dimensions[r].height = 22
    r += 1
r += 1
wsg.cell(r, 2, "Vendas do colaborador no mês (InovaFarma):").font = font(10, True, "1F3864")
r += 1
for rot, colb in [("Venda bruta", "C"), ("Descontos concedidos", "D"), ("Venda líquida", "E"),
                  ("Comissão apurada", "F"), ("Total de incentivos", "J")]:
    wsg.cell(r, 2, rot).font = font(10); wsg.cell(r, 2).border = BOX
    v = wsg.cell(r, 3, f"=IFERROR(INDEX({BASE}!${colb}${first}:${colb}${last},MATCH($C$4,{BASE}!$B${first}:$B${last},0)),0)")
    v.number_format = MONEY; v.font = font(10, False, GREEN); v.border = BOX
    r += 1
r += 1
for t in ["A comissão apurada acima é a apuração do sistema. O que entra na folha é a linha \"Comissão sobre vendas\" — pode ser diferente (valor fixo acordado, PDV antigos ou quem não recebe comissão).",
          "Os valores de INSS e IRRF são calculados pela contabilidade e só aparecem aqui depois de preenchidos na aba HOLERITE AGO.26.",
          "Para mandar para o colaborador: selecione o nome acima e imprima esta aba em PDF (ou tire um print).",
          "Na aba LISTA CONTABILIDADE também dá para filtrar por FUNCIONÁRIO e ver só os lançamentos de uma pessoa."]:
    wsg.cell(r, 2, t).font = font(9, it=True)
    r += 1
wsg.print_area = f"A1:D{r}"
wsg.page_setup.fitToWidth = 1
wsg.page_setup.fitToHeight = 1
wsg.sheet_properties.pageSetUpPr.fitToPage = True


# ================================================== aba PONTO
PONTO = [  # (funcionario, tipo, qtde, observacao)
 ("EDEY",    "HE50", 24, "24 horas extras normais em agosto/2026."),
 ("EDEY",    "VALOR", 350.0, "Adicional noturno lançado como valor fixo desde jul/26 — CONFERIR com o apontamento do ponto."),
 ("NATI",    "NOT",  17, "17 horas noturnas."),
 ("NATI",    "FER",   1, "Feriado de 15/08/2026 trabalhado, sem folga compensatória — pago em dobro (1 dia a mais)."),
 ("CAMILA",  "FER",   1, "Dia 15/08/2026 (feriado) trabalhado."),
 ("SARA",    "FER",   1, "Dia 15/08/2026 (feriado) trabalhado."),
 ("VALÉRIA", "FER",   1, "Dia 15/08/2026 (feriado) trabalhado."),
 ("VALÉRIA", "NOT", None, "2 madrugadas trabalhadas, com folga normal no dia seguinte — cabe só o adicional noturno. INFORMAR o total de horas noturnas das 2 madrugadas."),
 ("AGNOR",   "NOT", None, "3 madrugadas trabalhadas, com folga normal — cabe só o adicional noturno. INFORMAR o total de horas noturnas das 3 madrugadas."),
 ("DEAN",    "FER_COMP", 1, "Dia 15/08/2026 trabalhado, com folga compensatória — nada a pagar, registro apenas."),
]
TIPOS = {"HE50": ("Horas extras 50%", "horas", "HORAS EXTRAS"),
         "NOT": ("Adicional noturno", "horas", "ADICIONAL NOTURNO"),
         "FER": ("Feriado trabalhado (dobra)", "dias", "HORAS EXTRAS"),
         "FER_COMP": ("Feriado com folga compensatória", "dias", "—"),
         "VALOR": ("Adicional noturno (valor fixo)", "valor", "ADICIONAL NOTURNO")}

wsp = wb.create_sheet("PONTO AGO.26")
wsp.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHI", [14, 32, 10, 10, 14, 14, 14, 22, 62]):
    wsp.column_dimensions[col].width = w
title_block(wsp, "APONTAMENTO DO PONTO — AGOSTO/2026",
            "Horas extras, adicional noturno e feriado trabalhado · alimenta as rubricas da aba HOLERITE AGO.26", 9)
for i, h in enumerate(["FUNCIONÁRIO", "OCORRÊNCIA", "QTDE", "UNIDADE", "SALÁRIO BASE",
                       "VALOR UNITÁRIO", "TOTAL", "RUBRICA DESTINO", "OBSERVAÇÃO"], 1):
    c = wsp.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
wsp.row_dimensions[4].height = 26
SALROW = rows["SALÁRIO BASE"]
P_HORAS = f"PARÂMETROS!$B${PROW['Horas mensais (base do salário-hora)']}"
P_HE = f"PARÂMETROS!$B${PROW['Adicional de hora extra']}"
P_NOT = f"PARÂMETROS!$B${PROW['Adicional noturno']}"
P_DIAS = f"PARÂMETROS!$B${PROW['Dias do mês para o salário-dia']}"
pr = 5
for func, tipo, qtd, obs in PONTO:
    rot, uni, dest = TIPOS[tipo]
    wsp.cell(pr, 1, func).font = font(10, True)
    wsp.cell(pr, 2, rot).font = font(10)
    c = wsp.cell(pr, 3, qtd if tipo != "VALOR" else 1)
    c.font = font(10, False, BLUE); c.fill = FILL_IN
    c.alignment = Alignment(horizontal="center")
    wsp.cell(pr, 4, uni).alignment = Alignment(horizontal="center")
    c = wsp.cell(pr, 5, f"=IFERROR(INDEX({HOL}!$B${SALROW}:${get_column_letter(CT-1)}${SALROW},1,"
                        f"MATCH(A{pr},{HOL}!$B$4:${get_column_letter(CT-1)}$4,0)),0)")
    c.number_format = MONEY; c.font = font(10, False, GREEN)
    if tipo == "HE50":
        f_unit = f"=ROUND(E{pr}/{P_HORAS}*(1+{P_HE}),4)"
    elif tipo == "NOT":
        f_unit = f"=ROUND(E{pr}/{P_HORAS}*{P_NOT},4)"
    elif tipo == "FER":
        f_unit = f"=ROUND(E{pr}/{P_DIAS},2)"
    elif tipo == "FER_COMP":
        f_unit = 0
    else:
        f_unit = qtd
    c = wsp.cell(pr, 6, f_unit); c.number_format = MONEY
    c.font = font(10, False, BLUE if tipo == "VALOR" else "000000")
    if tipo == "VALOR":
        c.fill = FILL_CONF
    c = wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)"); c.number_format = MONEY
    c.font = font(10, True); c.fill = FILL_TOT
    wsp.cell(pr, 8, dest).font = font(9, True, "1F3864")
    c = wsp.cell(pr, 9, obs); c.font = font(9, it=True)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    for i in range(1, 10):
        wsp.cell(pr, i).border = BOX
    pr += 1
P_INI = 5
for extra in range(6):
    for i in range(1, 10):
        cell = wsp.cell(pr, i); cell.border = BOX
        if i in (1, 2, 3, 4, 8, 9):
            cell.font = font(10, False, BLUE); cell.fill = FILL_IN
    wsp.cell(pr, 5, f"=IFERROR(INDEX({HOL}!$B${SALROW}:${get_column_letter(CT-1)}${SALROW},1,"
                    f"MATCH(A{pr},{HOL}!$B$4:${get_column_letter(CT-1)}$4,0)),0)").number_format = MONEY
    wsp.cell(pr, 5).font = font(10, False, GREEN)
    wsp.cell(pr, 6).number_format = MONEY
    wsp.cell(pr, 6).font = font(10, False, BLUE); wsp.cell(pr, 6).fill = FILL_IN
    wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)").number_format = MONEY
    wsp.cell(pr, 7).font = font(10, True); wsp.cell(pr, 7).fill = FILL_TOT
    pr += 1
P_FIM = pr - 1
wsp.cell(pr, 2, "TOTAL").font = font(10, True)
c = wsp.cell(pr, 7, f"=SUM(G{P_INI}:G{P_FIM})"); c.number_format = MONEY; c.font = font(10, True)
for i in range(1, 10):
    wsp.cell(pr, i).fill = FILL_TOT; wsp.cell(pr, i).border = BOX
wsp.auto_filter.ref = f"A4:I{P_FIM}"
wsp.freeze_panes = "A5"
pr += 2
for t in ["Salário-hora = salário base ÷ 220. Hora extra = salário-hora × 1,5. Adicional noturno = salário-hora × 20%. Feriado trabalhado sem folga = 1 salário-dia a mais (salário base ÷ 30).",
          "Confira: 40 horas extras dão R$ 442,09 e 24 horas dão R$ 265,25 no salário de R$ 1.621,00 — é a mesma conta usada na planilha de julho.",
          "As linhas em branco no fim da tabela são para acrescentar outras ocorrências: preencha funcionário, ocorrência, quantidade, valor unitário e a rubrica de destino (HORAS EXTRAS ou ADICIONAL NOTURNO).",
          "Os totais desta aba entram sozinhos nas linhas HORAS EXTRAS e ADICIONAL NOTURNO da aba HOLERITE AGO.26 — não digite nada por cima lá."]:
    wsp.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=9)
    c = wsp.cell(pr, 1, t); c.font = font(9, it=True)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    wsp.row_dimensions[pr].height = 24
    pr += 1

# as rubricas do holerite passam a somar o apontamento
wsh = wb["HOLERITE AGO.26"]
for rub, linha_hol in (("HORAS EXTRAS", rows["HORAS EXTRAS"]), ("ADICIONAL NOTURNO", rows["ADICIONAL NOTURNO"])):
    for i, n in enumerate(EMP):
        c = wsh.cell(linha_hol, C0 + i,
                     f"=SUMIFS('PONTO AGO.26'!$G${P_INI}:$G${P_FIM},'PONTO AGO.26'!$A${P_INI}:$A${P_FIM},"
                     f"{get_column_letter(C0+i)}$4,'PONTO AGO.26'!$H${P_INI}:$H${P_FIM},\"{rub}\")")
        c.number_format = MONEY; c.font = font(10, False, GREEN); c.fill = PatternFill()
    wsh.cell(linha_hol, CO).value = ("Somado automaticamente da aba PONTO AGO.26 (apontamento de agosto/2026). "
                                     "Para mudar, edite lá.")

# ============================================================== aba CAPA
ws = wb.create_sheet("CAPA", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 118
title_block(ws, "RELATÓRIO DE FOLHA PARA A CONTABILIDADE", "", 2)
ws["A2"] = ""
blocos = [
 ("SEC", "IDENTIFICAÇÃO"),
 ("T", "Empresa: Farmácia Tropical Multi Econômica"),
 ("T", "Competência (mês de referência): AGOSTO/2026 — vendas de 01/08/2026 a 31/08/2026"),
 ("T", "Data do pagamento: 05/09/2026"),
 ("T", "Fonte das comissões e incentivos: InovaFarma — Relatório detalhado de comissão de vendedor, extraído em 03/09/2026"),
 ("T", "Funcionários: DEAN, AGNOR, EDEY, JOEL, VALÉRIA, SARA, CAMILA e NATI"),
 ("SEC", "COMO USAR ESTE ARQUIVO"),
 ("T", "1. Abra a aba HOLERITE AGO.26 — é o relatório que vai para a contabilidade."),
 ("T", "2. Preencha as células AMARELAS (horas extras, prêmios, vales, convênio, faltas)."),
 ("T", "3. Confira as células LARANJAS — são valores repetidos de julho/2026 que podem ter mudado."),
 ("T", "4. Comissões, incentivos e DSR já vêm calculados; não digite por cima (texto verde/preto)."),
 ("T", "5. A aba LISTA CONTABILIDADE traz os mesmos valores em formato de lista, rubrica por rubrica."),
 ("T", "6. INSS e IRRF ficam em branco: são calculados pela contabilidade."),
 ("SEC", "O QUE JÁ ESTÁ APURADO (AGOSTO/2026)"),
 ("T", "Comissão total apurada no InovaFarma para os 8 funcionários: R$ 4.856,68 · Incentivos: R$ 685,00"),
 ("T", "Venda bruta geral do mês: R$ 300.371,95 · Descontos concedidos: R$ 90.276,65 · Venda líquida: R$ 210.095,30"),
 ("T", "DSR de agosto/2026: 5 domingos (02, 09, 16, 23 e 30) ÷ 26 dias úteis = fator 0,192307"),
 ("SEC", "APONTAMENTO DE AGOSTO/2026 (aba PONTO AGO.26)"),
 ("T", "EDEY: 24 horas extras normais → R$ 265,25."),
 ("T", "NATI: 17 horas noturnas → R$ 25,05 de adicional; e o feriado de 15/08 trabalhado sem folga → R$ 54,03."),
 ("T", "CAMILA, SARA e VALÉRIA: dia 15/08 (feriado) trabalhado → R$ 54,03 cada."),
 ("T", "VALÉRIA: 2 madrugadas com folga no dia seguinte — cabe só o adicional noturno; falta informar as horas."),
 ("T", "AGNOR: 3 madrugadas com folga — cabe só o adicional noturno; falta informar as horas."),
 ("T", "DEAN: dia 15/08 trabalhado com folga compensatória — nada a pagar."),
 ("SEC", "JÁ DEFINIDO PELA EMPRESA"),
 ("T", "AGNOR: mantida a comissão fixa de R$ 2.000,00 — a linha COMPLEMENTO / COMISSÃO PDV calcula sozinha a diferença (R$ 1.090,49) sobre o apurado de R$ 909,51."),
 ("T", "JOEL: por acordo interno saiu em 04/08 e volta em 04/09, mas o recibo de férias saiu como 01 a 31/08 e já foi pago no início de agosto. Por isso ele tem comissão dos dias 01 e 02/08 (R$ 103,31) e fica sem salário e sem desconto de vale-transporte neste holerite; férias e 1/3 NÃO se repetem aqui."),
 ("T", "CAMILA: contratada recentemente; agosto é o primeiro mês completo, com salário integral de R$ 1.621,00."),
 ("T", "SARA: a comissão dela na loja Centro (R$ 2,65) foi somada aqui, na linha COMPLEMENTO / COMISSÃO PDV."),
 ("T", "DEAN: não recebe comissão sobre vendas — a rubrica fica zerada na folha (a apurada, R$ 975,65, aparece só na aba BASE). Os incentivos dele continuam sendo pagos normalmente."),
 ("SEC", "PENDÊNCIAS — CONFIRMAR ANTES DE ENVIAR"),
 ("P", "VALÉRIA e AGNOR: informar na aba PONTO AGO.26 quantas horas noturnas tiveram as madrugadas trabalhadas (2 e 3 madrugadas)."),
 ("P", "EDEY: o adicional noturno está lançado como valor fixo de R$ 350,00 — conferir com o apontamento do ponto."),
 ("P", "Confirmar o critério do feriado trabalhado: está sendo pago 1 salário-dia a mais (R$ 54,03) por feriado sem folga."),
 ("P", "Prêmio cota geral e prêmio pré-vencidos de agosto."),
 ("P", "Vales adiantados, convênio e faltas de agosto."),
 ("P", "JOEL: confirmar se há algum desconto para agosto (convênio, vale ou adiantamento) — hoje o holerite dele fica só com a comissão de R$ 103,31 e o DSR."),
 ("P", "JOEL: como o retorno é em 04/09, o vale-transporte de setembro deve ser proporcional e o holerite de setembro ainda pega os dias 01 a 03/09 de férias, já pagos no recibo de agosto."),
 ("P", "Comissões de PDV antigos, que em julho eram somadas à comissão do sistema."),
 ("P", "Desconto de vale-transporte: vem sendo lançado R$ 84,29; 6% do salário de R$ 1.621,00 seria R$ 97,26."),
 ("P", "DSR de julho/2026 foi calculado com o fator de agosto (5/26) em vez de 4/27 — ver aba JULHO.26 REVISADO e decidir se ajusta em setembro."),
 ("SEC", "OBSERVAÇÃO TÉCNICA"),
 ("T", "As células de total e de cálculo são fórmulas. Ao abrir o arquivo no Excel ou no Google Planilhas os valores aparecem calculados automaticamente; em visualizadores simples (prévia de celular, por exemplo) elas podem aparecer em branco até o arquivo ser aberto de fato."),
 ("SEC", "ABAS DO ARQUIVO"),
 ("T", "HOLERITE AGO.26 — relatório principal da competência agosto/2026."),
 ("T", "PONTO AGO.26 — apontamento de horas extras, adicional noturno e feriado trabalhado; alimenta o holerite."),
 ("T", "GANHOS DO COLABORADOR — escolha o nome na lista e saia o demonstrativo individual, pronto para imprimir/mandar."),
 ("T", "LISTA CONTABILIDADE — os mesmos lançamentos em formato de lista por funcionário."),
 ("T", "BASE INOVAFARMA AGO.26 — apuração de comissões e incentivos por vendedor."),
 ("T", "JULHO.26 REVISADO — a folha de julho reorganizada e conferida (valores pagos preservados)."),
 ("T", "PARÂMETROS — calendário do mês, fator do DSR e valores fixos."),
]
r = 4
for tipo, txt in blocos:
    if tipo == "SEC":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(r, 1, txt); c.font = font(11, True, "1F3864"); c.fill = FILL_SEC
        ws.row_dimensions[r].height = 20
    else:
        ws.cell(r, 1, "•" if tipo == "P" else "").font = font(10, True, "C00000")
        c = ws.cell(r, 2, txt); c.font = font(10)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if tipo == "P":
            c.fill = FILL_ALERT
        ws.row_dimensions[r].height = 17
    r += 1
r += 1
ws.cell(r, 2, "Documento gerado a partir da planilha CONISSÕES PLAN.AGOSTO/JULHO ARRAIAL 2026 e do relatório de comissões do InovaFarma.").font = font(9, it=True)

wb.save(OUT)
print("ok", OUT)
