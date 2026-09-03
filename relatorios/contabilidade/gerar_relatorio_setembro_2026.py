# -*- coding: utf-8 -*-
"""Relatorio de folha (holerite) AGOSTO/2026 - lojas CENTRO e TRANCOSO.
Pagamento em 05/09/2026. Mesma estrutura usada na loja ARRAIAL."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = "/home/user/site-farmacia/relatorios/contabilidade/"
F = "Arial"
MONEY = 'R$ #,##0.00;-R$ #,##0.00;"-"'

def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)

FILL_TIT   = PatternFill("solid", fgColor="1F3864")
FILL_SEC   = PatternFill("solid", fgColor="D9E2F3")
FILL_HDR   = PatternFill("solid", fgColor="2E5496")
FILL_IN    = PatternFill("solid", fgColor="FFF2CC")
FILL_CONF  = PatternFill("solid", fgColor="FCE4D6")
FILL_TOT   = PatternFill("solid", fgColor="E2EFDA")
FILL_LIQ   = PatternFill("solid", fgColor="C6E0B4")
FILL_ALERT = PatternFill("solid", fgColor="FFF0F0")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BLUE, GREEN = "0000FF", "008000"

# ------------------------------------------------------------------ CONFIG
# Setembro/2026 ainda não terminou: comissões e incentivos entram depois,
# com o relatório do InovaFarma extraído no início de outubro/2026.
ZERO = (0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0)

ARRAIAL = dict(
 loja="ARRAIAL", arquivo="RELATORIO_HOLERITE_SETEMBRO_2026_ARRAIAL_pgto_05-10-2026.xlsx",
 mult=1.0,
 EMP=["DEAN", "AGNOR", "EDEY", "JOEL", "VALÉRIA", "SARA", "CAMILA", "NATI"],
 INOVA={n: (cod,) + ZERO[1:] for n, cod in
        [("DEAN", 242), ("AGNOR", 53), ("EDEY", 272), ("JOEL", 181),
         ("VALÉRIA", 292), ("SARA", 43), ("CAMILA", 312), ("NATI", 212)]},
 OUTROS_COD=[],
 OUTROS_NOTA="Conferir também o código 43 (SARA) no relatório da loja Centro: o saldo dela vem para esta folha.",
 GERAL=dict(bruta=0.0, desc=0.0, liq=0.0, com=0.0, inc=0.0),
 SALARIO={"DEAN": 5648.41, "AGNOR": 1621.0, "EDEY": 1621.0, "JOEL": 1458.90,
          "VALÉRIA": 1621.0, "SARA": 1621.0, "CAMILA": 1621.0, "NATI": 1621.0},
 NOTURNO={}, AUXGER={"AGNOR": 810.5, "SARA": 810.5},
 METACX={"VALÉRIA": 307.99, "NATI": 307.99},
 VT6={"EDEY": -84.29, "JOEL": -84.29, "VALÉRIA": -84.29, "NATI": -84.29},
 SEM_COMISSAO=["DEAN"],
 COM_FIXA={"AGNOR": 2000.00},
 PONTO=[("EDEY", "VALOR", 350.0, "Adicional noturno lançado como valor fixo desde jul/26 — CONFERIR com o apontamento do ponto."),
        ("JOEL", "INFO", 3, "Férias de 01 a 03/09/2026 (fim do período iniciado em 04/08), já pagas no recibo de agosto — nada a pagar aqui. Ele voltou ao trabalho em 04/09.")],
 DEFINIDO=["DEAN: não recebe comissão sobre vendas — a rubrica fica zerada.",
           "AGNOR: comissão fixa de R$ 2.000,00 — o complemento sobre o apurado é calculado sozinho.",
           "JOEL: salário proporcional a 27 dias (R$ 1.458,90), porque as férias foram até 03/09 e ele voltou em 04/09.",
           "SARA: o saldo dela na loja Centro deve ser somado aqui, na linha COMPLEMENTO / COMISSÃO PDV."],
 PENDENCIAS=["Extrair o relatório de comissões de setembro no InovaFarma e preencher a aba BASE INOVAFARMA SET.26.",
             "JOEL: conferir o salário proporcional e lançar o vale-transporte de outubro cheio (ele já voltou).",
             "CAMILA: confirmar se segue com salário integral."],
)

CENTRO = dict(
 loja="CENTRO", arquivo="RELATORIO_HOLERITE_SETEMBRO_2026_CENTRO_pgto_05-10-2026.xlsx",
 mult=1.0,
 EMP=["ARIANE", "ELIANA", "GENECIR", "RENALDO", "THAYANE", "PEDRO"],
 INOVA={n: (cod,) + ZERO[1:] for n, cod in
        [("ARIANE", 211), ("ELIANA", 81), ("GENECIR", 241), ("RENALDO", 151),
         ("THAYANE", 61), ("PEDRO", 41)]},
 OUTROS_COD=[("11", "GILSON MOURA SANTOS", 0.0, 0.0, 0.0),
             ("13", "UILLIAN SILVA SANTOS", 0.0, 0.0, 0.0),
             ("31", "FABIANA RODRIGUES", 0.0, 0.0, 0.0),
             ("43", "SARA", 0.0, 0.0, 0.0)],
 OUTROS_NOTA="GILSON, UILLIAN e FABIANA não recebem comissão. O saldo da SARA vai para a folha do Arraial.",
 FOLGUISTAS=[("SERGIO", "51", 150.00, 0.0, 0.0, 0.0),
             ("ANA CELIA", "131", 100.00, 0.0, 0.0, 0.0)],
 GERAL=dict(bruta=0.0, desc=0.0, liq=0.0, com=0.0, inc=0.0),
 SALARIO={"ARIANE": 5648.41, "ELIANA": 1621.0, "GENECIR": 1621.0,
          "RENALDO": 1621.0, "THAYANE": 1621.0, "PEDRO": 1621.0},
 NOTURNO={}, AUXGER={}, METACX={},
 VT6={"ELIANA": -84.29, "GENECIR": -84.29, "RENALDO": -74.29, "THAYANE": -84.29},
 SEM_COMISSAO=["ARIANE", "PEDRO"],
 PONTO=[("GENECIR", "INFO", 0, "Voltou das férias (01 a 30/08/2026) — salário integral em setembro.")],
 DEFINIDO=["ARIANE e PEDRO: não recebem comissão sobre vendas.",
           "GENECIR: voltou das férias, salário integral em setembro.",
           "SERGIO e ANA CELIA seguem como folguistas (diária de R$ 150,00 e R$ 100,00), na aba FOLGUISTAS."],
 PENDENCIAS=["Extrair o relatório de comissões de setembro no InovaFarma e preencher a aba BASE INOVAFARMA SET.26.",
             "Preencher a quantidade de diárias de SERGIO e ANA CELIA."],
)

TRANCOSO = dict(
 loja="TRANCOSO", arquivo="RELATORIO_HOLERITE_SETEMBRO_2026_TRANCOSO_pgto_05-10-2026.xlsx",
 mult=2.0,
 EMP=["UILLIAN", "MANOEL", "VALDICK", "INIURLE", "TAMILES"],
 INOVA={n: (cod,) + ZERO[1:] for n, cod in
        [("UILLIAN", 13), ("MANOEL", 92), ("VALDICK", 162), ("INIURLE", 83), ("TAMILES", 103)]},
 OUTROS_COD=[], OUTROS_NOTA="",
 GERAL=dict(bruta=0.0, desc=0.0, liq=0.0, com=0.0, inc=0.0),
 SALARIO={"UILLIAN": 5648.41, "MANOEL": 1621.0, "VALDICK": 1621.0,
          "INIURLE": 1621.0, "TAMILES": 1621.0},
 NOTURNO={}, AUXGER={}, METACX={}, VT6={},
 SEM_COMISSAO=["UILLIAN"],
 PONTO=[("TAMILES", "ATESTADO", 7, "Atestado médico de 02 a 08/09/2026, entregue em 02/09. Até 15 dias é abonado pela empresa: falta justificada, SEM desconto de salário nem de DSR. Anexar o atestado à pasta da funcionária.")],
 DEFINIDO=["TRANCOSO PAGA O DOBRO DA COMISSÃO APURADA: a linha COMISSÃO PRODUTOS já multiplica por 2 o valor do InovaFarma.",
           "Os incentivos são simples: entram pelo valor apurado, sem dobrar.",
           "A diferença gerada pela comissão dobrada é reduzida nas premiações.",
           "UILLIAN: não recebe comissão sobre vendas.",
           "TAMILES: atestado de 02 a 08/09 abonado — salário integral, sem desconto."],
 PENDENCIAS=["Extrair o relatório de comissões de setembro no InovaFarma e preencher a aba BASE INOVAFARMA SET.26.",
             "TAMILES: se o atestado for prorrogado além de 15 dias corridos, a partir do 16º dia passa a ser auxílio-doença do INSS.",
             "Lançar na linha PRÊMIO COTA GERAL o valor já com a redução combinada."],
)

# ------------------------------------------------------------------ BUILDER
def build(cfg):
    EMP = cfg["EMP"]; NCOL = len(EMP)
    C0 = 2; CT = C0 + NCOL; CO = CT + 1
    LO = get_column_letter(CO)
    INOVA = cfg["INOVA"]
    wb = openpyxl.Workbook()

    def title_block(ws, text, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text); c.font = font(14, True, "FFFFFF"); c.fill = FILL_TIT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(2, 1, sub); c.font = font(10, True, "1F3864")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

    # ---------------------------------------------------------- PARÂMETROS
    ws = wb.active; ws.title = "PARÂMETROS"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABC", [46, 18, 74]):
        ws.column_dimensions[col].width = w
    title_block(ws, "PARÂMETROS DE CÁLCULO",
                f"Farmácia Tropical Multi Econômica — loja {cfg['loja']} — competência SETEMBRO/2026", 3)
    P = [("SEÇÃO", "COMPETÊNCIA / PAGAMENTO", ""),
         ("Competência (mês de referência)", "SETEMBRO/2026", "Vendas de 01/09/2026 a 30/09/2026 (relatório InovaFarma)."),
         ("Data do pagamento", "05/10/2026", "Folha a ser lançada no holerite pago em 05/10/2026."),
         ("Data de extração do InovaFarma", "a extrair", f"Extrair o relatório da loja {cfg['loja']} no início de outubro/2026 e lançar na aba BASE INOVAFARMA SET.26."),
         ("SEÇÃO", "CALENDÁRIO DO MÊS (base do DSR)", ""),
         ("Dias do mês", 30, "Setembro/2026."),
         ("Domingos + feriados", 5, "Domingos: 06, 13, 20 e 27/09/2026, mais o feriado de 07/09 (segunda-feira). CONFERIR feriados municipais."),
         ("Dias úteis (inclui sábados)", None, "Dias do mês menos domingos e feriados."),
         ("Fator DSR (domingos ÷ dias úteis)", None, "Domingos ÷ dias úteis. Usado na rubrica REPOUSO REMUNERADO / DSR."),
         ("SEÇÃO", "VALORES FIXOS / RECORRENTES", ""),
         ("Salário base (piso 2026)", 1621.00, "Valor praticado em ago/2026."),
         ("Salário do gerente", 5648.41, "Conforme ago/2026."),
         ("Horas mensais (base do salário-hora)", 220, "Salário-hora = salário base ÷ 220."),
         ("Adicional de hora extra", 0.5, "50% sobre a hora normal (hora extra = salário-hora × 1,5)."),
         ("Adicional noturno", 0.2, "20% sobre a hora normal, nas horas entre 22h e 5h."),
         ("Dias do mês para o salário-dia", 30, "Salário-dia = salário base ÷ 30. Usado no feriado trabalhado."),
         ("Multiplicador da comissão", cfg["mult"],
          "TRANCOSO paga o dobro do valor apurado no InovaFarma." if cfg["mult"] != 1
          else "1 = paga o valor apurado no InovaFarma, sem multiplicador."),
         ("SEÇÃO", "INSS / IRRF", ""),
         ("INSS", "a calcular", "Calculado pela contabilidade sobre o total de proventos (tabela progressiva vigente)."),
         ("IRRF", "a calcular", "Calculado pela contabilidade após dedução do INSS e dependentes.")]
    r = 4; PROW = {}
    for a, b, c in P:
        if a == "SEÇÃO":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            cc = ws.cell(r, 1, b); cc.font = font(11, True, "1F3864"); cc.fill = FILL_SEC
            ws.row_dimensions[r].height = 20
        else:
            ws.cell(r, 1, a).font = font(10, True)
            cc = ws.cell(r, 2, b); cc.font = font(10, False, BLUE); cc.fill = FILL_IN
            cc.alignment = Alignment(horizontal="center"); cc.border = BOX
            if isinstance(b, float):
                cc.number_format = MONEY if b > 100 else "0.00"
            PROW[a] = r
            ws.cell(r, 3, c).font = font(9, it=True)
            ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    r_dias = PROW["Dias do mês"]; r_dom = PROW["Domingos + feriados"]
    r_ute = PROW["Dias úteis (inclui sábados)"]; r_fat = PROW["Fator DSR (domingos ÷ dias úteis)"]
    ws.cell(r_ute, 2).value = f"=B{r_dias}-B{r_dom}"
    cf = ws.cell(r_fat, 2, f"=B{r_dom}/B{r_ute}")
    cf.number_format = "0.000000"; cf.font = font(10, True); cf.fill = FILL_TOT
    cf.alignment = Alignment(horizontal="center"); cf.border = BOX
    FATOR = f"PARÂMETROS!$B${r_fat}"
    MULT = f"PARÂMETROS!$B${PROW['Multiplicador da comissão']}"

    # ---------------------------------------------------------- BASE INOVA
    ws = wb.create_sheet("BASE INOVAFARMA SET.26")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([8, 26, 16, 16, 16, 15, 13, 13, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    title_block(ws, "APURAÇÃO DE COMISSÕES E INCENTIVOS — INOVAFARMA",
                f"Loja {cfg['loja']} · vendas de 01/09/2026 a 30/09/2026 · PREENCHER com o relatório extraído no início de outubro", 10)
    hdr = ["CÓD.", "VENDEDOR", "VENDA BRUTA", "DESCONTOS", "VENDA LÍQUIDA", "COMISSÃO",
           "INCENT. APLIC.", "INCENT. VITAM.", "OUTROS INCENT.", "TOTAL INCENT."]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
    ws.row_dimensions[4].height = 28
    r = 5; first = r
    for n in EMP:
        cod, vb, ds, vl, com, inj, vit, out = INOVA[n]
        ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, n).font = font(10, True)
        for i, v in enumerate([vb, ds, vl, com, inj, vit, out], 3):
            c = ws.cell(r, i, v); c.number_format = MONEY; c.font = font(10, False, BLUE)
        ws.cell(r, 10, f"=SUM(G{r}:I{r})").number_format = MONEY
        ws.cell(r, 10).font = font(10, True)
        for i in range(1, 11):
            ws.cell(r, i).border = BOX
        r += 1
    last = r - 1
    ws.cell(r, 2, "TOTAL FUNCIONÁRIOS").font = font(10, True)
    for i in range(3, 11):
        L = get_column_letter(i)
        c = ws.cell(r, i, f"=SUM({L}{first}:{L}{last})")
        c.number_format = MONEY; c.font = font(10, True)
    for i in range(1, 11):
        ws.cell(r, i).fill = FILL_TOT; ws.cell(r, i).border = BOX
    ws.auto_filter.ref = f"A4:J{last}"
    ws.freeze_panes = "A5"
    r += 2
    ws.cell(r, 1, "Códigos sem vínculo de folha nesta loja:").font = font(9, True, "1F3864")
    if cfg.get("OUTROS_NOTA"):
        ws.cell(r, 3, cfg["OUTROS_NOTA"]).font = font(9, it=True)
    r += 1
    for cod, nome, vb, com, inc in cfg["OUTROS_COD"]:
        ws.cell(r, 1, cod).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, nome).font = font(9)
        ws.cell(r, 3, vb).number_format = MONEY; ws.cell(r, 3).font = font(9)
        ws.cell(r, 6, com).number_format = MONEY; ws.cell(r, 6).font = font(9)
        ws.cell(r, 10, inc).number_format = MONEY; ws.cell(r, 10).font = font(9)
        r += 1
    r += 1
    g = cfg["GERAL"]
    notas = [f"CONFERÊNCIA — total geral do relatório: venda bruta R$ {g['bruta']:,.2f} · descontos R$ {g['desc']:,.2f} · "
             f"venda líquida R$ {g['liq']:,.2f} · comissão R$ {g['com']:,.2f} · incentivo R$ {g['inc']:,.2f}."
             .replace(",", "X").replace(".", ",").replace("X", "."),
             "INCENT. APLIC. = incentivo do grupo INJETÁVEIS (aplicações). INCENT. VITAM. = grupo APLICAÇÃO E VITAMINAS INCENTIVO. OUTROS INCENT. = populares, oficinais e similar normal.",
             "PREENCHER esta aba com o relatório de comissões de setembro/2026, extraído do InovaFarma no início de outubro. Enquanto estiver zerada, o holerite fica só com salário e rubricas fixas.",
             "A comissão desta aba é a APURADA pelo sistema, sem multiplicador. O valor que vai para a folha está na aba HOLERITE SET.26."]
    for n in cfg["SEM_COMISSAO"]:
        notas.append(f"{n} não recebe comissão em folha: no caso dele(a) a comissão apurada acima é apenas informativa.")
    if cfg["mult"] != 1:
        notas.append("ATENÇÃO: nesta loja a comissão paga é o DOBRO do valor apurado acima.")
    for t in notas:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(r, 1, t); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1
    BASE = "'BASE INOVAFARMA SET.26'"
    BROW = {n: first + i for i, n in enumerate(EMP)}

    # ---------------------------------------------------------- helpers folha
    def cabecalho(ws, titulo, sub):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 38
        for i in range(C0, CT + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13.5
        ws.column_dimensions[LO].width = 62
        ws.freeze_panes = "B5"
        title_block(ws, titulo, sub, CO)
        ws.cell(4, 1, "RUBRICA").font = font(10, True, "FFFFFF")
        ws.cell(4, 1).fill = FILL_HDR
        for i, n in enumerate(EMP):
            c = ws.cell(4, C0 + i, n); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(4, CT, "TOTAL"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(4, CO, "ORIGEM / OBSERVAÇÃO"); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR
        for i in range(1, CO + 1):
            ws.cell(4, i).border = BOX
        ws.row_dimensions[4].height = 20
        return 5

    def sec(ws, r, texto):
        for i in range(1, CO + 1):
            ws.cell(r, i).fill = FILL_SEC
        c = ws.cell(r, 1, texto); c.font = font(11, True, "1F3864")
        ws.row_dimensions[r].height = 19
        return r + 1

    def linha(ws, r, rotulo, valores, obs="", kind="in", bold=False, fill=None):
        c = ws.cell(r, 1, rotulo); c.font = font(10, bold); c.border = BOX
        for i, n in enumerate(EMP):
            cell = ws.cell(r, C0 + i)
            v = valores.get(n)
            if v is not None:
                cell.value = v
            cell.number_format = MONEY; cell.border = BOX
            if kind == "in":
                cell.font = font(10, bold, BLUE); cell.fill = fill or FILL_IN
            elif kind == "link":
                cell.font = font(10, bold, GREEN)
                if fill: cell.fill = fill
            else:
                cell.font = font(10, bold)
                if fill: cell.fill = fill
        t = ws.cell(r, CT, f"=SUM({get_column_letter(C0)}{r}:{get_column_letter(CT-1)}{r})")
        t.number_format = MONEY; t.font = font(10, True); t.border = BOX
        t.fill = fill or FILL_TOT
        o = ws.cell(r, CO, obs); o.font = font(9, it=True); o.border = BOX
        o.alignment = Alignment(wrap_text=True, vertical="center")
        return r + 1

    # ---------------------------------------------------------- HOLERITE AGO
    ws = wb.create_sheet("HOLERITE SET.26")
    r = cabecalho(ws, "RELATÓRIO PARA A CONTABILIDADE — LANÇAMENTO EM HOLERITE",
                  f"Loja {cfg['loja']} · competência SETEMBRO/2026 (01/08 a 31/08/2026) · pagamento em 05/09/2026")
    rows = {}
    sem = cfg["SEM_COMISSAO"]
    r = sec(ws, r, "PROVENTOS")
    rows["SALÁRIO BASE"] = r
    sal = dict(cfg["SALARIO"])
    obs_sal = "Piso 2026 R$ 1.621,00; gerente conforme ago/2026. CONFERIR admissões, afastamentos e férias do mês."
    for n, motivo in cfg.get("SALARIO_ZERO", {}).items():
        sal[n] = 0.0
        obs_sal += f" {n}: zerado — {motivo}."
    r = linha(ws, r, "SALÁRIO BASE", sal, obs_sal, kind="in", fill=FILL_CONF)
    rows["ADICIONAL NOTURNO"] = r
    r = linha(ws, r, "ADICIONAL NOTURNO", cfg["NOTURNO"], "Conforme apontamento do ponto.", kind="in")
    rows["HORAS EXTRAS"] = r
    r = linha(ws, r, "HORAS EXTRAS", {}, "PREENCHER com o apurado no ponto de agosto/2026.", kind="in")
    rows["COM INOVA"] = r
    obs_com = "Apurado no InovaFarma (aba BASE INOVAFARMA SET.26)"
    if cfg["mult"] != 1:
        obs_com += ", JÁ MULTIPLICADO POR 2 (multiplicador na aba PARÂMETROS)"
    if sem:
        obs_com += ". " + " e ".join(sem) + " não recebe(m) comissão — lançado zero"
    r = linha(ws, r, "COMISSÃO PRODUTOS (INOVAFARMA)",
              {n: (0 if n in sem else f"=ROUND({BASE}!F{BROW[n]}*{MULT},2)") for n in EMP},
              obs_com + ".", kind="link")
    rows["COM PDV"] = r
    fixa = cfg.get("COM_FIXA", {})
    obs_pdv = "PREENCHER: PDV antigos e complementos de comissão acordados, se houver."
    vals_pdv = {}
    for n, v in fixa.items():
        i = EMP.index(n)
        vals_pdv[n] = (f"=ROUND(MAX(0,{v}-{get_column_letter(C0+i)}{rows['COM INOVA']}),2)")
        obs_pdv += f" {n}: complemento até a comissão fixa de R$ {v:,.2f}.".replace(",", "X").replace(".", ",").replace("X", ".")
    r = linha(ws, r, "COMPLEMENTO / COMISSÃO PDV", vals_pdv, obs_pdv,
              kind="in", fill=FILL_CONF if fixa else None)
    rows["COMISSÃO TOTAL"] = r
    r = linha(ws, r, "= COMISSÃO TOTAL",
              {n: f"=SUM({get_column_letter(C0+i)}{rows['COM INOVA']}:{get_column_letter(C0+i)}{rows['COM PDV']})"
               for i, n in enumerate(EMP)},
              "Base do DSR junto com as horas extras.", kind="calc", bold=True, fill=FILL_TOT)
    rows["DSR"] = r
    r = linha(ws, r, "REPOUSO REMUNERADO / DSR",
              {n: (f"=ROUND(({get_column_letter(C0+i)}{rows['COMISSÃO TOTAL']}"
                   f"+{get_column_letter(C0+i)}{rows['HORAS EXTRAS']})*{FATOR},2)") for i, n in enumerate(EMP)},
              "= (comissão total + horas extras) × (4 domingos + feriado 07/09) ÷ 25 dias úteis de setembro/2026.", kind="calc")
    rows["AUXÍLIO GERÊNCIA"] = r
    r = linha(ws, r, "AUXÍLIO GERÊNCIA", cfg["AUXGER"], "Conforme acordo da loja.", kind="in")
    rows["META CAIXA"] = r
    r = linha(ws, r, "ADICIONAL PRÊMIO META CAIXA", cfg["METACX"], "Conforme apuração do caixa.", kind="in")
    rows["PRÊMIO COTA GERAL"] = r
    obs_cota = "PREENCHER conforme a tabela de metas da loja (aba METAS da planilha original)."
    if cfg["mult"] != 1:
        obs_cota += " Lançar já com a redução combinada para compensar o aumento da comissão dobrada."
    r = linha(ws, r, "PRÊMIO COTA GERAL", {}, obs_cota, kind="in", fill=FILL_CONF if cfg["mult"] != 1 else None)
    rows["PRÊMIO PRÉ-VENCIDOS"] = r
    r = linha(ws, r, "PRÊMIO PRÉ-VENCIDOS", {}, "PREENCHER conforme a apuração de pré-vencidos de setembro.", kind="in")
    rows["FÉRIAS"] = r
    r = linha(ws, r, "FÉRIAS (DIAS GOZADOS)", {}, "PREENCHER se houve férias no mês; informar o período.", kind="in")
    rows["1/3 FÉRIAS"] = r
    r = linha(ws, r, "1/3 CONSTITUCIONAL DE FÉRIAS", {}, "Calculado pela contabilidade sobre o valor das férias.", kind="in")
    rows["INC APLIC"] = r
    inc_obs = "Incentivo de injetáveis apurado no InovaFarma."
    if cfg["mult"] != 1:
        inc_obs += " Incentivo é simples: NÃO entra em dobro."
    r = linha(ws, r, "INCENTIVO APLICAÇÕES", {n: f"={BASE}!G{BROW[n]}" for n in EMP}, inc_obs, kind="link")
    rows["INC VIT"] = r
    r = linha(ws, r, "INCENTIVO VITAMINAS", {n: f"={BASE}!H{BROW[n]}" for n in EMP},
              "Grupo APLICAÇÃO E VITAMINAS INCENTIVO no InovaFarma.", kind="link")
    rows["INC OUTROS"] = r
    r = linha(ws, r, "OUTROS INCENTIVOS", {n: f"={BASE}!I{BROW[n]}" for n in EMP},
              "Populares, oficinais e similar normal.", kind="link")
    rows["TOTAL PROVENTOS"] = r
    r = linha(ws, r, "TOTAL DE PROVENTOS",
              {n: (f"=SUM({get_column_letter(C0+i)}{rows['SALÁRIO BASE']}:{get_column_letter(C0+i)}{rows['INC OUTROS']})"
                   f"-{get_column_letter(C0+i)}{rows['COM INOVA']}-{get_column_letter(C0+i)}{rows['COM PDV']}")
               for i, n in enumerate(EMP)},
              "Soma das rubricas acima (a linha COMISSÃO TOTAL já engloba a comissão do InovaFarma e o complemento).",
              kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "DESCONTOS  (lançar com sinal negativo)")
    rows["VALES"] = r
    r = linha(ws, r, "ADIANTAMENTO / VALES", {}, "PREENCHER com os vales adiantados durante setembro.", kind="in")
    rows["VALES INC"] = r
    r = linha(ws, r, "ADIANT. VALES INCENT. E APLIC.",
              {n: (f"=-({get_column_letter(C0+i)}{rows['INC APLIC']}+{get_column_letter(C0+i)}{rows['INC VIT']}"
                   f"+{get_column_letter(C0+i)}{rows['INC OUTROS']})") for i, n in enumerate(EMP)},
              "Estorno dos incentivos já adiantados em dinheiro no mês. AJUSTAR se algum incentivo não foi adiantado.",
              kind="calc", fill=FILL_CONF)
    rows["CONVÊNIO"] = r
    r = linha(ws, r, "CONVÊNIO", {}, "PREENCHER conforme o convênio de setembro.", kind="in")
    rows["FALTAS"] = r
    r = linha(ws, r, "DESCONTO FALTAS / ATRASOS", {}, "PREENCHER conforme o ponto.", kind="in")
    rows["VT6"] = r
    r = linha(ws, r, "DESCONTO VALE TRANSPORTE 6%", cfg["VT6"],
              "Valores praticados em ago/26. 6% sobre R$ 1.621,00 seria R$ 97,26 — CONFERIR a base usada.",
              kind="in", fill=FILL_CONF)
    rows["INSS"] = r
    r = linha(ws, r, "DESCONTO INSS", {}, "A CALCULAR PELA CONTABILIDADE sobre o total de proventos.", kind="in")
    rows["IRRF"] = r
    r = linha(ws, r, "DESCONTO IRRF", {}, "A CALCULAR PELA CONTABILIDADE.", kind="in")
    rows["TOTAL DESC"] = r
    r = linha(ws, r, "TOTAL DE DESCONTOS",
              {n: f"=SUM({get_column_letter(C0+i)}{rows['VALES']}:{get_column_letter(C0+i)}{rows['IRRF']})"
               for i, n in enumerate(EMP)}, "", kind="calc", bold=True, fill=FILL_TOT)
    r = sec(ws, r, "LÍQUIDO")
    rows["LIQ"] = r
    r = linha(ws, r, "LÍQUIDO A RECEBER (05/10/2026)",
              {n: f"={get_column_letter(C0+i)}{rows['TOTAL PROVENTOS']}+{get_column_letter(C0+i)}{rows['TOTAL DESC']}"
               for i, n in enumerate(EMP)},
              "Total de proventos menos descontos. Só fica definitivo depois do INSS/IRRF da contabilidade.",
              kind="calc", bold=True, fill=FILL_LIQ)
    r = sec(ws, r, "INFORMATIVO — PAGO PELA EMPRESA, NÃO ENTRA NO HOLERITE")
    rows["VT"] = r
    r = linha(ws, r, "VALE TRANSPORTE (compra out/26)", {}, "PREENCHER com as passagens compradas para outubro/2026.", kind="in")
    rows["VA"] = r
    r = linha(ws, r, "VALE ALIMENTAÇÃO FERIADOS E DOMINGOS", {}, "PREENCHER conforme escala de domingos e feriados.", kind="in")
    ws.auto_filter.ref = f"A4:{LO}{r-1}"
    r += 1
    for a, b in [("LEGENDA", ""),
                 ("Célula amarela, texto azul", "valor digitado — PREENCHER ou conferir antes de enviar."),
                 ("Célula laranja", "valor repetido de agosto/2026 — CONFERIR se continua válido em setembro."),
                 ("Texto verde", "valor puxado da aba BASE INOVAFARMA SET.26 (não digitar por cima)."),
                 ("Texto preto em célula verde", "resultado de fórmula — não alterar.")]:
        ws.cell(r, 1, a).font = font(9, True, "1F3864")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=CO)
        ws.cell(r, 2, b).font = font(9, it=True)
        r += 1
    HOL = "'HOLERITE SET.26'"

    # ---------------------------------------------------------- LISTA
    ws = wb.create_sheet("LISTA CONTABILIDADE")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [16, 40, 14, 16, 46]):
        ws.column_dimensions[col].width = w
    title_block(ws, f"LANÇAMENTOS POR FUNCIONÁRIO — {cfg['loja']} — SETEMBRO/2026 (pagamento 05/10/2026)",
                "Mesmos valores da aba HOLERITE SET.26, em formato de lista (vinculados por fórmula)", 5)
    for i, h in enumerate(["FUNCIONÁRIO", "RUBRICA", "TIPO", "VALOR", "OBSERVAÇÃO"], 1):
        c = ws.cell(4, i, h); c.font = font(10, True, "FFFFFF"); c.fill = FILL_HDR; c.border = BOX
    LISTA = [("SALÁRIO BASE", "SALÁRIO BASE", "Provento"),
             ("ADICIONAL NOTURNO", "ADICIONAL NOTURNO", "Provento"),
             ("HORAS EXTRAS", "HORAS EXTRAS", "Provento"),
             ("COMISSÃO TOTAL", "COMISSÃO SOBRE VENDAS", "Provento"),
             ("DSR", "REPOUSO REMUNERADO / DSR", "Provento"),
             ("AUXÍLIO GERÊNCIA", "AUXÍLIO GERÊNCIA", "Provento"),
             ("META CAIXA", "ADICIONAL PRÊMIO META CAIXA", "Provento"),
             ("PRÊMIO COTA GERAL", "PRÊMIO COTA GERAL", "Provento"),
             ("PRÊMIO PRÉ-VENCIDOS", "PRÊMIO PRÉ-VENCIDOS", "Provento"),
             ("FÉRIAS", "FÉRIAS", "Provento"),
             ("1/3 FÉRIAS", "1/3 CONSTITUCIONAL DE FÉRIAS", "Provento"),
             ("INC APLIC", "INCENTIVO APLICAÇÕES", "Provento"),
             ("INC VIT", "INCENTIVO VITAMINAS", "Provento"),
             ("INC OUTROS", "OUTROS INCENTIVOS", "Provento"),
             ("TOTAL PROVENTOS", "TOTAL DE PROVENTOS", "Subtotal"),
             ("VALES", "ADIANTAMENTO / VALES", "Desconto"),
             ("VALES INC", "ADIANT. VALES INCENT. E APLIC.", "Desconto"),
             ("CONVÊNIO", "CONVÊNIO", "Desconto"),
             ("FALTAS", "DESCONTO FALTAS / ATRASOS", "Desconto"),
             ("VT6", "DESCONTO VALE TRANSPORTE 6%", "Desconto"),
             ("INSS", "DESCONTO INSS", "Desconto"),
             ("IRRF", "DESCONTO IRRF", "Desconto"),
             ("TOTAL DESC", "TOTAL DE DESCONTOS", "Subtotal"),
             ("LIQ", "LÍQUIDO A RECEBER", "Líquido")]
    r = 5
    for i, n in enumerate(EMP):
        col = get_column_letter(C0 + i)
        for key, rot, tipo in LISTA:
            ws.cell(r, 1, n).font = font(10, tipo == "Líquido")
            ws.cell(r, 2, rot).font = font(10)
            ws.cell(r, 3, tipo).font = font(10)
            c = ws.cell(r, 4, f"={HOL}!{col}{rows[key]}")
            c.number_format = MONEY; c.font = font(10, tipo in ("Subtotal", "Líquido"), GREEN)
            if tipo == "Líquido":
                for k in range(1, 6): ws.cell(r, k).fill = FILL_LIQ
            elif tipo == "Subtotal":
                for k in range(1, 6): ws.cell(r, k).fill = FILL_TOT
            for k in range(1, 6):
                ws.cell(r, k).border = BOX
            r += 1
        r += 1
    ws.auto_filter.ref = f"A4:E{r-2}"
    ws.freeze_panes = "A5"

    # ---------------------------------------------------------- GANHOS
    wsg = wb.create_sheet("GANHOS DO COLABORADOR")
    wsg.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [4, 44, 18, 58]):
        wsg.column_dimensions[col].width = w
    title_block(wsg, "DEMONSTRATIVO DE GANHOS DO COLABORADOR",
                f"Loja {cfg['loja']} · competência SETEMBRO/2026 · pagamento em 05/09/2026", 4)
    wsg.cell(4, 2, "Escolha o colaborador:").font = font(11, True, "1F3864")
    sel = wsg.cell(4, 3, EMP[0])
    sel.font = font(12, True, BLUE); sel.fill = FILL_IN; sel.border = BOX
    sel.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1='"' + ",".join(EMP) + '"', allow_blank=False)
    dv.error = "Escolha um nome da lista."
    dv.promptTitle = "Colaborador"; dv.prompt = "Selecione o colaborador para ver os ganhos dele."
    wsg.add_data_validation(dv); dv.add(sel)
    wsg.cell(4, 4, "Troque o nome aqui e a tabela abaixo muda sozinha — é o relatório para mandar para cada um.").font = font(9, it=True)
    COLREF = f"MATCH($C$4,{HOL}!$B$4:${get_column_letter(CT-1)}$4,0)"
    RECIBO = [("SEC", "PROVENTOS", None),
              ("L", "Salário base", rows["SALÁRIO BASE"]),
              ("L", "Adicional noturno", rows["ADICIONAL NOTURNO"]),
              ("L", "Horas extras", rows["HORAS EXTRAS"]),
              ("L", "Comissão sobre vendas", rows["COMISSÃO TOTAL"]),
              ("L", "Repouso remunerado / DSR", rows["DSR"]),
              ("L", "Auxílio gerência", rows["AUXÍLIO GERÊNCIA"]),
              ("L", "Adicional prêmio meta caixa", rows["META CAIXA"]),
              ("L", "Prêmio cota geral", rows["PRÊMIO COTA GERAL"]),
              ("L", "Prêmio pré-vencidos", rows["PRÊMIO PRÉ-VENCIDOS"]),
              ("L", "Férias", rows["FÉRIAS"]),
              ("L", "1/3 constitucional de férias", rows["1/3 FÉRIAS"]),
              ("L", "Incentivo aplicações", rows["INC APLIC"]),
              ("L", "Incentivo vitaminas", rows["INC VIT"]),
              ("L", "Outros incentivos", rows["INC OUTROS"]),
              ("T", "TOTAL DE PROVENTOS", rows["TOTAL PROVENTOS"]),
              ("SEC", "DESCONTOS", None),
              ("L", "Adiantamento / vales", rows["VALES"]),
              ("L", "Adiantamento de incentivos e aplicações", rows["VALES INC"]),
              ("L", "Convênio", rows["CONVÊNIO"]),
              ("L", "Faltas / atrasos", rows["FALTAS"]),
              ("L", "Vale transporte 6%", rows["VT6"]),
              ("L", "INSS", rows["INSS"]),
              ("L", "IRRF", rows["IRRF"]),
              ("T", "TOTAL DE DESCONTOS", rows["TOTAL DESC"]),
              ("SEC", "LÍQUIDO", None),
              ("Q", "LÍQUIDO A RECEBER EM 05/10/2026", rows["LIQ"])]
    r = 6
    for tipo, rot, lh in RECIBO:
        if tipo == "SEC":
            for i in range(2, 5):
                wsg.cell(r, i).fill = FILL_SEC
            wsg.cell(r, 2, rot).font = font(11, True, "1F3864")
            wsg.row_dimensions[r].height = 19
            r += 1
            continue
        c = wsg.cell(r, 2, rot)
        c.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q")); c.border = BOX
        v = wsg.cell(r, 3, f"=IFERROR(INDEX({HOL}!$B${lh}:${get_column_letter(CT-1)}${lh},1,{COLREF}),0)")
        v.number_format = MONEY; v.font = font(11 if tipo == "Q" else 10, tipo in ("T", "Q"), GREEN); v.border = BOX
        if tipo == "T":
            wsg.cell(r, 2).fill = FILL_TOT; v.fill = FILL_TOT
        if tipo == "Q":
            wsg.cell(r, 2).fill = FILL_LIQ; v.fill = FILL_LIQ
            wsg.row_dimensions[r].height = 22
        r += 1
    r += 1
    wsg.cell(r, 2, "Vendas do colaborador no mês (InovaFarma):").font = font(10, True, "1F3864")
    r += 1
    for rot, colb in [("Venda bruta", "C"), ("Descontos concedidos", "D"), ("Venda líquida", "E"),
                      ("Comissão apurada", "F"), ("Total de incentivos", "J")]:
        wsg.cell(r, 2, rot).font = font(10); wsg.cell(r, 2).border = BOX
        v = wsg.cell(r, 3, f"=IFERROR(INDEX({BASE}!${colb}${first}:${colb}${last},MATCH($C$4,{BASE}!$B${first}:$B${last},0)),0)")
        v.number_format = MONEY; v.font = font(10, False, GREEN); v.border = BOX
        r += 1
    r += 1
    rodape = ['A comissão apurada acima é a do sistema. O que entra na folha é a linha "Comissão sobre vendas".']
    if cfg["mult"] != 1:
        rodape.append("Nesta loja a comissão paga é o DOBRO da apurada — por isso os dois valores são diferentes.")
    rodape += ["Os valores de INSS e IRRF são calculados pela contabilidade e só aparecem depois de preenchidos na aba HOLERITE SET.26.",
               "Para mandar para o colaborador: selecione o nome acima e imprima esta aba em PDF (ou tire um print)."]
    for t in rodape:
        wsg.cell(r, 2, t).font = font(9, it=True)
        r += 1
    wsg.print_area = f"A1:D{r}"
    wsg.page_setup.fitToWidth = 1
    wsg.page_setup.fitToHeight = 1
    wsg.sheet_properties.pageSetUpPr.fitToPage = True

    # ---------------------------------------------------------- PONTO
    TIPOS = {"HE50": ("Horas extras 50%", "horas", "HORAS EXTRAS"),
             "NOT": ("Adicional noturno", "horas", "ADICIONAL NOTURNO"),
             "FER": ("Feriado trabalhado (dobra)", "dias", "HORAS EXTRAS"),
             "FER_COMP": ("Feriado com folga compensatória", "dias", "—"),
             "ATESTADO": ("Atestado médico (abonado)", "dias", "—"),
             "INFO": ("Registro informativo", "dias", "—"),
             "VALOR": ("Adicional noturno (valor fixo)", "valor", "ADICIONAL NOTURNO")}
    wsp = wb.create_sheet("PONTO SET.26")
    wsp.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [14, 32, 10, 10, 14, 14, 14, 22, 62]):
        wsp.column_dimensions[col].width = w
    title_block(wsp, "APONTAMENTO DO PONTO — SETEMBRO/2026",
                f"Loja {cfg['loja']} · horas extras, adicional noturno e feriado trabalhado · alimenta a aba HOLERITE SET.26", 9)
    for i, h in enumerate(["FUNCIONÁRIO", "OCORRÊNCIA", "QTDE", "UNIDADE", "SALÁRIO BASE",
                           "VALOR UNITÁRIO", "TOTAL", "RUBRICA DESTINO", "OBSERVAÇÃO"], 1):
        c = wsp.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
    wsp.row_dimensions[4].height = 26
    SALROW = rows["SALÁRIO BASE"]
    LEMP = get_column_letter(CT - 1)
    P_HORAS = f"PARÂMETROS!$B${PROW['Horas mensais (base do salário-hora)']}"
    P_HE = f"PARÂMETROS!$B${PROW['Adicional de hora extra']}"
    P_NOT = f"PARÂMETROS!$B${PROW['Adicional noturno']}"
    P_DIAS = f"PARÂMETROS!$B${PROW['Dias do mês para o salário-dia']}"
    pr = 5; P_INI = 5
    def linha_sal(pr):
        c = wsp.cell(pr, 5, f"=IFERROR(INDEX({HOL}!$B${SALROW}:${LEMP}${SALROW},1,"
                            f"MATCH(A{pr},{HOL}!$B$4:${LEMP}$4,0)),0)")
        c.number_format = MONEY; c.font = font(10, False, GREEN)
    for func, tipo, qtd, obs in cfg.get("PONTO", []):
        rot, uni, dest = TIPOS[tipo]
        wsp.cell(pr, 1, func).font = font(10, True)
        wsp.cell(pr, 2, rot).font = font(10)
        c = wsp.cell(pr, 3, qtd if tipo != "VALOR" else 1)
        c.font = font(10, False, BLUE); c.fill = FILL_IN
        c.alignment = Alignment(horizontal="center")
        wsp.cell(pr, 4, uni).alignment = Alignment(horizontal="center")
        linha_sal(pr)
        if tipo == "HE50":
            fu = f"=ROUND(E{pr}/{P_HORAS}*(1+{P_HE}),4)"
        elif tipo == "NOT":
            fu = f"=ROUND(E{pr}/{P_HORAS}*{P_NOT},4)"
        elif tipo == "FER":
            fu = f"=ROUND(E{pr}/{P_DIAS},2)"
        elif tipo in ("FER_COMP", "ATESTADO", "INFO"):
            fu = 0
        else:
            fu = qtd
        c = wsp.cell(pr, 6, fu); c.number_format = MONEY
        c.font = font(10, False, BLUE if tipo == "VALOR" else "000000")
        if tipo == "VALOR":
            c.fill = FILL_CONF
        c = wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)"); c.number_format = MONEY
        c.font = font(10, True); c.fill = FILL_TOT
        wsp.cell(pr, 8, dest).font = font(9, True, "1F3864")
        c = wsp.cell(pr, 9, obs); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        for i in range(1, 10):
            wsp.cell(pr, i).border = BOX
        pr += 1
    for extra in range(8):
        for i in range(1, 10):
            cell = wsp.cell(pr, i); cell.border = BOX
            if i in (1, 2, 3, 4, 8, 9):
                cell.font = font(10, False, BLUE); cell.fill = FILL_IN
        linha_sal(pr)
        wsp.cell(pr, 6).number_format = MONEY
        wsp.cell(pr, 6).font = font(10, False, BLUE); wsp.cell(pr, 6).fill = FILL_IN
        wsp.cell(pr, 7, f"=ROUND(C{pr}*F{pr},2)").number_format = MONEY
        wsp.cell(pr, 7).font = font(10, True); wsp.cell(pr, 7).fill = FILL_TOT
        pr += 1
    P_FIM = pr - 1
    wsp.cell(pr, 2, "TOTAL").font = font(10, True)
    c = wsp.cell(pr, 7, f"=SUM(G{P_INI}:G{P_FIM})"); c.number_format = MONEY; c.font = font(10, True)
    for i in range(1, 10):
        wsp.cell(pr, i).fill = FILL_TOT; wsp.cell(pr, i).border = BOX
    wsp.auto_filter.ref = f"A4:I{P_FIM}"
    wsp.freeze_panes = "A5"
    pr += 2
    for t in ["Salário-hora = salário base ÷ 220. Hora extra = salário-hora × 1,5. Adicional noturno = salário-hora × 20%. Feriado trabalhado sem folga = 1 salário-dia a mais (salário base ÷ 30).",
              "No salário de R$ 1.621,00: 40 horas extras dão R$ 442,09 e 24 horas dão R$ 265,25 — é a mesma conta usada na planilha de julho.",
              "As linhas em branco são para acrescentar ocorrências: preencha funcionário, ocorrência, quantidade, valor unitário e a rubrica de destino (HORAS EXTRAS ou ADICIONAL NOTURNO).",
              "Os totais desta aba entram sozinhos nas linhas HORAS EXTRAS e ADICIONAL NOTURNO da aba HOLERITE SET.26."]:
        wsp.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=9)
        c = wsp.cell(pr, 1, t); c.font = font(9, it=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        wsp.row_dimensions[pr].height = 24
        pr += 1
    if cfg.get("PONTO_PROX"):
        pr += 1
        for i in range(1, 10):
            wsp.cell(pr, i).fill = FILL_SEC
        c = wsp.cell(pr, 1, "OCORRÊNCIAS PARA A PRÓXIMA COMPETÊNCIA (SETEMBRO/2026) — NÃO ENTRAM NESTA FOLHA")
        c.font = font(11, True, "1F3864")
        wsp.row_dimensions[pr].height = 19
        pr += 1
        for func, ocorrencia, detalhe in cfg["PONTO_PROX"]:
            c = wsp.cell(pr, 1, func); c.font = font(10, True); c.border = BOX
            wsp.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=9)
            c = wsp.cell(pr, 2, ocorrencia); c.font = font(10, True, "C00000"); c.fill = FILL_ALERT
            c.alignment = Alignment(wrap_text=True, vertical="center")
            pr += 1
            wsp.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=9)
            c = wsp.cell(pr, 2, detalhe); c.font = font(9, it=True)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            wsp.row_dimensions[pr].height = 30
            pr += 1

    wsh = wb["HOLERITE SET.26"]
    for rub, lh in (("HORAS EXTRAS", rows["HORAS EXTRAS"]), ("ADICIONAL NOTURNO", rows["ADICIONAL NOTURNO"])):
        for i, n in enumerate(EMP):
            c = wsh.cell(lh, C0 + i,
                         f"=SUMIFS('PONTO SET.26'!$G${P_INI}:$G${P_FIM},'PONTO SET.26'!$A${P_INI}:$A${P_FIM},"
                         f"{get_column_letter(C0+i)}$4,'PONTO SET.26'!$H${P_INI}:$H${P_FIM},\"{rub}\")")
            c.number_format = MONEY; c.font = font(10, False, GREEN); c.fill = PatternFill()
        wsh.cell(lh, CO).value = "Somado automaticamente da aba PONTO SET.26 (apontamento de setembro/2026). Para mudar, edite lá."

    # ---------------------------------------------------------- FOLGUISTAS
    if cfg.get("FOLGUISTAS"):
        wf = wb.create_sheet("FOLGUISTAS")
        wf.sheet_view.showGridLines = False
        for col, w in zip("ABCDEFGH", [22, 8, 14, 12, 16, 16, 15, 15]):
            wf.column_dimensions[col].width = w
        title_block(wf, "FOLGUISTAS — CONTAS A PAGAR",
                    f"Loja {cfg['loja']} · setembro/2026 · NÃO entram no holerite: pagamento por diária, lançar no contas a pagar", 8)
        hdr = ["COLABORADOR", "CÓD.", "VALOR DIÁRIA", "Nº DIÁRIAS", "TOTAL DIÁRIAS",
               "COMISSÃO APURADA", "INCENTIVOS", "TOTAL A PAGAR"]
        for i, h in enumerate(hdr, 1):
            c = wf.cell(4, i, h); c.font = font(9, True, "FFFFFF"); c.fill = FILL_HDR
            c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
        wf.row_dimensions[4].height = 28
        rr = 5; ini = rr
        for nome, cod, diaria, bruta, com, inc in cfg["FOLGUISTAS"]:
            wf.cell(rr, 1, nome).font = font(10, True)
            wf.cell(rr, 2, cod).alignment = Alignment(horizontal="center")
            c = wf.cell(rr, 3, diaria); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 4); c.font = font(10, False, BLUE); c.fill = FILL_IN
            c.alignment = Alignment(horizontal="center")
            wf.cell(rr, 5, f"=C{rr}*D{rr}").number_format = MONEY
            c = wf.cell(rr, 6, com); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 7, inc); c.number_format = MONEY; c.font = font(10, False, BLUE); c.fill = FILL_IN
            c = wf.cell(rr, 8, f"=E{rr}+F{rr}+G{rr}"); c.number_format = MONEY; c.font = font(10, True)
            c.fill = FILL_LIQ
            for i in range(1, 9):
                wf.cell(rr, i).border = BOX
            rr += 1
        fim = rr - 1
        wf.cell(rr, 1, "TOTAL").font = font(10, True)
        for i in (5, 6, 7, 8):
            L = get_column_letter(i)
            c = wf.cell(rr, i, f"=SUM({L}{ini}:{L}{fim})")
            c.number_format = MONEY; c.font = font(10, True)
        for i in range(1, 9):
            wf.cell(rr, i).fill = FILL_TOT; wf.cell(rr, i).border = BOX
        wf.auto_filter.ref = f"A4:H{fim}"
        wf.freeze_panes = "A5"
        rr += 2
        for t in ["PREENCHER a coluna Nº DIÁRIAS com os dias trabalhados em setembro/2026.",
                  "Valor da diária combinado: SERGIO R$ 150,00 e ANA CELIA R$ 100,00.",
                  "Comissão e incentivos vêm do relatório do InovaFarma (códigos 51 e 131, quando o relatório de setembro for extraído) — CONFIRMAR se o folguista recebe esses valores além da diária; se não receber, zerar as colunas.",
                  "Estes valores não entram no holerite: são pagamento de prestação de serviço, para baixa no contas a pagar."]:
            wf.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
            c = wf.cell(rr, 1, t); c.font = font(9, it=True)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            rr += 1

    # ---------------------------------------------------------- CAPA
    ws = wb.create_sheet("CAPA", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 118
    title_block(ws, f"RELATÓRIO DE FOLHA PARA A CONTABILIDADE — LOJA {cfg['loja']}", "", 2)
    blocos = [("SEC", "IDENTIFICAÇÃO"),
              ("T", f"Empresa: Farmácia Tropical Multi Econômica — loja {cfg['loja']}"),
              ("T", "Competência (mês de referência): SETEMBRO/2026 — vendas de 01/09/2026 a 30/09/2026"),
              ("T", "Data do pagamento: 05/10/2026"),
              ("T", "Fonte das comissões e incentivos: InovaFarma — relatório de comissão de vendedor, a extrair no início de outubro/2026"),
              ("T", "Funcionários: " + ", ".join(EMP)),
              ("SEC", "COMO USAR ESTE ARQUIVO"),
              ("T", "1. Abra a aba HOLERITE SET.26 — é o relatório que vai para a contabilidade."),
              ("T", "2. Preencha as células AMARELAS (horas extras, prêmios, vales, convênio, faltas)."),
              ("T", "3. Confira as células LARANJAS — são valores repetidos de julho/2026 que podem ter mudado."),
              ("T", "4. Comissões, incentivos e DSR já vêm calculados; não digite por cima."),
              ("T", "5. A aba GANHOS DO COLABORADOR monta o demonstrativo individual: escolha o nome na lista."),
              ("T", "6. INSS e IRRF ficam em branco: são calculados pela contabilidade."),
              ("SEC", "APURAÇÃO DO MÊS"),
              ("T", f"Venda bruta geral da loja: R$ {g['bruta']:,.2f} · descontos concedidos: R$ {g['desc']:,.2f} · venda líquida: R$ {g['liq']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
              ("T", f"Comissão apurada no InovaFarma: R$ {g['com']:,.2f} · incentivos: R$ {g['inc']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
              ("T", "DSR de setembro/2026: 4 domingos + o feriado de 07/09 ÷ 25 dias úteis = fator 0,200000"),
              ("SEC", "DEFINIÇÕES DA LOJA")]
    for t in cfg["DEFINIDO"]:
        blocos.append(("T", t))
    blocos.append(("SEC", "PENDÊNCIAS — CONFIRMAR ANTES DE ENVIAR"))
    for t in ["Horas extras, adicional noturno e feriados trabalhados de setembro — lançar na aba PONTO SET.26.",
              "Prêmio cota geral e prêmio pré-vencidos, conforme a tabela de metas da loja.",
              "Vales adiantados, convênio e faltas de agosto.",
              "Férias, afastamentos e admissões que mudem o salário do mês."] + cfg["PENDENCIAS"]:
        blocos.append(("P", t))
    blocos += [("SEC", "OBSERVAÇÃO TÉCNICA"),
               ("T", "As células de total e de cálculo são fórmulas. Ao abrir no Excel ou no Google Planilhas elas aparecem calculadas; em visualizadores simples podem aparecer em branco até o arquivo ser aberto."),
               ("SEC", "ABAS DO ARQUIVO"),
               ("T", "HOLERITE SET.26 — relatório principal da competência setembro/2026."),
               ("T", "GANHOS DO COLABORADOR — demonstrativo individual, pronto para imprimir/mandar."),
               ("T", "LISTA CONTABILIDADE — os mesmos lançamentos em formato de lista."),
               ("T", "BASE INOVAFARMA SET.26 — apuração de comissões e incentivos por vendedor."),
               ("T", "PARÂMETROS — calendário do mês, fator do DSR e valores fixos."),
               ("T", "PONTO SET.26 — horas extras, adicional noturno e feriado trabalhado; alimenta o holerite.")]
    if cfg.get("FOLGUISTAS"):
        blocos.append(("T", "FOLGUISTAS — diárias de SERGIO e ANA CELIA para o contas a pagar (fora do holerite)."))
    r = 4
    for tipo, txt in blocos:
        if tipo == "SEC":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            c = ws.cell(r, 1, txt); c.font = font(11, True, "1F3864"); c.fill = FILL_SEC
            ws.row_dimensions[r].height = 20
        else:
            ws.cell(r, 1, "•" if tipo == "P" else "").font = font(10, True, "C00000")
            c = ws.cell(r, 2, txt); c.font = font(10)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if tipo == "P":
                c.fill = FILL_ALERT
            ws.row_dimensions[r].height = 17
        r += 1

    wb.save(DIR + cfg["arquivo"])
    return DIR + cfg["arquivo"]

if __name__ == "__main__":
    for cfg in (ARRAIAL, CENTRO, TRANCOSO):
        print("ok", build(cfg))
