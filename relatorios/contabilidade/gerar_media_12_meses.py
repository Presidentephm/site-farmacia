# -*- coding: utf-8 -*-
"""Media de comissoes dos ultimos 12 meses - base de calculo de ferias.
Fonte: InovaFarma, relatorio RESUMIDO POR GRUPO, extraido em 03/09/2026."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/home/user/site-farmacia/relatorios/contabilidade/MEDIA_COMISSOES_12_MESES_FERIAS.xlsx"
F = "Arial"
MONEY = 'R$ #,##0.00;-R$ #,##0.00;"-"'
def font(sz=10, b=False, color="000000", it=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=it)
FILL_TIT = PatternFill("solid", fgColor="1F3864")
FILL_SEC = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")
FILL_LIQ = PatternFill("solid", fgColor="C6E0B4")
FILL_ALERT = PatternFill("solid", fgColor="FFF0F0")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BLUE = "0000FF"

CASOS = [
 dict(aba="VALDICK", nome="VALDIC GONÇALVES RODRIGUES", cod="162", loja="TRANCOSO",
      periodo="01/09/2025 a 31/08/2026", ferias="férias de 01 a 30/09/2026",
      itens=17252, bruta=562715.75, desc=224944.06, liq=337771.69,
      com=10200.3462, inc=2702.00, salario=1621.00, dobro=True),
 dict(aba="RENALDO", nome="RENALDO RODRIGUES", cod="151", loja="CENTRO (MATRIZ)",
      periodo="01/09/2025 a 31/08/2026", ferias="férias de 01 a 30/09/2026",
      itens=15854, bruta=481964.11, desc=191683.4454, liq=290280.6646,
      com=9714.1688, inc=1676.00, salario=1621.00, dobro=False),
 dict(aba="MANOEL", nome="MANOEL SILVA", cod="92", loja="TRANCOSO",
      periodo="01/05/2025 a 30/04/2026", ferias="férias gozadas em maio/2026",
      itens=11545, bruta=363522.79, desc=133067.63, liq=230455.16,
      com=7253.0803, inc=421.00, salario=1621.00, dobro=True),
]

wb = openpyxl.Workbook()
primeiro = True
for c_ in CASOS:
    ws = wb.active if primeiro else wb.create_sheet()
    ws.title = c_["aba"]; primeiro = False
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABC", [46, 20, 66]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:C1")
    c = ws.cell(1, 1, f"MÉDIA DOS ÚLTIMOS 12 MESES — {c_['nome']}")
    c.font = font(13, True, "FFFFFF"); c.fill = FILL_TIT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:C2")
    c = ws.cell(2, 1, f"Loja {c_['loja']} · código {c_['cod']} · período {c_['periodo']} · base para {c_['ferias']}")
    c.font = font(10, True, "1F3864"); c.alignment = Alignment(horizontal="center")

    pos = [4]
    def sec(txt):
        r = pos[0]
        for i in range(1, 4):
            ws.cell(r, i).fill = FILL_SEC
        ws.cell(r, 1, txt).font = font(11, True, "1F3864")
        ws.row_dimensions[r].height = 19
        pos[0] += 1
    def linha(rot, val, obs="", fill=None, bold=False, fmt=MONEY, cor=None):
        r = pos[0]
        c = ws.cell(r, 1, rot); c.font = font(10, bold); c.border = BOX
        c2 = ws.cell(r, 2, val); c2.number_format = fmt
        c2.font = font(11 if bold else 10, bold, cor or "000000"); c2.border = BOX
        if fill:
            c.fill = fill; c2.fill = fill
        c3 = ws.cell(r, 3, obs); c3.font = font(9, it=True)
        c3.alignment = Alignment(wrap_text=True, vertical="center"); c3.border = BOX
        pos[0] += 1
        return r

    sec("APURADO NO INOVAFARMA — 12 MESES")
    linha("Total de itens vendidos", c_["itens"], "", fmt='#,##0')
    linha("Venda bruta", c_["bruta"])
    linha("Descontos concedidos", -c_["desc"])
    linha("Venda líquida", c_["liq"], "", fill=FILL_TOT)
    l_com = linha("Comissão apurada (12 meses)", c_["com"], "Total do relatório RESUMIDO POR GRUPO.", fill=FILL_TOT, bold=True)
    l_inc = linha("Incentivos (12 meses)", c_["inc"], "Injetáveis, populares, oficinais e demais incentivos.", fill=FILL_TOT, bold=True)

    sec("MÉDIA MENSAL (÷ 12)")
    m_com = linha("Média da comissão apurada", f"=ROUND(B{l_com}/12,2)", "", bold=True)
    m_inc = linha("Média dos incentivos", f"=ROUND(B{l_inc}/12,2)", "", bold=True)
    if c_["dobro"]:
        m_dob = linha("Média da comissão em dobro", f"=ROUND(B{m_com}*2,2)",
                      "Só vale se a loja pagou em dobro no período — ver o aviso no rodapé.", fill=FILL_ALERT)
    m_tot = linha("MÉDIA DA PARTE VARIÁVEL", f"=ROUND(B{m_com}+B{m_inc},2)",
                  "Comissão apurada + incentivos. É esta média que entra na base das férias.",
                  fill=FILL_LIQ, bold=True)

    sec("BASE DE FÉRIAS — SIMULAÇÃO PARA O CONTAS A PAGAR (30 dias)")
    l_sal = linha("Salário base", c_["salario"], "Piso 2026. Conferir se era outro no mês das férias.", cor=BLUE, fill=FILL_IN)
    l_base = linha("Base de férias (salário + média variável)", f"=ROUND(B{l_sal}+B{m_tot},2)", "", fill=FILL_TOT, bold=True)
    l_ter = linha("1/3 constitucional", f"=ROUND(B{l_base}/3,2)", "", fill=FILL_TOT)
    linha("TOTAL BRUTO DAS FÉRIAS", f"=ROUND(B{l_base}+B{l_ter},2)",
          "Valor bruto. INSS e IRRF sobre férias são calculados pela contabilidade.",
          fill=FILL_LIQ, bold=True)

    r = pos[0] + 1
    notas = ["OBSERVAÇÕES:",
             f"Fonte: InovaFarma, relatório PRODUTOS VENDIDOS RESUMIDO POR GRUPO, vendedor {c_['nome']}, período {c_['periodo']}, extraído em 03/09/2026.",
             "A média de férias é a soma dos 12 meses dividida por 12 — o relatório do período inteiro já dá essa soma, não precisa dos meses separados.",
             "Só a parte variável entra nesta média. Horas extras, adicional noturno e prêmios, se houver, têm média própria e devem ser somados pela contabilidade.",
             "Esta aba é uma simulação para o contas a pagar. O cálculo oficial do recibo de férias é da contabilidade."]
    if c_["dobro"]:
        notas.insert(2, "ATENÇÃO — TRANCOSO: o pagamento em dobro da comissão começou em agosto/2026, e em julho a comissão foi lançada por média. Nos meses anteriores dessa janela o que foi pago pode ter sido diferente do apurado. A média correta é sobre o que foi EFETIVAMENTE PAGO: se o dobro valeu em poucos meses, use a média da comissão apurada; se valeu no período todo, use a média em dobro. Confira com as planilhas de cada mês antes de fechar o recibo.")
    for t in notas:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, t)
        alerta = t.startswith("ATENÇÃO")
        c.font = font(9, t.endswith(":") or alerta, "C00000" if alerta or t.endswith(":") else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if alerta:
            c.fill = FILL_ALERT
            ws.row_dimensions[r].height = 40
        r += 1

# aba comparativa
ws = wb.create_sheet("COMPARATIVO", 0)
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [26, 12, 24, 16, 16, 18, 18]):
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:G1")
c = ws.cell(1, 1, "MÉDIA DOS 12 MESES — RESUMO DOS TRÊS CASOS")
c.font = font(13, True, "FFFFFF"); c.fill = FILL_TIT
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:G2")
c = ws.cell(2, 1, "Base de cálculo das férias · fonte InovaFarma, extraído em 03/09/2026")
c.font = font(10, True, "1F3864"); c.alignment = Alignment(horizontal="center")
hdr = ["FUNCIONÁRIO", "CÓD.", "PERÍODO", "COMISSÃO 12M", "INCENTIVOS 12M",
       "MÉDIA MENSAL", "BASE DE FÉRIAS"]
for i, h in enumerate(hdr, 1):
    c = ws.cell(4, i, h); c.font = font(9, True, "FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E5496")
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = BOX
ws.row_dimensions[4].height = 28
r = 5
for c_ in CASOS:
    ws.cell(r, 1, c_["nome"]).font = font(10, True)
    ws.cell(r, 2, c_["cod"]).alignment = Alignment(horizontal="center")
    ws.cell(r, 3, c_["periodo"]).font = font(10)
    for i, v in ((4, c_["com"]), (5, c_["inc"])):
        cc = ws.cell(r, i, v); cc.number_format = MONEY
    cc = ws.cell(r, 6, f"=ROUND((D{r}+E{r})/12,2)")
    cc.number_format = MONEY; cc.font = font(10, True); cc.fill = FILL_TOT
    cc = ws.cell(r, 7, f"=ROUND(F{r}+{c_['salario']},2)")
    cc.number_format = MONEY; cc.font = font(10, True); cc.fill = FILL_LIQ
    for i in range(1, 8):
        ws.cell(r, i).border = BOX
    r += 1
r += 1
for t in ["BASE DE FÉRIAS = salário base + média mensal da parte variável (comissão + incentivos).",
          "Cada aba traz o detalhe do funcionário, com o 1/3 constitucional e o total bruto das férias.",
          "TRANCOSO (VALDICK e MANOEL): conferir o efeito do pagamento em dobro antes de fechar — ver o aviso na aba de cada um."]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(r, 1, t); c.font = font(9, it=True)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1
wb.save(OUT)
print("ok", OUT)
